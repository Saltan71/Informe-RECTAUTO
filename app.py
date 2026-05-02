import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime, timedelta
import io
import zipfile
from fpdf import FPDF
import matplotlib.pyplot as plt
import os
import hashlib
import tempfile
import shutil
import uuid
import getpass
from PIL import Image
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, ColumnsAutoSizeMode
import math
from concurrent.futures import ThreadPoolExecutor, as_completed

# === NUEVA CLASE PARA ENTORNO DE USUARIO ===
class UserEnvironment:
    def __init__(self):
        try:
            self.username = getpass.getuser()
        except:
            self.username = "unknown_user"
        self.session_id = f"{self.username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
        self.working_dir = os.path.join(tempfile.gettempdir(), f"rectauto_{self.session_id}")
        os.makedirs(self.working_dir, exist_ok=True)
        print(f"User environment created: {self.working_dir}")  # Para debugging
        
    def get_cache_key(self, base_key):
        return f"{self.session_id}_{base_key}"
    
    def get_temp_path(self, filename):
        return os.path.join(self.working_dir, filename)
    
    def cleanup(self):
        """Limpia los archivos temporales (opcional)"""
        try:
            shutil.rmtree(self.working_dir)
        except:
            pass

# Inicializar el entorno de usuario
user_env = UserEnvironment()

# Constantes
FECHA_REFERENCIA = datetime(2022, 11, 1)
HOJA = "Sheet1"
ESTADOS_PENDIENTES = ["Abierto"]
CACHE_TTL = 14800  # 4 horas en segundos
CACHE_TTL_STATIC = 86400  # 24 horas para datos estáticos
CACHE_TTL_DYNAMIC = 7200  # 2 hora para datos dinámicos
COL_WIDTHS_OPTIMIZED = [28, 11, 11, 8, 16, 11, 11, 16, 11, 20, 20, 10, 18, 11, 14, 10, 24, 20, 11]

# Test file en directorio único por usuario
test_file = user_env.get_temp_path("test_write_access.tmp")
with open(test_file, 'w') as f:
    f.write("test")

st.set_page_config(page_title="Informe Rectauto", layout="wide", page_icon=Image.open("icono.ico"))

# CSS ACTUALIZADO CON ESTILOS PARA BOTONES Y MENÚS DESPLEGABLES
st.markdown("""
<style>
    [data-testid="stSidebar"] {
        background-color: #007933 !important;
    }
    
    .main .block-container {
        background-color: #C4DDCA !important;
        padding: 2rem;
        border-radius: 10px;
    }
    
    .stApp {
        background-color: #92C88F !important;
    }
    
    [data-testid="stSidebar"] * {
        color: white !important;
    }
    
    .main .stMarkdown, .main h1, .main h2, .main h3 {
        color: #333333 !important;
    }
    
    /* ESTILOS ESPECÍFICOS PARA BOTONES EN LA BARRA LATERAL */
    [data-testid="stSidebar"] button {
        background-color: #005a25 !important;
        color: white !important;
        border: 1px solid white !important;
        border-radius: 5px !important;
        padding: 0.5rem 1rem !important;
        font-weight: bold !important;
    }
    
    [data-testid="stSidebar"] button:hover {
        background-color: #003d1a !important;
        color: white !important;
        border: 1px solid white !important;
    }
    
    [data-testid="stSidebar"] button:focus {
        background-color: #003d1a !important;
        color: white !important;
        border: 2px solid #ffffff !important;
        box-shadow: 0 0 0 0.2rem rgba(255, 255, 255, 0.25) !important;
    }
    
    [data-testid="stSidebar"] button[kind="primary"] {
        background-color: #00802b !important;
        color: white !important;
        border: 2px solid #ffffff !important;
    }
    
    [data-testid="stSidebar"] button[kind="primary"]:hover {
        background-color: #006622 !important;
        color: white !important;
    }
    
    /* Estilos para los botones de navegación de semanas en KPI */
    [data-testid="stSidebar"] .stButton button {
        background-color: #005a25 !important;
        color: white !important;
        border: 1px solid white !important;
        border-radius: 5px !important;
        padding: 0.5rem 1rem !important;
        font-weight: bold !important;
        width: 100% !important;
    }
    
    [data-testid="stSidebar"] .stButton button:hover {
        background-color: #003d1a !important;
        color: white !important;
        border: 1px solid white !important;
    }
    
    /* ESTILOS PARA MENÚS DESPLEGABLES EN BARRA LATERAL */
    [data-testid="stSidebar"] .stSelectbox > div > div {
        background-color: #009945 !important;
        color: white !important;
        border: 1px solid white !important;
        border-radius: 5px !important;
    }
    
    [data-testid="stSidebar"] .stSelectbox > div > div:hover {
        background-color: #007933 !important;
        color: white !important;
        border: 1px solid white !important;
    }
    
    [data-testid="stSidebar"] .stSelectbox input {
        color: white !important;
    }
    
    [data-testid="stSidebar"] .stMultiSelect > div > div {
        background-color: #009945 !important;
        color: white !important;
        border: 1px solid white !important;
        border-radius: 5px !important;
    }
    
    [data-testid="stSidebar"] .stMultiSelect > div > div:hover {
        background-color: #007933 !important;
        color: white !important;
        border: 1px solid white !important;
    }
    
    [data-testid="stSidebar"] .stMultiSelect input {
        color: white !important;
    }
    
    /* Estilos para las opciones del menú desplegable */
    [data-testid="stSidebar"] .stSelectbox [data-testid="stMarkdownContainer"] p,
    [data-testid="stSidebar"] .stMultiSelect [data-testid="stMarkdownContainer"] p {
        color: white !important;
    }
    
    /* Estilos para las opciones seleccionadas en multiselect */
    [data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"] {
        background-color: #005a25 !important;
        color: white !important;
        border: 1px solid white !important;
        border-radius: 12px !important;
    }
    
    [data-testid="stSidebar"] .stMultiSelect [data-baseweb="tag"]:hover {
        background-color: #003d1a !important;
        color: white !important;
    }
    
    /* Estilos para el menú desplegable abierto */
    [data-testid="stSidebar"] div[data-baseweb="popover"] {
        background-color: #009945 !important;
        border: 1px solid white !important;
        border-radius: 5px !important;
    }
    
    [data-testid="stSidebar"] div[data-baseweb="menu"] {
        background-color: #009945 !important;
        color: white !important;
    }
    
    [data-testid="stSidebar"] div[data-baseweb="menu"] li {
        background-color: #009945 !important;
        color: white !important;
    }
    
    [data-testid="stSidebar"] div[data-baseweb="menu"] li:hover {
        background-color: #007933 !important;
        color: white !important;
    }
    
    [data-testid="stSidebar"] div[data-baseweb="menu"] li:focus {
        background-color: #005a25 !important;
        color: white !important;
    }
    
    /* Estilos para el texto dentro de los menús desplegables */
    [data-testid="stSidebar"] div[data-baseweb="popover"] * {
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

# Logo
st.sidebar.image("Logo Atrian.png", width=260)

# Menú principal reorganizado
menu = ["Carga de Archivos", "Vista de Expedientes", "Indicadores clave (KPI)", "Análisis del Rendimiento", "Informes y Correos"]
eleccion = st.sidebar.selectbox("Menú", menu)

# Función para obtener información de la semana actual
def obtener_info_semana_actual(df_combinado):
    """Obtiene la información de la semana actual para mostrar en todas las páginas"""
    if df_combinado is None or df_combinado.empty:
        return None, None, None
    
    try:
        # Obtener la columna de fecha (asumiendo que está en la posición 13)
        columna_fecha = df_combinado.columns[13]
        df_combinado[columna_fecha] = pd.to_datetime(df_combinado[columna_fecha], errors='coerce')
        fecha_max = df_combinado[columna_fecha].max()
        
        if pd.isna(fecha_max):
            return None, None, None
        
        # Ajustar la fecha al viernes de esa semana
        # Si la fecha no es viernes, encontrar el viernes siguiente
        # Lunes=0, Domingo=6 → Viernes=4
        if fecha_max.weekday() != 4:  # 4 representa viernes
            # Calcular cuántos días faltan para el próximo viernes
            dias_hasta_viernes = (4 - fecha_max.weekday()) % 7
            # Si es sábado (5) o domingo (6), %7 nos da el correcto
            fecha_viernes = fecha_max + pd.Timedelta(days=dias_hasta_viernes)
        else:
            fecha_viernes = fecha_max
        
        # Calcular días transcurridos desde FECHA_REFERENCIA hasta el viernes
        dias_transcurridos = (fecha_viernes - FECHA_REFERENCIA).days
        
        if dias_transcurridos < 0:
            return None, None, None
            
        num_semana = dias_transcurridos // 7 + 1
        fecha_max_str = fecha_viernes.strftime("%d/%m/%Y")
        fecha_max = fecha_viernes
        
        return num_semana, fecha_max_str, fecha_max
    except Exception as e:
        print(f"Error en obtener_info_semana_actual: {e}")
        return None, None, None

# Mostrar información de la semana actual en todas las páginas
if "df_combinado" in st.session_state:
    num_semana, fecha_max_str, fecha_max = obtener_info_semana_actual(st.session_state["df_combinado"])
    if num_semana and fecha_max_str:
        st.title(f"📊 Seguimiento Equipo Regional RECTAUTO - Semana {num_semana} a {fecha_max_str}")
    else:
        st.title("📊 Seguimiento Equipo Regional RECTAUTO")
else:
    st.title("📊 Seguimiento Equipo Regional RECTAUTO")

# Inicializar variables de sesión para documentación
if 'mostrar_descarga' not in st.session_state:
    st.session_state.mostrar_descarga = False
if 'documentos_actualizados' not in st.session_state:
    st.session_state.documentos_actualizados = None
if 'cambios_documentacion' not in st.session_state:
    st.session_state.cambios_documentacion = {}

# Clase PDF (mejorada con ordenamiento)
class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 8)
        self.cell(0, 5, 'Informe de Expedientes Pendientes', 0, 1, 'C')
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 6)
        self.cell(0, 5, f'Página {self.page_no()}', 0, 0, 'C')
    
    def aplicar_formato_condicional_pdf(self, df_original, idx, col_name, col_width, altura_fila, x, y):
        """Aplica formato condicional a celdas específicas en el PDF con NUEVAS condiciones"""
        try:
            if idx >= len(df_original):
                return
                
            fila = df_original.iloc[idx]
            
            # Condición 1: USUARIO vs USUARIO-CSV
            if col_name == 'USUARIO-CSV':
                usuario_principal = fila.get('USUARIO', '')
                usuario_csv = fila.get('USUARIO-CSV', '')
                
                # Comparar los valores (manejar NaN y tipos diferentes)
                if pd.notna(usuario_principal) and pd.notna(usuario_csv):
                    if str(usuario_principal).strip() != str(usuario_csv).strip():
                        # Fondo rojo cuando USUARIO es distinto de USUARIO-CSV
                        self.set_fill_color(255, 0, 0)
                        self.rect(x, y, col_width, altura_fila, 'F')
                        self.set_fill_color(255, 255, 255)
            
            # Condición 2: RUE con MÚLTIPLES condiciones específicas
            elif col_name == 'RUE':
                etiq_penultimo = fila.get('ETIQ. PENÚLTIMO TRAM.', '')
                fecha_notif = fila.get('FECHA NOTIFICACIÓN', None)
                docum_incorp = fila.get('DOCUM.INCORP.', '')
                
                es_amarillo = False
                
                # CONDICIÓN 2.1: "80 PROPRES" con fecha límite superada
                if (str(etiq_penultimo).strip() == "80 PROPRES" and 
                    pd.notna(fecha_notif) and 
                    isinstance(fecha_notif, (pd.Timestamp, datetime))):
                    
                    fecha_limite = fecha_notif + timedelta(days=23)
                    if datetime.now() > fecha_limite:
                        es_amarillo = True
                
                # NUEVA CONDICIÓN 2.2: "50 REQUERIR" con fecha límite superada
                elif (str(etiq_penultimo).strip() == "50 REQUERIR" and 
                    pd.notna(fecha_notif) and 
                    isinstance(fecha_notif, (pd.Timestamp, datetime))):
                    
                    fecha_limite = fecha_notif + timedelta(days=23)
                    if datetime.now() > fecha_limite:
                        es_amarillo = True
                
                # NUEVA CONDICIÓN 2.3: "70 ALEGACI" o "60 CONTESTA"
                elif str(etiq_penultimo).strip() in ["70 ALEGACI", "60 CONTESTA"]:
                    es_amarillo = True
                
                # NUEVA CONDICIÓN 2.4: DOCUM.INCORP. no vacío Y distinto de "SOLICITUD" o "REITERA SOLICITUD"
                elif (pd.notna(docum_incorp) and 
                    str(docum_incorp).strip() != '' and
                    str(docum_incorp).strip().upper() not in ["SOLICITUD", "REITERA SOLICITUD"]):
                    es_amarillo = True
                
                if es_amarillo:
                    # Fondo amarillo cuando se cumple alguna condición
                    self.set_fill_color(255, 255, 0)
                    self.rect(x, y, col_width, altura_fila, 'F')
                    self.set_fill_color(255, 255, 255)
            
            # Condición 3: Resaltar DOCUM.INCORP. cuando tiene valor
            elif col_name == 'DOCUM.INCORP.':
                docum_incorp = fila.get('DOCUM.INCORP.', '')
                if pd.notna(docum_incorp) and str(docum_incorp).strip() != '':
                    # Fondo azul claro cuando hay documentación incorporada
                    self.set_fill_color(173, 216, 230)
                    self.rect(x, y, col_width, altura_fila, 'F')
                    self.set_fill_color(255, 255, 255)
                        
        except Exception as e:
            # Silenciar errores para no interrumpir la generación del PDF
            pass

# === CLASE PDFResumenKPI CORREGIDA (CON FUENTE QUE SOPORTA CARACTERES ESPECIALES) ===
class PDFResumenKPI(FPDF):
    def __init__(self):
        super().__init__()
        # Agregar soporte para caracteres latinos
        self.add_page()
        
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'RECTAUTO - Resumen de KPIs Semanales', 0, 1, 'C')
        self.ln(5)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Pagina {self.page_no()}', 0, 0, 'C')
    
    def add_section_title(self, title):
        self.set_font('Arial', 'B', 10)
        # Reemplazar caracteres problemáticos
        title_safe = title.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
        self.cell(0, 8, title_safe, 0, 1, 'L')
        self.ln(2)
    
    def add_metric(self, label, value, explanation=""):
        # Limpiar caracteres especiales de las etiquetas
        label_safe = label.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
        
        self.set_font('Arial', 'B', 9)
        self.cell(60, 6, label_safe, 0, 0)
        self.set_font('Arial', '', 9)
        
        # Limpiar también el valor si es texto
        if isinstance(value, str):
            value_safe = value.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
        else:
            value_safe = value
            
        self.cell(40, 6, str(value_safe), 0, 1)
        if explanation:
            self.set_font('Arial', 'I', 8)
            explanation_safe = explanation.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
            self.cell(0, 4, explanation_safe, 0, 1)
            self.ln(1)

# === FUNCIONES OPTIMIZADAS ===

# Funciones optimizadas con cache
@st.cache_data(ttl=CACHE_TTL, show_spinner="Procesando archivo Excel...")
def cargar_y_procesar_rectauto(archivo, _user_key=user_env.session_id):
    """Carga y procesa el archivo RECTAUTO con cache de 2 horas"""
    df = pd.read_excel(
        archivo, 
        sheet_name=HOJA, 
        header=0, 
        thousands='.', 
        decimal=',', 
        engine="openpyxl" if archivo.name.endswith("xlsx") else "xlrd"
    )
    df.columns = [col.upper().strip() for col in df.columns]
    columnas = [0, 1, 2, 3, 4, 5, 6, 12, 14, 15, 16, 17, 18, 20, 21, 23, 26, 27]
    df = df.iloc[:, columnas]
    return df

@st.cache_data(ttl=CACHE_TTL)
def cargar_y_procesar_notifica(archivo, _user_key=user_env.session_id):
    """Carga y procesa el archivo NOTIFICA"""
    try:
        df = pd.read_excel(archivo, sheet_name=HOJA)
        df.columns = [col.upper().strip() for col in df.columns]
        # Ordenar por RUE ORIGEN (ascendente) y FECHA APERTURA (descendente)
        if 'RUE ORIGEN' in df.columns and 'FECHA APERTURA' in df.columns:
            df['FECHA APERTURA'] = pd.to_datetime(df['FECHA APERTURA'], errors='coerce')
            df = df.sort_values(['RUE ORIGEN', 'FECHA APERTURA'], ascending=[True, False])
        
        # Mantener solo columnas relevantes
        columnas_a_mantener = ['RUE ORIGEN', 'FECHA NOTIFICACIÓN']
        columnas_existentes = [col for col in columnas_a_mantener if col in df.columns]
        df = df[columnas_existentes]
        
        return df
    except Exception as e:
        st.error(f"Error procesando NOTIFICA: {e}")
        return None

@st.cache_data(ttl=CACHE_TTL)
def cargar_y_procesar_triaje(archivo, _user_key=user_env.session_id):
    """Carga y procesa el archivo TRIAJE"""
    try:
        df = pd.read_excel(archivo, sheet_name='Triaje')
        df.columns = [col.upper().strip() for col in df.columns]
        
        # Crear RUE a partir de las primeras 4 columnas
        if df.shape[1] >= 4:
            # Formatear la cuarta columna a 6 dígitos
            df['RUE_TEMP'] = df.iloc[:, 3].astype(str).str.zfill(6)
            
            # Concatenar las cuatro primeras columnas
            df['RUE'] = (
                df.iloc[:, 0].astype(str) + 
                df.iloc[:, 1].astype(str) + 
                df.iloc[:, 2].astype(str) + 
                df['RUE_TEMP']
            )
            
            # Mantener solo columnas relevantes
            columnas_a_mantener = ['USUARIO-CSV', 'CALIFICACIÓN', 'OBSERVACIONES', 'FECHA ASIG']
            # Normalizar también los nombres de las columnas a mantener
            columnas_a_mantener = [col.upper().strip() for col in columnas_a_mantener]
            columnas_existentes = [col for col in columnas_a_mantener if col in df.columns]
            df = df[['RUE'] + columnas_existentes]
            
            return df
        else:
            st.warning("TRIAJE no tiene al menos 4 columnas")
            return None
    except Exception as e:
        st.error(f"Error procesando TRIAJE: {e}")
        return None

@st.cache_data(ttl=CACHE_TTL)
def cargar_y_procesar_usuarios(archivo, _user_key=user_env.session_id):
    """Carga y procesa el archivo USUARIOS"""
    try:
        df = pd.read_excel(archivo, sheet_name=HOJA)
        df.columns = [col.upper().strip() for col in df.columns]
        return df
    except Exception as e:
        st.error(f"Error procesando USUARIOS: {e}")
        return None

@st.cache_data(ttl=CACHE_TTL)
def cargar_y_procesar_documentos(archivo, _user_key=user_env.session_id):
    """Carga y procesa el archivo DOCUMENTOS"""
    try:
        # Cargar hoja DOCU para los valores del desplegable
        df_docu = pd.read_excel(archivo, sheet_name='DOCU')
        opciones_docu = df_docu.iloc[:, 0].dropna().tolist()
        
        # Cargar hoja DOCUMENTOS para los valores guardados
        df_documentos = pd.read_excel(archivo, sheet_name='DOCUMENTOS')
        df_documentos.columns = [col.upper().strip() for col in df_documentos.columns]
        
        return {
            'opciones': opciones_docu,
            'documentos': df_documentos,
            'archivo': archivo
        }
    except Exception as e:
        st.error(f"Error procesando DOCUMENTOS: {e}")
        return None

@st.cache_data(ttl=CACHE_TTL, show_spinner="Procesando archivos...")
def procesar_archivos_combinado(archivos_dict, _user_key=user_env.session_id):
    """Procesa todos los archivos en una sola función optimizada"""
    try:
        # Cargar RECTAUTO primero
        df_rectauto = cargar_y_procesar_rectauto(archivos_dict['rectauto'])
        
        # Cargar otros archivos en paralelo (si están disponibles)
        resultados = {}
        for nombre, archivo in archivos_dict.items():
            if archivo and nombre != 'rectauto':
                if nombre == 'notifica':
                    resultados['notifica'] = cargar_y_procesar_notifica(archivo)
                elif nombre == 'triaje':
                    resultados['triaje'] = cargar_y_procesar_triaje(archivo)
                elif nombre == 'usuarios':
                    resultados['usuarios'] = cargar_y_procesar_usuarios(archivo)
                elif nombre == 'documentos':
                    resultados['documentos'] = cargar_y_procesar_documentos(archivo)
        
        # Combinar todo
        df_combinado = combinar_archivos(
            df_rectauto, 
            resultados.get('notifica'),
            resultados.get('triaje'), 
            resultados.get('usuarios'),
            resultados.get('documentos')
        )
        
        return df_combinado, resultados.get('usuarios'), resultados.get('documentos')
        
    except Exception as e:
        st.error(f"Error en procesamiento combinado: {e}")
        return df_rectauto, None, None

def guardar_documentos_actualizados(archivo_original, df_documentos_actualizado):
    """Guarda los datos actualizados en el archivo DOCUMENTOS.xlsx"""
    try:
        # Usar directorio temporal único por usuario
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=user_env.working_dir)
        tmp_path = tmp_file.name
        tmp_file.close()  # 🔹 Muy importante: libera el archivo en Windows

        # Escribir las dos hojas en el archivo temporal
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            # Hoja DOCUMENTOS actualizada - SOLO los registros actuales
            df_documentos_actualizado.to_excel(writer, sheet_name="DOCUMENTOS", index=False)

            # Hoja DOCU - MANTENER las opciones del desplegable del original
            archivo_original.seek(0)
            df_docu_original = pd.read_excel(archivo_original, sheet_name="DOCU")
            df_docu_original.to_excel(writer, sheet_name="DOCU", index=False)

        # Leer contenido ya guardado
        with open(tmp_path, "rb") as f:
            contenido = f.read()

        # Eliminar el archivo temporal (libre ya de bloqueos)
        os.remove(tmp_path)

        return contenido

    except Exception as e:
        st.error(f"Error guardando DOCUMENTOS: {e}")
        return None

@st.cache_data(ttl=CACHE_TTL)
def combinar_archivos(rectauto_df, notifica_df=None, triaje_df=None, usuarios_df=None, documentos_data=None, _user_key=user_env.session_id):
    """Combina los archivos en un único DataFrame incluyendo DOCUM.INCORP."""
    df_combinado = rectauto_df.copy()
    
    # Combinar con NOTIFICA - VERSIÓN CORREGIDA (dejar vacío si no hay notificación válida)
    if notifica_df is not None and 'RUE ORIGEN' in notifica_df.columns:
        # Asegurar que las fechas están en formato datetime
        notifica_df['FECHA NOTIFICACIÓN'] = pd.to_datetime(notifica_df['FECHA NOTIFICACIÓN'], errors='coerce')
        
        # Ordenar por RUE ORIGEN (ascendente) y FECHA NOTIFICACIÓN (descendente)
        notifica_df = notifica_df.sort_values(['RUE ORIGEN', 'FECHA NOTIFICACIÓN'], ascending=[True, False])
        
        # Preparar dataframe para el merge
        notificaciones_finales = []
        
        # Primero necesitamos las fechas del penúltimo trámite de RECTAUTO
        if 'FECHA PENÚLTIMO TRAM.' in df_combinado.columns:
            # Crear un diccionario RUE -> FECHA PENÚLTIMO TRAM.
            rue_fecha_penultimo = df_combinado.set_index('RUE')['FECHA PENÚLTIMO TRAM.'].to_dict()
            
            # Para cada RUE en NOTIFICA, encontrar la notificación válida
            for rue, rue_group in notifica_df.groupby('RUE ORIGEN'):
                # Obtener la fecha del penúltimo trámite para este RUE
                fecha_penultimo = rue_fecha_penultimo.get(rue)
                
                if pd.isna(fecha_penultimo):
                    # Si no hay fecha de penúltimo trámite, NO tomar ninguna notificación
                    notificaciones_finales.append({
                        'RUE ORIGEN': rue,
                        'FECHA NOTIFICACIÓN': pd.NaT  # Dejar vacío
                    })
                else:
                    # Filtrar solo notificaciones con fecha igual o posterior al penúltimo trámite
                    notificaciones_validas = rue_group[rue_group['FECHA NOTIFICACIÓN'] >= fecha_penultimo]
                    
                    if not notificaciones_validas.empty:
                        # Tomar la más reciente de las válidas
                        notificacion_valida = notificaciones_validas.iloc[0]
                        notificaciones_finales.append({
                            'RUE ORIGEN': rue,
                            'FECHA NOTIFICACIÓN': notificacion_valida['FECHA NOTIFICACIÓN']
                        })
                    else:
                        # Si no hay notificaciones válidas, dejar vacío
                        notificaciones_finales.append({
                            'RUE ORIGEN': rue,
                            'FECHA NOTIFICACIÓN': pd.NaT  # Dejar vacío
                        })
        else:
            # Si no hay columna de fecha de penúltimo trámite, NO tomar ninguna notificación
            st.sidebar.warning("ℹ️ No se encontró columna 'FECHA PENÚLTIMO TRAM.', se dejarán vacías las notificaciones")
            # Crear estructura vacía
            for rue in notifica_df['RUE ORIGEN'].unique():
                notificaciones_finales.append({
                    'RUE ORIGEN': rue,
                    'FECHA NOTIFICACIÓN': pd.NaT  # Dejar vacío
                })
        
        # Convertir a DataFrame
        notifica_filtrada = pd.DataFrame(notificaciones_finales)
        
        # Realizar el merge con las notificaciones filtradas
        df_combinado = pd.merge(
            df_combinado, 
            notifica_filtrada, 
            left_on='RUE', 
            right_on='RUE ORIGEN', 
            how='left'
        )
        
        # Eliminar la columna RUE ORIGEN ya que ya tenemos RUE
        if 'RUE ORIGEN' in df_combinado.columns:
            df_combinado.drop('RUE ORIGEN', axis=1, inplace=True)
        
        # Contar notificaciones válidas
        notificaciones_validas_count = notifica_filtrada['FECHA NOTIFICACIÓN'].notna().sum()
        st.sidebar.info(f"✅ NOTIFICA combinado: {notificaciones_validas_count} notificaciones válidas de {len(notifica_filtrada)} RUEs")
    else:
        if notifica_df is not None:
            st.sidebar.warning("ℹ️ NOTIFICA no tiene columna 'RUE ORIGEN'")
    
    # Combinar con TRIAJE
    # Antes de cada merge, verificar duplicados
    if triaje_df is not None and 'RUE' in triaje_df.columns:
        # Verificar duplicados en triaje_df
        duplicados_triaje = triaje_df['RUE'].duplicated().sum()
        if duplicados_triaje > 0:
            st.sidebar.warning(f"⚠️ TRIAJE tiene {duplicados_triaje} RUEs duplicados")
            # Mantener el último registro por RUE
            triaje_df = triaje_df.drop_duplicates(subset='RUE', keep='last')
        
        df_combinado = pd.merge(
            df_combinado, 
            triaje_df, 
            on='RUE', 
            how='left',
            validate='one_to_one'  # Validar que no haya múltiples matches
        )
        st.sidebar.info(f"✅ TRIAJE combinado: {len(triaje_df)} registros")
    
    # CORRECCIÓN: Rellenar FECHA ASIG vacías con FECHA APERTURA
    if 'FECHA ASIG' in df_combinado.columns and 'FECHA APERTURA' in df_combinado.columns:
        # Contar cuántas fechas ASIG están vacías
        asig_vacias_antes = df_combinado['FECHA ASIG'].isna().sum()
        
        # Rellenar las vacías con FECHA APERTURA
        mask_vacias = df_combinado['FECHA ASIG'].isna()
        df_combinado.loc[mask_vacias, 'FECHA ASIG'] = df_combinado.loc[mask_vacias, 'FECHA APERTURA']
        
        # Contar cuántas se rellenaron
        asig_vacias_despues = df_combinado['FECHA ASIG'].isna().sum()
        rellenadas = asig_vacias_antes - asig_vacias_despues
        
        if rellenadas > 0:
            st.sidebar.info(f"📅 {rellenadas} fechas ASIG vacías rellenadas con FECHA APERTURA")
    
    # Añadir columna DOCUM.INCORP. después de FECHA NOTIFICACIÓN
    if 'FECHA NOTIFICACIÓN' in df_combinado.columns:
        # Buscar la posición de FECHA NOTIFICACIÓN
        columnas = df_combinado.columns.tolist()
        pos_fecha_notif = columnas.index('FECHA NOTIFICACIÓN')
        
        # Insertar DOCUM.INCORP. después de FECHA NOTIFICACIÓN
        columnas.insert(pos_fecha_notif + 1, 'DOCUM.INCORP.')
        df_combinado = df_combinado.reindex(columns=columnas)
        
        # Si tenemos datos de documentación, rellenar los valores
        if documentos_data is not None and not documentos_data['documentos'].empty:
            df_documentos = documentos_data['documentos']
            # Crear un mapeo RUE -> DOCUMENTACION
            mapeo_documentos = df_documentos.set_index('RUE')['DOCUM.INCORP.'].to_dict()
            # Aplicar el mapeo a la nueva columna
            df_combinado['DOCUM.INCORP.'] = df_combinado['RUE'].map(mapeo_documentos)
    
    return df_combinado

@st.cache_data(ttl=CACHE_TTL)
def convertir_fechas(df, _user_key=user_env.session_id):
    """Convierte columnas con 'FECHA' en el nombre a datetime"""
    for col in df.columns:
        if 'FECHA' in col.upper():
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df

# === FUNCIONES DE CÁLCULO OPTIMIZADAS ===

def calcular_despachados_optimizado(_df, inicio_semana, fin_semana, fecha_inicio_totales):
    """Calcula despachados de forma optimizada usando operaciones vectorizadas"""
    if not all(col in _df.columns for col in ['FECHA RESOLUCIÓN', 'ESTADO', 'FECHA CIERRE']):
        return 0, 0

    fecha_9999 = pd.to_datetime('9999-09-09', errors='coerce')

    # Expedientes con FECHA RESOLUCIÓN real (distinta de 9999 y no nula) dentro del rango semanal
    mask_despachados_semana_reales = (
        _df['FECHA RESOLUCIÓN'].notna() & 
        (_df['FECHA RESOLUCIÓN'] != fecha_9999) &
        (_df['FECHA RESOLUCIÓN'] >= inicio_semana) &
        (_df['FECHA RESOLUCIÓN'] <= fin_semana)
    )

    # Expedientes CERRADOS con FECHA RESOLUCIÓN = 9999-09-09 o vacía
    mask_despachados_semana_cerrados = (
        (_df['ESTADO'] == 'Cerrado') &
        (_df['FECHA RESOLUCIÓN'].isna() | (_df['FECHA RESOLUCIÓN'] == fecha_9999)) &
        _df['FECHA CIERRE'].notna() &
        (_df['FECHA CIERRE'] >= inicio_semana) &
        (_df['FECHA CIERRE'] <= fin_semana)
    )

    mask_despachados_semana = mask_despachados_semana_reales | mask_despachados_semana_cerrados
    despachados_semana = _df[mask_despachados_semana].shape[0]

    # Totales: igual pero usando fecha_inicio_totales
    mask_despachados_totales_reales = (
        _df['FECHA RESOLUCIÓN'].notna() & 
        (_df['FECHA RESOLUCIÓN'] != fecha_9999) &
        (_df['FECHA RESOLUCIÓN'] >= fecha_inicio_totales) &
        (_df['FECHA RESOLUCIÓN'] <= fin_semana)
    )

    mask_despachados_totales_cerrados = (
        (_df['ESTADO'] == 'Cerrado') &
        (_df['FECHA RESOLUCIÓN'].isna() | (_df['FECHA RESOLUCIÓN'] == fecha_9999)) &
        _df['FECHA CIERRE'].notna() &
        (_df['FECHA CIERRE'] >= fecha_inicio_totales) &
        (_df['FECHA CIERRE'] <= fin_semana)
    )

    mask_despachados_totales = mask_despachados_totales_reales | mask_despachados_totales_cerrados
    despachados_totales = _df[mask_despachados_totales].shape[0] + 6  # HAY 6 EXPEDIENTES QUE SE CERRARON ANTES DEL 1/11/2022

    return despachados_semana, despachados_totales

# === FUNCIÓN OPTIMIZADA PARA CÁLCULO DE TIEMPOS ===
@st.cache_data(ttl=CACHE_TTL)
def calcular_tiempos_optimizado(_df, fecha_inicio_totales, fin_semana):
    """Versión completamente optimizada del cálculo de tiempos - MÁS RÁPIDO"""
    
    resultados = {
        'tiempo_medio_despachados': 0,
        'percentil_90_despachados': 0,
        'percentil_180_despachados': 0,
        'percentil_120_despachados': 0,
        'tiempo_medio_cerrados': 0,
        'percentil_90_cerrados': 0,
        'percentil_180_cerrados': 0,
        'percentil_120_cerrados': 0,
        'percentil_90_abiertos': 0,
        'percentil_180_abiertos': 0,
        'percentil_120_abiertos': 0,
        'percentil_90_abiertos_no_despachados': 0,  # NUEVO
        'percentil_180_abiertos_no_despachados': 0,  # NUEVO
        'percentil_120_abiertos_no_despachados': 0   # NUEVO
    }
    
    try:
        # VERIFICACIÓN RÁPIDA DE COLUMNAS NECESARIAS
        columnas_necesarias = ['FECHA INICIO TRAMITACIÓN']
        if not all(col in _df.columns for col in columnas_necesarias):
            return resultados
        
        # CREAR COPIA RÁPIDA con solo las columnas necesarias
        columnas_calculo = ['FECHA RESOLUCIÓN', 'FECHA INICIO TRAMITACIÓN', 'ESTADO', 'FECHA CIERRE', 'FECHA APERTURA']
        columnas_existentes = [col for col in columnas_calculo if col in _df.columns]
        
        if not columnas_existentes:
            return resultados
            
        # Las fechas ya vienen convertidas desde combinar_archivos/convertir_fechas.
        df_temp = _df[columnas_existentes]

        fecha_9999 = pd.to_datetime('9999-09-09', errors='coerce')
        fin_semana = pd.to_datetime(fin_semana)
        fecha_inicio_totales = pd.to_datetime(fecha_inicio_totales)
        
        # ===== TIEMPOS DESPACHADOS - VECTORIZADO =====
        if all(col in df_temp.columns for col in ['FECHA RESOLUCIÓN', 'FECHA INICIO TRAMITACIÓN', 'ESTADO', 'FECHA CIERRE']):
            # MÁSCARAS VECTORIZADAS
            mask_resolucion_valida = (
                df_temp['FECHA RESOLUCIÓN'].notna() & 
                (df_temp['FECHA RESOLUCIÓN'] != fecha_9999) &
                (df_temp['FECHA RESOLUCIÓN'] >= fecha_inicio_totales) &
                (df_temp['FECHA RESOLUCIÓN'] <= fin_semana)
            )
            
            mask_cierre_valido = (
                (df_temp['ESTADO'] == 'Cerrado') &
                (df_temp['FECHA RESOLUCIÓN'].isna() | (df_temp['FECHA RESOLUCIÓN'] == fecha_9999)) &
                df_temp['FECHA CIERRE'].notna() &
                (df_temp['FECHA CIERRE'] >= fecha_inicio_totales) &
                (df_temp['FECHA CIERRE'] <= fin_semana)
            )
            
            mask_despachados = mask_resolucion_valida | mask_cierre_valido
            
            if mask_despachados.any():
                df_despachados = df_temp[mask_despachados]
                
                # CALCULAR FECHA FINAL VECTORIZADO
                df_despachados['FECHA_FINAL'] = df_despachados['FECHA RESOLUCIÓN']
                mask_usar_cierre = df_despachados['FECHA RESOLUCIÓN'].isna() | (df_despachados['FECHA RESOLUCIÓN'] == fecha_9999)
                df_despachados.loc[mask_usar_cierre, 'FECHA_FINAL'] = df_despachados.loc[mask_usar_cierre, 'FECHA CIERRE']
                
                # CALCULAR DÍAS VECTORIZADO
                df_despachados['dias_tramitacion'] = (df_despachados['FECHA_FINAL'] - df_despachados['FECHA INICIO TRAMITACIÓN']).dt.days + 1
                dias_validos = df_despachados['dias_tramitacion'][df_despachados['dias_tramitacion'] >= 0]
                
                if not dias_validos.empty:
                    resultados['tiempo_medio_despachados'] = round(dias_validos.mean(), 0)
                    resultados['percentil_90_despachados'] = dias_validos.quantile(0.9, interpolation='higher')
                    resultados['percentil_180_despachados'] = round((dias_validos <= 180).mean() * 100, 2)
                    resultados['percentil_120_despachados'] = round((dias_validos <= 120).mean() * 100, 2)
        
        # ===== TIEMPOS CERRADOS - VECTORIZADO =====
        if all(col in df_temp.columns for col in ['FECHA CIERRE', 'FECHA INICIO TRAMITACIÓN']):
            mask_cerrados = (
                df_temp['FECHA CIERRE'].notna() &
                (df_temp['FECHA CIERRE'] >= fecha_inicio_totales) & 
                (df_temp['FECHA CIERRE'] <= fin_semana)
            )
            
            if mask_cerrados.any():
                df_cerrados = df_temp[mask_cerrados]
                df_cerrados['dias_tramitacion'] = (df_cerrados['FECHA CIERRE'] - df_cerrados['FECHA INICIO TRAMITACIÓN']).dt.days + 1
                dias_validos = df_cerrados['dias_tramitacion'][df_cerrados['dias_tramitacion'] >= 0]
                
                if not dias_validos.empty:
                    resultados['tiempo_medio_cerrados'] = round(dias_validos.mean(), 0)
                    resultados['percentil_90_cerrados'] = dias_validos.quantile(0.9, interpolation='higher')
                    resultados['percentil_180_cerrados'] = round((dias_validos <= 180).mean() * 100, 2)
                    resultados['percentil_120_cerrados'] = round((dias_validos <= 120).mean() * 100, 2)
        
        # ===== TIEMPOS ABIERTOS - VECTORIZADO =====
        if all(col in df_temp.columns for col in ['FECHA INICIO TRAMITACIÓN', 'FECHA APERTURA', 'FECHA CIERRE']):
            mask_abiertos = (
                (df_temp['FECHA APERTURA'] <= fin_semana) & 
                ((df_temp['FECHA CIERRE'] > fin_semana) | (df_temp['FECHA CIERRE'].isna()))
            )
            
            if mask_abiertos.any():
                df_abiertos = df_temp[mask_abiertos]
                df_abiertos['dias_tramitacion'] = (fin_semana - df_abiertos['FECHA INICIO TRAMITACIÓN']).dt.days + 1
                dias_validos = df_abiertos['dias_tramitacion'][df_abiertos['dias_tramitacion'] >= 0]
                
                if not dias_validos.empty:
                    resultados['percentil_90_abiertos'] = dias_validos.quantile(0.9, interpolation='higher')
                    resultados['percentil_180_abiertos'] = round((dias_validos <= 180).mean() * 100, 2)
                    resultados['percentil_120_abiertos'] = round((dias_validos <= 120).mean() * 100, 2)

        # ===== TIEMPOS ABIERTOS NO DESPACHADOS - VECTORIZADO =====
        if all(col in df_temp.columns for col in ['FECHA INICIO TRAMITACIÓN', 'FECHA APERTURA', 'FECHA CIERRE', 'FECHA RESOLUCIÓN', 'ESTADO']):
            # Expedientes abiertos
            mask_abiertos = (
                (df_temp['FECHA APERTURA'] <= fin_semana) & 
                ((df_temp['FECHA CIERRE'] > fin_semana) | (df_temp['FECHA CIERRE'].isna()))
            )
            
            # Expedientes despachados hasta fin_semana
            mask_despachados_reales = (
                df_temp['FECHA RESOLUCIÓN'].notna() & 
                (df_temp['FECHA RESOLUCIÓN'] != fecha_9999) &
                (df_temp['FECHA RESOLUCIÓN'] >= fecha_inicio_totales) &
                (df_temp['FECHA RESOLUCIÓN'] <= fin_semana)
            )
            
            mask_despachados_cerrados = (
                (df_temp['ESTADO'] == 'Cerrado') &
                (df_temp['FECHA RESOLUCIÓN'].isna() | (df_temp['FECHA RESOLUCIÓN'] == fecha_9999)) &
                df_temp['FECHA CIERRE'].notna() &
                (df_temp['FECHA CIERRE'] >= fecha_inicio_totales) &
                (df_temp['FECHA CIERRE'] <= fin_semana)
            )
            
            mask_despachados_total = mask_despachados_reales | mask_despachados_cerrados
            
            # Expedientes abiertos no despachados = Abiertos - Despachados
            mask_abiertos_no_despachados = mask_abiertos & (~mask_despachados_total)
            
            if mask_abiertos_no_despachados.any():
                df_abiertos_no_desp = df_temp[mask_abiertos_no_despachados]
                df_abiertos_no_desp['dias_tramitacion'] = (fin_semana - df_abiertos_no_desp['FECHA INICIO TRAMITACIÓN']).dt.days + 1
                dias_validos = df_abiertos_no_desp['dias_tramitacion'][df_abiertos_no_desp['dias_tramitacion'] >= 0]
                
                if not dias_validos.empty:
                    resultados['percentil_90_abiertos_no_despachados'] = dias_validos.quantile(0.9, interpolation='higher')
                    resultados['percentil_180_abiertos_no_despachados'] = round((dias_validos <= 180).mean() * 100, 2)
                    resultados['percentil_120_abiertos_no_despachados'] = round((dias_validos <= 120).mean() * 100, 2)
        
    except Exception as e:
        # Error silencioso para no interrumpir
        pass
    
    return resultados

# === FUNCIÓN OPTIMIZADA PARA CÁLCULO DE KPIs POR SEMANA ===
@st.cache_data(ttl=CACHE_TTL)
def calcular_kpis_para_semana_optimizado(_df, semana_fin, es_semana_actual=False):
    """Versión ultra optimizada del cálculo de KPIs"""
    
    # DETERMINAR RANGO SEMANAL
    if es_semana_actual:
        inicio_semana = semana_fin - timedelta(days=7)
        fin_semana = semana_fin
        dias_semana = 8
    else:
        inicio_semana = semana_fin - timedelta(days=7)
        fin_semana = semana_fin - timedelta(days=1)
        dias_semana = 7
    
    fecha_inicio_totales = datetime(2022, 11, 1)
    
    # CONVERSIONES RÁPIDAS
    inicio_semana = pd.to_datetime(inicio_semana)
    fin_semana = pd.to_datetime(fin_semana)
    
    # PRE-CÁLCULO DE MÁSCARAS REUTILIZABLES
    resultados = {
        'nuevos_expedientes': 0,
        'nuevos_expedientes_totales': 0,
        'despachados_semana': 0,
        'despachados_totales': 0,
        'expedientes_cerrados': 0,
        'expedientes_cerrados_totales': 0,
        'total_abiertos': 0,
        'total_abiertos_no_despachados': 0,  # NUEVO KPI
        'total_rehabilitados': 0,
        'expedientes_especiales': 0,
        'porcentaje_especiales': 0,
        'inicio_semana': inicio_semana,
        'fin_semana': fin_semana,
        'dias_semana': dias_semana,
        'es_semana_actual': es_semana_actual
    }

    try:
        # NUEVOS EXPEDIENTES
        if 'FECHA ASIG' in _df.columns:
            mask_semana = (_df['FECHA ASIG'] >= inicio_semana) & (_df['FECHA ASIG'] <= fin_semana)
            mask_totales = (_df['FECHA APERTURA'] >= fecha_inicio_totales) & (_df['FECHA APERTURA'] <= fin_semana)
            
            resultados['nuevos_expedientes'] = mask_semana.sum()
            resultados['nuevos_expedientes_totales'] = mask_totales.sum()
        
        # EXPEDIENTES DESPACHADOS
        despachados_semana, despachados_totales = calcular_despachados_optimizado(_df, inicio_semana, fin_semana, fecha_inicio_totales)
        resultados['despachados_semana'] = despachados_semana
        resultados['despachados_totales'] = despachados_totales
        
        # EXPEDIENTES CERRADOS
        if 'FECHA CIERRE' in _df.columns:
            mask_cerrados_semana = (_df['FECHA CIERRE'] >= inicio_semana) & (_df['FECHA CIERRE'] <= fin_semana)
            mask_cerrados_totales = (_df['FECHA CIERRE'] >= fecha_inicio_totales) & (_df['FECHA CIERRE'] <= fin_semana)
            
            resultados['expedientes_cerrados'] = mask_cerrados_semana.sum()
            resultados['expedientes_cerrados_totales'] = mask_cerrados_totales.sum()+6 # HAY 6 EXPEDIENTES ASIGNADOS QUE ESTÁN CERRADOS ANTES DEL 1/11/2022
        
        # COEFICIENTES DE ABSORCIÓN
        if resultados['nuevos_expedientes'] > 0:
            resultados['c_abs_despachados_sem'] = (resultados['despachados_semana'] / resultados['nuevos_expedientes'] * 100)
            resultados['c_abs_cerrados_sem'] = (resultados['expedientes_cerrados'] / resultados['nuevos_expedientes'] * 100)
        else:
            resultados['c_abs_despachados_sem'] = 0
            resultados['c_abs_cerrados_sem'] = 0
            
        if resultados['nuevos_expedientes_totales'] > 0:
            resultados['c_abs_despachados_tot'] = (resultados['despachados_totales'] / resultados['nuevos_expedientes_totales'] * 100)
            resultados['c_abs_cerrados_tot'] = (resultados['expedientes_cerrados_totales'] / resultados['nuevos_expedientes_totales'] * 100)
        else:
            resultados['c_abs_despachados_tot'] = 0
            resultados['c_abs_cerrados_tot'] = 0
        
        # EXPEDIENTES ABIERTOS
        if 'FECHA CIERRE' in _df.columns and 'FECHA APERTURA' in _df.columns:
            mask_abiertos = (_df['FECHA APERTURA'] <= fin_semana) & ((_df['FECHA CIERRE'] > fin_semana) | (_df['FECHA CIERRE'].isna()))
            resultados['total_abiertos'] = mask_abiertos.sum()

        # EXPEDIENTES ABIERTOS NO DESPACHADOS (NUEVO KPI)
        if all(col in _df.columns for col in ['FECHA APERTURA', 'FECHA CIERRE', 'FECHA RESOLUCIÓN', 'ESTADO']):
            # Fecha 9999 para comparación
            fecha_9999 = pd.to_datetime('9999-09-09', errors='coerce')
            
            # Expedientes despachados hasta fin_semana (misma lógica que calcular_despachados_optimizado)
            # 1. Expedientes con FECHA RESOLUCIÓN real (distinta de 9999 y no nula) hasta fin_semana
            mask_despachados_reales = (
                _df['FECHA RESOLUCIÓN'].notna() & 
                (_df['FECHA RESOLUCIÓN'] != fecha_9999) &
                (_df['FECHA RESOLUCIÓN'] <= fin_semana)
            )
            
            # 2. Expedientes CERRADOS con FECHA RESOLUCIÓN = 9999-09-09 o vacía hasta fin_semana
            mask_despachados_cerrados = (
                (_df['ESTADO'] == 'Cerrado') &
                (_df['FECHA RESOLUCIÓN'].isna() | (_df['FECHA RESOLUCIÓN'] == fecha_9999)) &
                _df['FECHA CIERRE'].notna() &
                (_df['FECHA CIERRE'] <= fin_semana)
            )
            
            mask_despachados_total = mask_despachados_reales | mask_despachados_cerrados
            
            # Expedientes abiertos no despachados = Abiertos - Despachados
            mask_abiertos_no_despachados = mask_abiertos & (~mask_despachados_total)
            resultados['total_abiertos_no_despachados'] = mask_abiertos_no_despachados.sum()
        
        # EXPEDIENTES REHABILITADOS
        if 'ETIQ. PENÚLTIMO TRAM.' in _df.columns:
            mask_rehabilitados = (
                ~_df['FECHA CIERRE'].isna() &
                _df['ESTADO'].isin(['Abierto'])
            )
            resultados['total_rehabilitados'] = mask_rehabilitados.sum()
        
        # EXPEDIENTES ESPECIALES
        if 'ETIQ. PENÚLTIMO TRAM.' in _df.columns:
            mask_especiales = (
                ~_df['ETIQ. PENÚLTIMO TRAM.'].isin(['1 APERTURA', '10 DATEXPTE']) &
                _df['ESTADO'].isin(['Abierto'])
            )
            mask_abiertos_ultima_semana = _df['ESTADO'].isin(['Abierto'])
            
            resultados['expedientes_especiales'] = mask_especiales.sum()
            total_abiertos_ultima_semana = mask_abiertos_ultima_semana.sum()
            
            if total_abiertos_ultima_semana > 0:
                resultados['porcentaje_especiales'] = (resultados['expedientes_especiales'] / total_abiertos_ultima_semana * 100)
        
        # CÁLCULO DE TIEMPOS (USANDO FUNCIÓN OPTIMIZADA)
        tiempos = calcular_tiempos_optimizado(_df, fecha_inicio_totales, fin_semana)
        resultados.update(tiempos)
        
    except Exception as e:
        # En caso de error, devolver resultados básicos
        pass
    
    return resultados

def identificar_filas_prioritarias(df):
    """Versión optimizada usando operaciones vectorizadas - MEJORADA"""
    try:
        # Crear copia para no modificar el original
        df_priorizado = df.copy(deep=True)
        
        # Inicializar todas las prioridades como 0
        df_priorizado['_prioridad'] = 0
        
        # Verificar que las columnas necesarias existen
        columnas_necesarias = ['ETIQ. PENÚLTIMO TRAM.', 'FECHA NOTIFICACIÓN', 'DOCUM.INCORP.']
        if not all(col in df_priorizado.columns for col in columnas_necesarias):
            return df_priorizado
        
        # Crear máscaras vectorizadas para cada condición
        etiq_penultimo = df_priorizado['ETIQ. PENÚLTIMO TRAM.'].astype(str).str.strip()
        fecha_notif = pd.to_datetime(df_priorizado['FECHA NOTIFICACIÓN'], errors='coerce')
        docum_incorp = df_priorizado['DOCUM.INCORP.'].astype(str).str.strip()
        
        # Inicializar máscara de prioritarios
        mask_prioritarios = pd.Series(False, index=df_priorizado.index)
        
        # CONDICIÓN 1: "80 PROPRES" con fecha límite superada
        mask_80_propres = (etiq_penultimo == "80 PROPRES")
        if mask_80_propres.any():
            fechas_limite_80 = fecha_notif[mask_80_propres] + timedelta(days=23)
            mask_80_vencido = mask_80_propres & (datetime.now() > fechas_limite_80)
            mask_prioritarios = mask_prioritarios | mask_80_vencido
        
        # CONDICIÓN 2: "50 REQUERIR" con fecha límite superada  
        mask_50_requerir = (etiq_penultimo == "50 REQUERIR")
        if mask_50_requerir.any():
            fechas_limite_50 = fecha_notif[mask_50_requerir] + timedelta(days=23)
            mask_50_vencido = mask_50_requerir & (datetime.now() > fechas_limite_50)
            mask_prioritarios = mask_prioritarios | mask_50_vencido
        
        # CONDICIÓN 3: "70 ALEGACI" o "60 CONTESTA"
        mask_alegaci_contesta = etiq_penultimo.isin(["70 ALEGACI", "60 CONTESTA"])
        mask_prioritarios = mask_prioritarios | mask_alegaci_contesta
        
        # CONDICIÓN 4: DOCUM.INCORP. válido
        mask_docum_valido = (
            docum_incorp.notna() & 
            (docum_incorp != '') &
            (docum_incorp != 'nan') &
            (~docum_incorp.str.upper().isin(["SOLICITUD", "REITERA SOLICITUD"]))
        )
        mask_prioritarios = mask_prioritarios | mask_docum_valido
        
        # Asignar prioridad 1 a los que cumplen alguna condición
        df_priorizado.loc[mask_prioritarios, '_prioridad'] = 1
        
        return df_priorizado
    
    except Exception as e:
        st.error(f"❌ Error al identificar filas prioritarias: {e}")
        # En caso de error, devolver DataFrame con prioridad 0
        df['_prioridad'] = 0
        return df

@st.cache_data(ttl=3600)
def dataframe_to_pdf_bytes(df_mostrar, title, df_original):
    """Versión optimizada de generación de PDFs"""
    try:
        pdf = PDF('L', 'mm', 'A4')
        pdf.add_page()
        
        pdf.set_font("Arial", "B", 8)
        pdf.cell(0, 5, title, 0, 1, 'C')
        pdf.ln(5)

        # PRE-CALCULAR estructura de columnas
        columnas_a_excluir = {
            idx for idx, col_name in enumerate(df_mostrar.columns) 
            if "FECHA DE ACTUALIZACIÓN DATOS" in col_name.upper()
        }
        
        # Filtrar columnas una sola vez
        columnas_visibles = [
            (idx, col_name, col_width) 
            for idx, (col_name, col_width) in enumerate(zip(df_mostrar.columns, COL_WIDTHS_OPTIMIZED))
            if idx not in columnas_a_excluir
        ]

        ALTURA_ENCABEZADO = 13
        ALTURA_LINEA = 3
        ALTURA_BASE_FILA = 2

        # --- Encabezados de tabla ---
        def imprimir_encabezados():
            pdf.set_font("Arial", "", 5)
            pdf.set_fill_color(200, 220, 255)
            y_inicio = pdf.get_y()

            for i, header in enumerate(df_mostrar.columns):
                # EXCLUIR columna "FECHA DE ACTUALIZACIÓN DATOS"
                if "FECHA DE ACTUALIZACIÓN DATOS" in header.upper():
                    continue
                    
                x = pdf.get_x()
                y = pdf.get_y()
                pdf.cell(COL_WIDTHS_OPTIMIZED[i], ALTURA_ENCABEZADO, "", 1, 0, 'C', True)
                pdf.set_xy(x, y)

                texto = str(header)
                ancho_texto = pdf.get_string_width(texto)

                if ancho_texto <= COL_WIDTHS_OPTIMIZED[i] - 2:
                    altura_texto = 3
                    y_pos = y + (ALTURA_ENCABEZADO - altura_texto) / 2
                    pdf.set_xy(x, y_pos)
                    pdf.cell(COL_WIDTHS_OPTIMIZED[i], altura_texto, texto, 0, 0, 'C')
                else:
                    pdf.set_xy(x, y + 1)
                    pdf.multi_cell(COL_WIDTHS_OPTIMIZED[i], 2.5, texto, 0, 'C')

                pdf.set_xy(x + COL_WIDTHS_OPTIMIZED[i], y)

            pdf.set_xy(pdf.l_margin, y_inicio + ALTURA_ENCABEZADO)

        imprimir_encabezados()
        pdf.set_font("Arial", "", 5)

        # --- PRE-CÁLCULO: columnas visibles con anchos acumulados ---
        cols_info = [
            (i, col_name, COL_WIDTHS_OPTIMIZED[i])
            for i, col_name in enumerate(df_mostrar.columns)
            if "FECHA DE ACTUALIZACIÓN DATOS" not in col_name.upper()
            and i < len(COL_WIDTHS_OPTIMIZED)
        ]
        anchos_acum = []
        acum = 0
        for _, _, w in cols_info:
            anchos_acum.append(acum)
            acum += w

        ANCHO_MIN = min(COL_WIDTHS_OPTIMIZED) - 2

        # --- PRE-CALCULAR MASKS DE COLOR (vectorizado, una sola vez) ---
        n = len(df_original)
        mask_amarillo = [False] * n
        mask_rojo     = [False] * n
        mask_azul     = [False] * n

        try:
            if 'ETIQ. PENÚLTIMO TRAM.' in df_original.columns:
                _prio = identificar_filas_prioritarias(df_original)
                mask_amarillo = (_prio['_prioridad'] == 1).tolist()
        except Exception:
            pass

        try:
            if 'USUARIO' in df_original.columns and 'USUARIO-CSV' in df_original.columns:
                _u1 = df_original['USUARIO'].astype(str).str.strip()
                _u2 = df_original['USUARIO-CSV'].astype(str).str.strip()
                mask_rojo = (
                    df_original['USUARIO'].notna() &
                    df_original['USUARIO-CSV'].notna() &
                    (_u1 != _u2)
                ).tolist()
        except Exception:
            pass

        try:
            if 'DOCUM.INCORP.' in df_original.columns:
                _d = df_original['DOCUM.INCORP.'].astype(str).str.strip()
                mask_azul = (
                    df_original['DOCUM.INCORP.'].notna() &
                    (_d != '') & (_d != 'nan') &
                    (~_d.str.upper().isin(["SOLICITUD", "REITERA SOLICITUD"]))
                ).tolist()
        except Exception:
            pass

        # --- Convertir a lista Python pura para acceso O(1) ---
        datos_lista = df_mostrar.values.tolist()

        # --- Filas de datos ---
        for idx, row_datos in enumerate(datos_lista):
            # Calcular altura de fila
            max_lineas = 1
            for ci, (col_idx, col_name, col_width) in enumerate(cols_info):
                texto = str(row_datos[col_idx]).replace("\n", " ")
                if texto.lower() in ("nan", "") or not texto.strip():
                    continue
                ancho_texto = pdf.get_string_width(texto)
                if ancho_texto > ANCHO_MIN:
                    lineas = max(1, int(ancho_texto / ANCHO_MIN) + 1)
                    if lineas > max_lineas:
                        max_lineas = lineas

            altura_fila = ALTURA_BASE_FILA + ((max_lineas - 1) * ALTURA_LINEA) / 2

            if pdf.get_y() + altura_fila > 190:
                pdf.add_page()
                imprimir_encabezados()

            x_inicio = pdf.get_x()
            y_inicio = pdf.get_y()

            # Bordes
            for ci, (col_idx, col_name, col_width) in enumerate(cols_info):
                pdf.rect(x_inicio + anchos_acum[ci], y_inicio, col_width, altura_fila)

            # Contenido + color (usando masks pre-calculadas)
            es_am = mask_amarillo[idx] if idx < n else False
            es_ro = mask_rojo[idx]     if idx < n else False
            es_az = mask_azul[idx]     if idx < n else False

            for ci, (col_idx, col_name, col_width) in enumerate(cols_info):
                texto = str(row_datos[col_idx]).replace("\n", " ")
                if texto.lower() in ("nan", ""):
                    texto = ""

                x_celda = x_inicio + anchos_acum[ci]

                # Aplicar color de fondo sin llamar a iloc
                if col_name == 'RUE' and es_am:
                    pdf.set_fill_color(255, 255, 0)
                    pdf.rect(x_celda, y_inicio, col_width, altura_fila, 'F')
                    pdf.set_fill_color(255, 255, 255)
                elif col_name == 'USUARIO-CSV' and es_ro:
                    pdf.set_fill_color(255, 0, 0)
                    pdf.rect(x_celda, y_inicio, col_width, altura_fila, 'F')
                    pdf.set_fill_color(255, 255, 255)
                elif col_name == 'DOCUM.INCORP.' and es_az:
                    pdf.set_fill_color(173, 216, 230)
                    pdf.rect(x_celda, y_inicio, col_width, altura_fila, 'F')
                    pdf.set_fill_color(255, 255, 255)

                pdf.set_xy(x_celda, y_inicio)
                pdf.multi_cell(col_width, ALTURA_LINEA, texto, 0, 'L')

            pdf.set_xy(pdf.l_margin, y_inicio + altura_fila)

        # --- Exportar a bytes (compatible con todas las versiones de fpdf2) ---
        pdf_output = pdf.output(dest='S')

        # Normalizar salida: puede ser str, bytes o bytearray según versión de fpdf2
        if isinstance(pdf_output, str):
            pdf_bytes = pdf_output.encode('latin1')
        elif isinstance(pdf_output, (bytes, bytearray)):
            pdf_bytes = bytes(pdf_output)
        else:
            raise TypeError(f"Tipo inesperado devuelto por fpdf.output(): {type(pdf_output)}")

        return io.BytesIO(pdf_bytes).getvalue()

    except Exception as e:
        st.error(f"Error generando PDF: {e}")
        return None

# === FUNCIONES ORIGINALES (MANTENIDAS POR COMPATIBILIDAD) ===

def ordenar_dataframe_por_prioridad_y_antiguedad(df):
    """Ordena el DataFrame: RUE amarillos primero, luego por antigüedad descendente - CORREGIDA DEFINITIVAMENTE"""
    try:
        # Crear una COPIA PROFUNDA para no modificar el original
        df_priorizado = df.copy(deep=True)
        
        # Identificar filas prioritarias en la copia
        df_priorizado = identificar_filas_prioritarias(df_priorizado)
        
        # Buscar la columna de antigüedad
        columnas_antiguedad = [col for col in df_priorizado.columns if 'ANTIGÜEDAD' in col.upper() or 'DÍAS' in col.upper()]
        
        if columnas_antiguedad:
            columna_antiguedad = columnas_antiguedad[0]
            # 🔥 CORRECCIÓN CRÍTICA: NO MODIFICAR LA COLUMNA DE ANTIGÜEDAD
            # Solo asegurarnos de que es numérica para ordenar, pero sin cambiar los valores
            if not pd.api.types.is_numeric_dtype(df_priorizado[columna_antiguedad]):
                # Si no es numérica, convertir temporalmente solo para ordenar
                df_priorizado['_antiguedad_temp'] = pd.to_numeric(
                    df_priorizado[columna_antiguedad], 
                    errors='coerce'
                ).fillna(0)
                columna_para_ordenar = '_antiguedad_temp'
            else:
                columna_para_ordenar = columna_antiguedad
        else:
            st.warning("⚠️ No se encontró columna de antigüedad, usando orden por prioridad solamente")
            columna_para_ordenar = None
        
        # Ordenar por prioridad (True primero) y luego por antigüedad descendente si existe
        if columna_para_ordenar:
            df_ordenado = df_priorizado.sort_values(
                ['_prioridad', columna_para_ordenar], 
                ascending=[False, False]
            )
            # Eliminar columna temporal si se creó
            if '_antiguedad_temp' in df_ordenado.columns:
                df_ordenado = df_ordenado.drop('_antiguedad_temp', axis=1)
        else:
            df_ordenado = df_priorizado.sort_values('_prioridad', ascending=False)
        
        # Eliminar columna temporal de prioridad
        if '_prioridad' in df_ordenado.columns:
            df_ordenado = df_ordenado.drop('_prioridad', axis=1)
        
        return df_ordenado
    
    except Exception as e:
        st.error(f"❌ Error al ordenar DataFrame: {e}")
        return df

@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def generar_pdf_usuario(usuario, df_pendientes, num_semana, fecha_max_str):
    """Genera el PDF para un usuario específico con nombre único - CORREGIDA PARA DECIMALES"""
    # Crear copia para no modificar el original
    df_user = df_pendientes[df_pendientes["USUARIO"] == usuario].copy(deep=True)
    
    if df_user.empty:
        return None
    
    # ORDENAR el DataFrame: RUE amarillos primero Y luego por antigüedad
    df_user_ordenado = ordenar_dataframe_por_prioridad_y_antiguedad(df_user)
    
    # Procesar datos para PDF - mantener las columnas originales para el formato condicional
    indices_a_incluir = list(range(df_user_ordenado.shape[1]))
    indices_a_excluir = {1, 4, 5, 6, 13}
    
    # EXCLUIR también la columna "FECHA DE ACTUALIZACIÓN DATOS" si existe
    for idx, col_name in enumerate(df_user_ordenado.columns):
        if "FECHA DE ACTUALIZACIÓN DATOS" in col_name.upper():
            indices_a_excluir.add(idx)
    
    indices_finales = [i for i in indices_a_incluir if i not in indices_a_excluir]
    NOMBRES_COLUMNAS_PDF = df_user_ordenado.columns[indices_finales].tolist()

    # 🔥 CORRECCIÓN: Identificar columna de antigüedad
    columnas_antiguedad = [col for col in NOMBRES_COLUMNAS_PDF if 'ANTIGÜEDAD' in col.upper() or 'DÍAS' in col.upper()]
    
    # Crear DataFrame para mostrar (SOLO para visualización)
    df_pdf_mostrar = df_user_ordenado[NOMBRES_COLUMNAS_PDF].copy()
    
    # Formatear para visualización - VECTORIZADO
    for col in df_pdf_mostrar.columns:
        if 'fecha' in col.lower():
            df_pdf_mostrar[col] = pd.to_datetime(df_pdf_mostrar[col], errors='coerce')                                     .dt.strftime("%d/%m/%Y").fillna("")
        elif df_pdf_mostrar[col].dtype in ['float64', 'float32']:
            df_pdf_mostrar[col] = df_pdf_mostrar[col].fillna(0).round(0).astype(int).astype(str)
        elif df_pdf_mostrar[col].dtype in ['int64', 'int32']:
            df_pdf_mostrar[col] = df_pdf_mostrar[col].fillna(0).astype(int).astype(str)
        else:
            s = df_pdf_mostrar[col].astype(str)
            df_pdf_mostrar[col] = s.where(~s.str.lower().isin(["nan", "nat", "none", "<na>"]), "")

    num_expedientes = len(df_pdf_mostrar)
    
    titulo_pdf = f"{usuario} - Semana {num_semana} a {fecha_max_str} - Expedientes Pendientes ({num_expedientes})"
    
    return dataframe_to_pdf_bytes(df_pdf_mostrar, titulo_pdf, df_original=df_user_ordenado)

# === FUNCIONES AUXILIARES ===

def obtener_hash_archivo(archivo):
    """Genera un hash único del archivo para detectar cambios"""
    if archivo is None:
        return None
    archivo.seek(0)
    file_hash = hashlib.md5(archivo.read()).hexdigest()
    archivo.seek(0)
    return file_hash

@st.cache_data(ttl=CACHE_TTL)
def generar_pdf_equipo_prioritarios(equipo, df_pendientes, num_semana, fecha_max_str):
    """Genera el PDF para un equipo específico solo con expedientes prioritarios - CORREGIDA PARA DECIMALES"""
    # Crear copia para no modificar el original
    df_equipo = df_pendientes[df_pendientes["EQUIPO"] == equipo].copy(deep=True)
    
    if df_equipo.empty:
        return None
    
    # CORRECCIÓN: Usar la función de identificación de prioritarios
    df_prioritarios = identificar_filas_prioritarias(df_equipo)
    df_prioritarios = df_prioritarios[df_prioritarios['_prioridad'] == 1].copy()
    
    if df_prioritarios.empty:
        return None
    
    # Eliminar columna temporal de prioridad
    df_prioritarios = df_prioritarios.drop('_prioridad', axis=1)
    
    # 🔥 CORRECCIÓN: ORDENAR POR PRIORIDAD Y ANTIGÜEDAD
    df_prioritarios = ordenar_dataframe_por_prioridad_y_antiguedad(df_prioritarios)
    
    # Procesar datos para PDF
    indices_a_incluir = list(range(df_prioritarios.shape[1]))
    indices_a_excluir = {1, 4, 5, 6, 13}
    
    # EXCLUIR también la columna "FECHA DE ACTUALIZACIÓN DATOS" si existe
    for idx, col_name in enumerate(df_prioritarios.columns):
        if "FECHA DE ACTUALIZACIÓN DATOS" in col_name.upper():
            indices_a_excluir.add(idx)
    
    indices_finales = [i for i in indices_a_incluir if i not in indices_a_excluir]
    NOMBRES_COLUMNAS_PDF = df_prioritarios.columns[indices_finales].tolist()

    # 🔥 CORRECCIÓN: Identificar columna de antigüedad
    columnas_antiguedad = [col for col in NOMBRES_COLUMNAS_PDF if 'ANTIGÜEDAD' in col.upper() or 'DÍAS' in col.upper()]
    
    # Crear DataFrame para mostrar (SOLO para visualización)
    df_pdf_mostrar = df_prioritarios[NOMBRES_COLUMNAS_PDF].copy()
    
    # Formatear para visualización - VECTORIZADO
    for col in df_pdf_mostrar.columns:
        if 'fecha' in col.lower():
            df_pdf_mostrar[col] = pd.to_datetime(df_pdf_mostrar[col], errors='coerce')                                     .dt.strftime("%d/%m/%Y").fillna("")
        elif df_pdf_mostrar[col].dtype in ['float64', 'float32']:
            df_pdf_mostrar[col] = df_pdf_mostrar[col].fillna(0).round(0).astype(int).astype(str)
        elif df_pdf_mostrar[col].dtype in ['int64', 'int32']:
            df_pdf_mostrar[col] = df_pdf_mostrar[col].fillna(0).astype(int).astype(str)
        else:
            s = df_pdf_mostrar[col].astype(str)
            df_pdf_mostrar[col] = s.where(~s.str.lower().isin(["nan", "nat", "none", "<na>"]), "")

    num_expedientes = len(df_pdf_mostrar)
    
    titulo_pdf = f"{equipo} - Semana {num_semana} a {fecha_max_str} - Expedientes Prioritarios ({num_expedientes})"
    
    return dataframe_to_pdf_bytes(df_pdf_mostrar, titulo_pdf, df_original=df_prioritarios)

# === FUNCIÓN OPTIMIZADA PARA GENERAR PDF RESUMEN KPI CON GRÁFICOS ===
@st.cache_data(ttl=CACHE_TTL_DYNAMIC)
def generar_pdf_resumen_kpi_optimizado(df_kpis_semanales, num_semana, fecha_max_str, df_combinado, semanas_disponibles, FECHA_REFERENCIA, fecha_max):
    """Versión optimizada que incluye gráficos reutilizando cálculos de la página 3"""
    
    try:
        # FILTRAR DATOS DE LA SEMANA ACTUAL
        kpis_semana = df_kpis_semanales[df_kpis_semanales['semana_numero'] == num_semana].iloc[0]
        
        pdf = PDFResumenKPI()
        
        # TÍTULO PRINCIPAL
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, f'RESUMEN DE KPIs - SEMANA {num_semana}', 0, 1, 'C')
        pdf.cell(0, 5, f'Periodo: {fecha_max_str}', 0, 1, 'C')
        pdf.ln(3)
        
        # SECCIÓN 1: KPIs PRINCIPALES (MANTENIENDO FORMATO ORIGINAL)
        pdf.add_section_title("KPIs PRINCIPALES")
        
        # Semanales
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, "Semanales:", 0, 1)
        pdf.add_metric("-Nuevos Expedientes", f"{int(kpis_semana['nuevos_expedientes']):,}".replace(",", "."))
        pdf.add_metric("-Expedientes Despachados", f"{int(kpis_semana['despachados_semana']):,}".replace(",", "."))
        pdf.add_metric("-Coef. Absorcion (Desp/Nuevos)", f"{kpis_semana['c_abs_despachados_sem']:.2f}%".replace(".", ","))
        pdf.add_metric("-Expedientes Cerrados", f"{int(kpis_semana['expedientes_cerrados']):,}".replace(",", "."))
        pdf.add_metric("-Coef. Absorcion (Cer/Asig)", f"{kpis_semana['c_abs_cerrados_sem']:.2f}%".replace(".", ","))
        pdf.add_metric("-Expedientes Abiertos", f"{int(kpis_semana['total_abiertos']):,}".replace(",", "."))
        pdf.add_metric("-Expedientes Abiertos no Despachados", f"{int(kpis_semana.get('total_abiertos_no_despachados', 0)):,}".replace(",", "."))
        pdf.add_metric("-Expedientes Rehabilitados", f"{int(kpis_semana['total_rehabilitados']):,}".replace(",", "."))
        
        # Totales
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(0, 8, "Totales (desde 01/11/2022):", 0, 1)
        pdf.add_metric("-Nuevos Expedientes", f"{int(kpis_semana['nuevos_expedientes_totales']):,}".replace(",", "."))
        pdf.add_metric("-Expedientes Despachados", f"{int(kpis_semana['despachados_totales']):,}".replace(",", "."))
        pdf.add_metric("-Coef. Absorcion (Desp/Nuevos)", f"{kpis_semana['c_abs_despachados_tot']:.2f}%".replace(".", ","))
        pdf.add_metric("-Expedientes Cerrados", f"{int(kpis_semana['expedientes_cerrados_totales']):,}".replace(",", "."))
        pdf.add_metric("-Coef. Absorcion (Cer/Asig)", f"{kpis_semana['c_abs_cerrados_tot']:.2f}%".replace(".", ","))
        pdf.add_metric("-Total Expedientes Asignados", f"{int(kpis_semana['despachados_totales'] + kpis_semana['total_abiertos_no_despachados']):,}".replace(",", "."))
        
        pdf.ln(3)
        
        # SECCIÓN 2: EXPEDIENTES ESPECIALES
        pdf.add_section_title("EXPEDIENTES CON 029, 033, PRE, RSL, PENDIENTE DE FIRMA, DECISION O COMPLETAR TRAMITE")
        pdf.add_metric("Expedientes Especiales", f"{int(kpis_semana['expedientes_especiales']):,}".replace(",", "."))
        pdf.add_metric("Porcentaje sobre Abiertos", f"{kpis_semana['porcentaje_especiales']:.2f}%".replace(".", ","))
        
        pdf.ln(3)
        
        # SECCIÓN 3: TIEMPOS DE TRAMITACION - DISEÑO EN DOS COLUMNAS
        pdf.add_section_title("TIEMPOS DE TRAMITACION (en dias)")
        
        # Guardar posición Y inicial para las columnas
        y_inicial = pdf.get_y()
        
        # PRIMERA FILA - TÍTULOS
        pdf.set_font('Arial', 'B', 10)
        
        # Columna izquierda - Expedientes Despachados
        pdf.cell(95, 7, "Expedientes Despachados:", 0, 0, 'L')
        
        # Columna derecha - Expedientes Cerrados
        pdf.set_x(105)  # Posicionar en la segunda columna
        pdf.cell(95, 7, "Expedientes Cerrados:", 0, 1, 'L')
        
        # PRIMERA FILA - DATOS
        # Posicionar para columna izquierda (Despachados)
        pdf.set_y(y_inicial + 7)
        pdf.set_x(10)
        pdf.set_font('Arial', '', 9)
        
        # Despachados - Tiempo Medio
        pdf.cell(85, 6, "-Tiempo Medio:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['tiempo_medio_despachados']:.0f} dias", 0, 1, 'R')
        
        # Despachados - Percentil 90
        pdf.set_x(10)
        pdf.cell(85, 6, "-Percentil 90:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_90_despachados']:.0f} dias", 0, 1, 'R')
        
        # Despachados - <=180 dias
        pdf.set_x(10)
        pdf.cell(85, 6, " <=180 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_180_despachados']:.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Despachados - <=120 dias
        pdf.set_x(10)
        pdf.cell(85, 6, " <=120 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_120_despachados']:.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Posicionar para columna derecha (Cerrados)
        pdf.set_y(y_inicial + 7)
        pdf.set_x(105)
        
        # Cerrados - Tiempo Medio
        pdf.cell(85, 6, "-Tiempo Medio:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['tiempo_medio_cerrados']:.0f} dias", 0, 1, 'R')
        
        # Cerrados - Percentil 90
        pdf.set_x(105)
        pdf.cell(85, 6, "-Percentil 90:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_90_cerrados']:.0f} dias", 0, 1, 'R')
        
        # Cerrados - <=180 dias
        pdf.set_x(105)
        pdf.cell(85, 6, " <=180 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_180_cerrados']:.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Cerrados - <=120 dias
        pdf.set_x(105)
        pdf.cell(85, 6, " <=120 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_120_cerrados']:.2f}%".replace(".", ","), 0, 1, 'R')
        
        # SEGUNDA FILA - TÍTULOS
        y_segunda_fila = pdf.get_y() + 3  # Espacio entre filas
        
        pdf.set_y(y_segunda_fila)
        pdf.set_font('Arial', 'B', 10)
        
        # Columna izquierda - Expedientes Abiertos
        pdf.cell(95, 7, "Expedientes Abiertos:", 0, 0, 'L')
        
        # Columna derecha - Expedientes Abiertos no Despachados
        pdf.set_x(105)
        pdf.cell(95, 7, "Expedientes Abiertos no Despachados:", 0, 1, 'L')
        
        # SEGUNDA FILA - DATOS
        # Posicionar para columna izquierda (Abiertos)
        pdf.set_y(y_segunda_fila + 7)
        pdf.set_x(10)
        pdf.set_font('Arial', '', 9)
        
        # Abiertos - Percentil 90
        pdf.cell(85, 6, "-Percentil 90:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_90_abiertos']:.0f} dias", 0, 1, 'R')
        
        # Abiertos - <=180 dias
        pdf.set_x(10)
        pdf.cell(85, 6, " <=180 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_180_abiertos']:.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Abiertos - <=120 dias
        pdf.set_x(10)
        pdf.cell(85, 6, " <=120 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana['percentil_120_abiertos']:.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Posicionar para columna derecha (Abiertos no Despachados)
        pdf.set_y(y_segunda_fila + 7)
        pdf.set_x(105)
        
        # Abiertos no Despachados - Percentil 90
        pdf.cell(85, 6, "-Percentil 90:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana.get('percentil_90_abiertos_no_despachados', 0):.0f} dias", 0, 1, 'R')
        
        # Abiertos no Despachados - <=180 dias
        pdf.set_x(105)
        pdf.cell(85, 6, " <=180 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana.get('percentil_180_abiertos_no_despachados', 0):.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Abiertos no Despachados - <=120 dias
        pdf.set_x(105)
        pdf.cell(85, 6, " <=120 dias:", 0, 0, 'L')
        pdf.cell(10, 6, f"{kpis_semana.get('percentil_120_abiertos_no_despachados', 0):.2f}%".replace(".", ","), 0, 1, 'R')
        
        # Información del período
        pdf.ln(8)
        pdf.set_font('Arial', 'I', 8)
        if kpis_semana['es_semana_actual']:
            periodo_texto = f"Periodo de la semana (ACTUAL): {kpis_semana['inicio_semana'].strftime('%d/%m/%Y')} a {kpis_semana['fin_semana'].strftime('%d/%m/%Y')} - {kpis_semana['dias_semana']} dias"
        else:
            periodo_texto = f"Periodo de la semana: {kpis_semana['inicio_semana'].strftime('%d/%m/%Y')} a {kpis_semana['fin_semana'].strftime('%d/%m/%Y')} - {kpis_semana['dias_semana']} dias"
        
        pdf.cell(0, 5, periodo_texto, 0, 1)

        # ===== SECCIÓN DE GRÁFICOS OPTIMIZADA =====
        pdf.add_page()
        pdf.add_section_title("GRAFICOS DE EVOLUCION - SEMANA " + str(num_semana))

        try:
            # VERIFICACIÓN RÁPIDA DE KALEIDO
            import plotly.io as pio
            if not hasattr(pio, 'kaleido') or pio.kaleido.scope is None:
                st.warning("Kaleido no disponible para gráficos")
                raise ImportError("Kaleido no disponible")
            
            # USAR DATOS YA CALCULADOS - SIN RECÁLCULOS
            datos_grafico = df_kpis_semanales.copy()
            
            # CONFIGURACIÓN DE GRÁFICOS MÁS RÁPIDA
            config_rapida = {
                'displayModeBar': False,
                'staticPlot': True,  # Gráficos estáticos más rápidos
            }
            
            # GRÁFICO 1: Evolución de expedientes principales
            fig1 = px.line(
                datos_grafico,
                x='semana_numero',
                y=['nuevos_expedientes', 'despachados_semana', 'expedientes_cerrados'],
                title=f'Evolucion de Expedientes - Semana {num_semana}',
                labels={'semana_numero': 'Semana', 'value': 'Cantidad', 'variable': 'KPI'},
                color_discrete_map={
                    'nuevos_expedientes': '#1f77b4',
                    'despachados_semana': '#ff7f0e', 
                    'expedientes_cerrados': '#2ca02c'
                }
            )
            # CONFIGURACIÓN MÍNIMA PARA RAPIDEZ
            fig1.update_layout(
                height=300,  # Menor altura = más rápido
                showlegend=True,
                margin=dict(l=40, r=20, t=40, b=20)
            )
            fig1.add_vline(x=num_semana, line_dash="dash", line_color="red")
            
            # GUARDAR CON CONFIGURACIÓN RÁPIDA
            temp_chart1 = user_env.get_temp_path(f"chart1_{num_semana}.png")
            fig1.write_image(temp_chart1, engine="kaleido", scale=1, width=700, height=300)
            
            # INSERTAR EN PDF
            pdf.ln(3)
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "Evolucion de Expedientes (Nuevos, Despachados, Cerrados)", 0, 1)
            pdf.image(temp_chart1, x=10, w=190)
            
            pdf.ln(3)
            
            # GRÁFICO 2: Expedientes Abiertos
            fig2 = px.line(
                datos_grafico,
                x='semana_numero', 
                y=['total_abiertos'],
                title=f'Expedientes Abiertos - Semana {num_semana}',
                labels={'semana_numero': 'Semana', 'value': 'Cantidad'},
                color_discrete_sequence=['#d62728']
            )
            fig2.update_layout(height=300, showlegend=False, margin=dict(l=40, r=20, t=40, b=20))
            fig2.add_vline(x=num_semana, line_dash="dash", line_color="red")
            
            temp_chart2 = user_env.get_temp_path(f"chart2_{num_semana}.png")
            fig2.write_image(temp_chart2, engine="kaleido", scale=1, width=700, height=300)
            
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "Evolucion de Expedientes Abiertos", 0, 1)
            pdf.image(temp_chart2, x=10, w=190)
            
            pdf.ln(3)
            
            # GRÁFICO 3: Coeficientes de absorción
            fig3 = px.line(
                datos_grafico,
                x='semana_numero',
                y=[
                    'c_abs_despachados_sem', 'c_abs_despachados_tot',
                    'c_abs_cerrados_sem', 'c_abs_cerrados_tot'
                ],
                title='Evolución de Coeficientes de Absorción (%)',
                labels={'semana_numero': 'Semana', 'value': 'Porcentaje (%)', 'variable': 'Indicador'},
                color_discrete_map={
                    'c_abs_despachados_sem': '#9467bd',
                    'c_abs_despachados_tot': '#c5b0d5',
                    'c_abs_cerrados_sem': '#8c564b',
                    'c_abs_cerrados_tot': '#c49c94'
                }
            )
            fig3.update_layout(height=300, showlegend=True, margin=dict(l=40, r=20, t=40, b=20))
            fig3.add_vline(x=num_semana, line_dash="dash", line_color="red")
            
            temp_chart3 = user_env.get_temp_path(f"chart3_{num_semana}.png")
            fig3.write_image(temp_chart3, engine="kaleido", scale=1, width=700, height=300)
            
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "Coeficientes de Absorcion Semanales (%)", 0, 1)
            pdf.image(temp_chart3, x=10, w=190)
            
            pdf.ln(3)
            
            # GRÁFICO 4: Tiempos de tramitación
            fig4 = px.line(
                datos_grafico,
                x='semana_numero',
                y=[
                    'tiempo_medio_despachados', 'tiempo_medio_cerrados',
                    'percentil_90_despachados', 'percentil_90_cerrados'
                ],
                title='Tiempos Medios y Percentiles 90 (días)',
                labels={'semana_numero': 'Semana', 'value': 'Días', 'variable': 'Indicador'},
                color_discrete_map={
                    'tiempo_medio_despachados': '#ff7f0e',
                    'tiempo_medio_cerrados': '#2ca02c',
                    'percentil_90_despachados': '#ffbb78',
                    'percentil_90_cerrados': '#98df8a'
                }
            )
            fig4.update_layout(height=300, showlegend=True, margin=dict(l=40, r=20, t=40, b=20))
            fig4.add_vline(x=num_semana, line_dash="dash", line_color="red")
            
            temp_chart4 = user_env.get_temp_path(f"chart4_{num_semana}.png")
            fig4.write_image(temp_chart4, engine="kaleido", scale=1, width=700, height=300)
            
            pdf.add_page()
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "Tiempos de Tramitacion (Medios)", 0, 1)
            pdf.image(temp_chart4, x=10, w=190)
            
            pdf.ln(3)
            
            # GRÁFICO 5: Porcentajes 120/180 días
            fig5 = px.line(
                datos_grafico,
                x='semana_numero',
                y=[
                    'percentil_180_despachados', 'percentil_120_despachados',
                    'percentil_180_cerrados', 'percentil_120_cerrados'
                ],
                title='Porcentaje de Expedientes ≤120 y ≤180 días (%)',
                labels={'semana_numero': 'Semana', 'value': 'Porcentaje (%)', 'variable': 'Indicador'},
                color_discrete_map={
                    'percentil_180_despachados': '#ff7f0e',
                    'percentil_120_despachados': '#ffddaa',
                    'percentil_180_cerrados': '#2ca02c',
                    'percentil_120_cerrados': '#98df8a'
                }
            )
            fig5.update_layout(height=300, showlegend=True, margin=dict(l=40, r=20, t=40, b=20))
            fig5.add_vline(x=num_semana, line_dash="dash", line_color="red")
            
            temp_chart5 = user_env.get_temp_path(f"chart5_{num_semana}.png")
            fig5.write_image(temp_chart5, engine="kaleido", scale=1, width=700, height=300)
            
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 8, "Porcentaje de Expedientes Despachados dentro de Plazos (120/180 dias)", 0, 1)
            pdf.image(temp_chart5, x=10, w=190)
            
            # LIMPIAR TEMPORALES INMEDIATAMENTE
            for temp_file in [temp_chart1, temp_chart2, temp_chart3, temp_chart4, temp_chart5]:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass
                
        except Exception as chart_error:
            # FALLBACK: TABLA DE DATOS SIN GRÁFICOS
            pdf.ln(3)
            pdf.set_font('Arial', 'I', 8)
            pdf.cell(0, 5, f"Nota: No se pudieron incluir los graficos. Error: {str(chart_error)}", 0, 1)
            
            # TABLA ALTERNATIVA CON DATOS
            pdf.ln(3)
            pdf.set_font('Arial', 'B', 9)
            pdf.cell(0, 6, "Datos de evolucion (ultimas 8 semanas):", 0, 1)
            
            datos_tabla = datos_grafico.tail(8)[[
                'semana_numero', 'nuevos_expedientes', 'despachados_semana', 
                'expedientes_cerrados', 'total_abiertos'
            ]]
            pdf.set_font('Arial', '', 6)
            
            # ENCABEZADOS
            headers = ["Sem", "Nuevos", "Despach", "Cerrad", "Abiert"]
            widths = [15, 20, 20, 20, 20]
            
            for i, header in enumerate(headers):
                pdf.cell(widths[i], 5, header, 1)
            pdf.ln()
            
            # DATOS
            for _, row in datos_tabla.iterrows():
                pdf.cell(widths[0], 5, str(int(row['semana_numero'])), 1)
                pdf.cell(widths[1], 5, str(int(row['nuevos_expedientes'])), 1)
                pdf.cell(widths[2], 5, str(int(row['despachados_semana'])), 1)
                pdf.cell(widths[3], 5, str(int(row['expedientes_cerrados'])), 1)
                pdf.cell(widths[4], 5, str(int(row['total_abiertos'])), 1)
                pdf.ln()
        
        # EXPORTAR A BYTES
        pdf_output = pdf.output(dest='S')
        
        if isinstance(pdf_output, str):
            pdf_bytes = pdf_output.encode('latin1')
        elif isinstance(pdf_output, (bytes, bytearray)):
            pdf_bytes = bytes(pdf_output)
        else:
            pdf_bytes = b''
        
        return pdf_bytes

    except Exception as e:
        # FALLBACK COMPLETO EN CASO DE ERROR
        try:
            pdf = PDFResumenKPI()
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, f'RESUMEN KPIs - SEMANA {num_semana}', 0, 1, 'C')
            pdf.cell(0, 5, f'Periodo: {fecha_max_str}', 0, 1, 'C')
            pdf.ln(5)
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 8, f"Expedientes totales: {len(df_combinado)}", 0, 1)
            pdf.cell(0, 8, f"Semana calculada: {num_semana}", 0, 1)
            
            pdf_output = pdf.output(dest='S')
            if isinstance(pdf_output, str):
                return pdf_output.encode('latin1')
            elif isinstance(pdf_output, (bytes, bytearray)):
                return bytes(pdf_output)
        except:
            pass
        
        return None

# === FUNCIÓN OPTIMIZADA PARA KPIs DE TODAS LAS SEMANAS ===
@st.cache_data(ttl=CACHE_TTL, show_spinner="📊 Calculando KPIs históricos...")
def calcular_kpis_todas_semanas_optimizado(_df, _semanas, _fecha_referencia, _fecha_max, _user_key=user_env.session_id):
    """Versión optimizada con procesamiento por lotes"""
    
    datos_semanales = []
    
    for i, semana in enumerate(_semanas):
        es_semana_actual = (i == len(_semanas) - 1)
        
        kpis = calcular_kpis_para_semana_optimizado(_df, semana, es_semana_actual)
        num_semana = ((semana - _fecha_referencia).days) // 7 + 1
        
        datos_semanales.append({
            'semana_numero': num_semana,
            'semana_fin': semana,
            'semana_str': semana.strftime('%d/%m/%Y'),
            **kpis  # Desempaquetar todos los KPIs
        })
    
    return pd.DataFrame(datos_semanales)

def obtener_saludo():
    """Devuelve saludo según la hora actual"""
    hora_actual = datetime.now().hour
    
    if hora_actual < 12:
        return "Buenos días"
    elif hora_actual < 20:
        return "Buenas tardes"
    else:
        return "Buenas noches"

def enviar_correo_outlook(destinatario, asunto, cuerpo_mensaje, archivos_adjuntos, cc=None, bcc=None):
    """
    Envía correo usando MAPI - versión específicamente corregida para el resumen
    """
    try:
        import win32com.client
        import pythoncom
        import os
        import tempfile
        
        # Forzar inicialización COM
        pythoncom.CoInitialize()
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            
            mail.To = destinatario
            mail.Subject = asunto
            mail.Body = cuerpo_mensaje

            if cc and pd.notna(cc) and str(cc).strip():
                mail.CC = str(cc)
            if bcc and pd.notna(bcc) and str(bcc).strip():
                mail.BCC = str(bcc)

            # Adjuntar archivos - SOLUCIÓN ESPECÍFICA PARA RESUMEN
            for nombre_archivo, datos_archivo in archivos_adjuntos:
                try:
                    # Crear archivo temporal con el nombre EXACTO que queremos
                    temp_path = os.path.join(user_env.working_dir, nombre_archivo)
                    
                    # Escribir el archivo con el nombre correcto
                    with open(temp_path, 'wb') as temp_file:
                        temp_file.write(datos_archivo)
                    
                    # Adjuntar el archivo - método simple y directo
                    mail.Attachments.Add(temp_path)
                    
                    # 🔥 SOLUCIÓN ESPECÍFICA: Forzar el cierre del archivo antes de enviar
                    del temp_file  # Liberar el handle del archivo
                    
                except Exception as attach_error:
                    st.error(f"Error adjuntando archivo {nombre_archivo}: {attach_error}")
                    # Continuar con otros archivos

            # Enviar correo
            mail.Send()
            
            # Limpiar archivos temporales después del envío
            for nombre_archivo, datos_archivo in archivos_adjuntos:
                try:
                    temp_path = os.path.join(user_env.working_dir, nombre_archivo)
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)
                except:
                    pass  # Ignorar errores de limpieza
            
            return True
            
        except Exception as e:
            st.error(f"❌ Error Outlook al enviar a {destinatario}: {e}")
            return False
        finally:
            # Liberar objetos COM
            try:
                if 'mail' in locals():
                    del mail
                if 'outlook' in locals():
                    del outlook
            except:
                pass
            
    except Exception as e:
        st.error(f"❌ Error general al enviar a {destinatario}: {e}")
        return False
    finally:
        # Siempre desinicializar COM
        try:
            pythoncom.CoUninitialize()
        except:
            pass

# Función para gráficos dinámicos (SIN CACHE)
def crear_grafico_dinamico(_conteo, columna, titulo):
    """Crea gráficos dinámicos que responden a los filtros"""
    if _conteo.empty:
        return None
    
    fig = px.bar(_conteo, y=columna, x="Cantidad", title=titulo, text="Cantidad", 
                 color=columna, height=400)
    fig.update_traces(texttemplate='%{text:,}', textposition="auto")
    return fig

# === NUEVA CLASE PDF PARA RENDIMIENTO ===
class PDFRendimiento(FPDF):
    def __init__(self):
        super().__init__()
        # Agregar soporte para caracteres latinos
        self.add_page()
        
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'RECTAUTO - Informe de Rendimiento por Usuario', 0, 1, 'C')
        self.ln(5)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Pagina {self.page_no()}', 0, 0, 'C')
    
    def add_section_title(self, title):
        self.set_font('Arial', 'B', 10)
        # Reemplazar caracteres problemáticos
        title_safe = title.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
        self.cell(0, 8, title_safe, 0, 1, 'L')
        self.ln(2)
    
    def add_metric(self, label, value, explanation=""):
        # Limpiar caracteres especiales de las etiquetas
        label_safe = label.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
        
        self.set_font('Arial', 'B', 9)
        self.cell(60, 6, label_safe, 0, 0)
        self.set_font('Arial', '', 9)
        
        # Limpiar también el valor si es texto
        if isinstance(value, str):
            value_safe = value.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
        else:
            value_safe = value
            
        self.cell(40, 6, str(value_safe), 0, 1)
        if explanation:
            self.set_font('Arial', 'I', 8)
            explanation_safe = explanation.replace('Í', 'I').replace('É', 'E').replace('Á', 'A').replace('Ó', 'O').replace('Ú', 'U')
            self.cell(0, 4, explanation_safe, 0, 1)
            self.ln(1)

# === FUNCIÓN PARA GENERAR PDF DE RENDIMIENTO ===
@st.cache_data(ttl=CACHE_TTL_DYNAMIC)
def generar_pdf_rendimiento(df_rendimiento_completo, num_semana, fecha_max_str):
    """Genera un PDF con la tabla de rendimiento por usuario"""
    
    def fmt_es(valor, decimales=2):
        """Formato numerico espanol: 1.234,56"""
        try:
            if pd.isna(valor):
                return "0" if decimales == 0 else "0," + "0" * decimales
            v = float(valor)
            if decimales == 0:
                return f"{int(round(v)):,}".replace(",", ".")
            fmt = f"{v:,.{decimales}f}"
            izq, der = fmt.split(".")
            izq = izq.replace(",", ".")
            return izq + "," + der
        except Exception:
            return str(valor)

    try:
        # Verificar que haya datos
        if df_rendimiento_completo.empty:
            return None
        
        # Asegurar que solo tenemos usuarios ACTIVOS (doble verificación)
        df_rendimiento = df_rendimiento_completo.copy()
        if 'ESTADO' in df_rendimiento.columns:
            df_rendimiento = df_rendimiento[df_rendimiento['ESTADO'] == 'ACTIVO']
        
        if df_rendimiento.empty:
            return None
        
        pdf = PDFRendimiento()
        
        # TÍTULO PRINCIPAL
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, f'INFORME DE RENDIMIENTO - SEMANA {num_semana}', 0, 1, 'C')
        pdf.cell(0, 5, f'Periodo: {fecha_max_str}', 0, 1, 'C')
        pdf.ln(3)
        
        # SECCIÓN 1: RESUMEN GENERAL
        pdf.add_section_title("RESUMEN GENERAL")
        
        # Calcular totales
        total_usuarios = df_rendimiento_completo['USUARIO'].nunique()
        total_expedientes = df_rendimiento_completo['EXPEDIENTES_DESPACHADOS'].sum()
        total_semanas = df_rendimiento_completo['SEMANAS_EFECTIVAS'].sum()
        rendimiento_promedio = total_expedientes/total_semanas
        rendimiento_anual_promedio = df_rendimiento_completo['RENDIMIENTO_ANUAL'].mean()
        potencial_anual_promedio = df_rendimiento_completo['POTENCIAL_ANUAL'].mean()
        potencial_anual_conjunto = potencial_anual_promedio*total_usuarios
        
        # Obtener lista de equipos únicos
        todos_equipos = set()
        for equipos_str in df_rendimiento_completo['EQUIPOS'].dropna():
            equipos_lista = [eq.strip() for eq in str(equipos_str).split(',')]
            todos_equipos.update(equipos_lista)
        
        pdf.add_metric("Total de Usuarios",                fmt_es(total_usuarios, 0))
        pdf.add_metric("Total de Expedientes Despachados",  fmt_es(total_expedientes, 0))
        pdf.add_metric("Total de Semanas Efectivas",        fmt_es(total_semanas, 1))
        pdf.add_metric("Rendimiento Promedio",              fmt_es(rendimiento_promedio, 2))
        pdf.add_metric("Rendimiento Anual Promedio",        fmt_es(rendimiento_anual_promedio, 2))
        pdf.add_metric("Potencial Anual Promedio",          fmt_es(potencial_anual_promedio, 0))
        pdf.add_metric("Potencial Anual Conjunto",          fmt_es(potencial_anual_conjunto, 0))
        pdf.add_metric("Equipos Analizados",                len(todos_equipos))
        
        pdf.ln(5)
        
        # SECCIÓN 2: DISTRIBUCIÓN POR ESTADO
        pdf.add_section_title("DISTRIBUCIÓN POR ESTADO")
        
        conteo_estado = df_rendimiento_completo['ESTADO'].value_counts()
        for estado, cantidad in conteo_estado.items():
            pdf.add_metric(f"- {estado}", cantidad)
        
        pdf.ln(5)
        
        # SECCIÓN 3: TABLA COMPLETA DE USUARIOS
        pdf.add_section_title("TABLA COMPLETA DE RENDIMIENTO POR USUARIO")
        
        # Definir anchos de columna para la tabla
        column_widths = [15, 45, 15, 15, 15, 15, 15, 15, 15, 15, 15]
        headers = ['USUARIO', 'EQUIPOS', 'ESTADO', 'EXP.DESP.', 'SEM.EFEC.', 
                'REND.TOT.', 'REND.ANUAL', 'POT.ANUAL', 'REND.TRI.', 
                'REND.MES', 'REND.SEM.']
        
        # Configurar fuente para tabla
        pdf.set_font('Arial', 'B', 6)
        
        # Imprimir encabezados
        for i, header in enumerate(headers):
            pdf.cell(column_widths[i], 8, header, 1, 0, 'C')
        pdf.ln()
        
        # Imprimir datos
        pdf.set_font('Arial', '', 6)
        ALTURA_LINEA_REND = 6
        ALTURA_FILA_REND  = 12

        for _, row in df_rendimiento_completo.iterrows():
            usuario   = str(row['USUARIO'])[:6]
            equipos   = str(row['EQUIPOS']) if pd.notna(row['EQUIPOS']) else ""
            estado    = str(row['ESTADO'])
            exp_desp  = fmt_es(row['EXPEDIENTES_DESPACHADOS'], 0)
            sem_efec  = fmt_es(row['SEMANAS_EFECTIVAS'],       1)
            rend_tot  = fmt_es(row['RENDIMIENTO_TOTAL'],       2)
            rend_anual= fmt_es(row['RENDIMIENTO_ANUAL'],       2)
            pot_anual = fmt_es(row['POTENCIAL_ANUAL'],         0)
            rend_tri  = fmt_es(row['RENDIMIENTO_TRIMESTRAL'],  2)
            rend_mes  = fmt_es(row['RENDIMIENTO_MENSUAL'],     2)
            rend_sem  = fmt_es(row['RENDIMIENTO_SEMANAL'],     2)

            datos_fila = [usuario, equipos, estado, exp_desp, sem_efec,
                          rend_tot, rend_anual, pot_anual, rend_tri,
                          rend_mes, rend_sem]

            if pdf.get_y() + ALTURA_FILA_REND > pdf.h - pdf.b_margin:
                pdf.add_page()
                pdf.set_font('Arial', 'B', 6)
                for i, header in enumerate(headers):
                    pdf.cell(column_widths[i], 8, header, 1, 0, 'C')
                pdf.ln()
                pdf.set_font('Arial', '', 6)

            x_inicio = pdf.get_x()
            y_inicio = pdf.get_y()

            for i, dato in enumerate(datos_fila):
                x_celda = x_inicio + sum(column_widths[:i])
                pdf.set_xy(x_celda, y_inicio)
                if i == 1:  # EQUIPOS: borde fijo 2 lineas, texto centrado
                    pdf.rect(x_celda, y_inicio, column_widths[i], ALTURA_FILA_REND)
                    ancho_texto = pdf.get_string_width(dato)
                    n_lineas = min(2, math.ceil(ancho_texto / (column_widths[i] - 2)) if ancho_texto > 0 else 1)
                    y_texto = y_inicio + (ALTURA_FILA_REND - n_lineas * ALTURA_LINEA_REND) / 2
                    pdf.set_xy(x_celda, y_texto)
                    pdf.multi_cell(column_widths[i], ALTURA_LINEA_REND, dato, 0, 'C')
                else:
                    pdf.cell(column_widths[i], ALTURA_FILA_REND, dato, 1, 0, 'C')

            pdf.set_xy(pdf.l_margin, y_inicio + ALTURA_FILA_REND)
        
        pdf.ln(5)
        
        # SECCIÓN 4: TOP 10 USUARIOS
        pdf.add_section_title("TOP 10 USUARIOS POR RENDIMIENTO TOTAL")
        
        # Ordenar por rendimiento total
        df_top10 = df_rendimiento_completo.sort_values('RENDIMIENTO_TOTAL', ascending=False).head(10)
        
        pdf.set_font('Arial', 'B', 6)
        pdf.cell(30, 6, 'USUARIO', 1, 0, 'C')
        pdf.cell(20, 6, 'REND.TOTAL', 1, 0, 'C')
        pdf.cell(20, 6, 'REND.ANUAL', 1, 0, 'C')
        pdf.cell(20, 6, 'POT.ANUAL', 1, 0, 'C')
        pdf.cell(25, 6, 'EXPEDIENTES', 1, 0, 'C')
        pdf.ln()
        
        pdf.set_font('Arial', '', 6)
        for _, row in df_top10.iterrows():
            usuario = str(row['USUARIO'])[:25]
            pdf.cell(30, 6, usuario,                                  1, 0, 'L')
            pdf.cell(20, 6, fmt_es(row['RENDIMIENTO_TOTAL'],       2), 1, 0, 'C')
            pdf.cell(20, 6, fmt_es(row['RENDIMIENTO_ANUAL'],       2), 1, 0, 'C')
            pdf.cell(20, 6, fmt_es(row['POTENCIAL_ANUAL'],         0), 1, 0, 'C')
            pdf.cell(25, 6, fmt_es(row['EXPEDIENTES_DESPACHADOS'], 0), 1, 0, 'C')
            pdf.ln()
        
        # EXPORTAR A BYTES
        pdf_output = pdf.output(dest='S')
        
        if isinstance(pdf_output, str):
            pdf_bytes = pdf_output.encode('latin1')
        elif isinstance(pdf_output, (bytes, bytearray)):
            pdf_bytes = bytes(pdf_output)
        else:
            pdf_bytes = b''
        
        return pdf_bytes

    except Exception as e:
        st.error(f"❌ Error generando PDF de rendimiento: {e}")
        return None

@st.cache_data(ttl=CACHE_TTL)
def calcular_rendimiento_usuarios_agrupado(_df, _df_usuarios, _fecha_max):
    """Calcula rendimiento AGRUPADO POR USUARIO (sin duplicar por equipos)"""
    
    # 1. IDENTIFICAR EXPEDIENTES DESPACHADOS
    fecha_9999 = pd.to_datetime('9999-09-09', errors='coerce')
    fecha_inicio_totales = datetime(2022, 11, 1)
    
    # Expedientes con FECHA RESOLUCIÓN real (distinta de 9999 y no nula)
    mask_despachados_reales = (
        _df['FECHA RESOLUCIÓN'].notna() & 
        (_df['FECHA RESOLUCIÓN'] != fecha_9999) &
        (_df['FECHA RESOLUCIÓN'] >= fecha_inicio_totales) &
        (_df['FECHA RESOLUCIÓN'] <= _fecha_max)
    )
    
    # Expedientes CERRADOS con FECHA RESOLUCIÓN = 9999-09-09 o vacía
    mask_despachados_cerrados = (
        (_df['ESTADO'] == 'Cerrado') &
        (_df['FECHA RESOLUCIÓN'].isna() | (_df['FECHA RESOLUCIÓN'] == fecha_9999)) &
        _df['FECHA CIERRE'].notna() &
        (_df['FECHA CIERRE'] >= fecha_inicio_totales) &
        (_df['FECHA CIERRE'] <= _fecha_max)
    )
    
    mask_despachados = mask_despachados_reales | mask_despachados_cerrados
    df_despachados = _df[mask_despachados].copy()
    
    # 2. CALCULAR DESPACHADOS POR USUARIO (agrupando todos los equipos)
    if 'USUARIO' not in df_despachados.columns:
        st.error("❌ No se encuentra la columna USUARIO en los datos")
        return pd.DataFrame()
        
    despachados_por_usuario = df_despachados.groupby('USUARIO').size().reset_index(name='EXPEDIENTES_DESPACHADOS')
    
    # 3. OBTENER EQUIPOS POR USUARIO (para mostrar en la tabla)
    equipos_por_usuario = df_despachados.groupby('USUARIO')['EQUIPO'].apply(
        lambda x: ', '.join(sorted(set(x.dropna().astype(str))))
    ).reset_index(name='EQUIPOS')
    
    # 4. PREPARAR DATOS DE USUARIOS
    usuarios_data = []
    
    # Verificar columnas en el archivo de usuarios
    columnas_usuarios = _df_usuarios.columns.tolist()
    st.info(f"📋 Columnas en archivo USUARIOS: {', '.join(columnas_usuarios)}")
    
    # Buscar nombres alternativos para las columnas
    columna_usuario = None
    columna_fecha_inicio = None
    columna_fecha_fin = None
    columna_semanas_baja = None
    
    # Mapeo de posibles nombres de columnas
    mapeo_columnas = {
        'usuario': ['USUARIOS', 'USUARIO', 'NOMBRE', 'NOMBRE USUARIO'],
        'fecha_inicio': ['FECHA INICIO', 'INICIO', 'FECHA_ALTA', 'ALTA'],
        'fecha_fin': ['FECHA FIN', 'FIN', 'FECHA_BAJA', 'BAJA', 'FECHA SALIDA'],
        'semanas_baja': ['SEMANAS DE BAJA', 'SEMANAS_BAJA', 'BAJAS', 'DIAS BAJA']
    }
    
    for col_tipo, posibles_nombres in mapeo_columnas.items():
        for nombre in posibles_nombres:
            if nombre in _df_usuarios.columns:
                if col_tipo == 'usuario':
                    columna_usuario = nombre
                elif col_tipo == 'fecha_inicio':
                    columna_fecha_inicio = nombre
                elif col_tipo == 'fecha_fin':
                    columna_fecha_fin = nombre
                elif col_tipo == 'semanas_baja':
                    columna_semanas_baja = nombre
                break
    
    if not columna_usuario:
        st.error("❌ No se encuentra la columna de usuarios en el archivo USUARIOS")
        st.info("💡 Las columnas disponibles son: " + ", ".join(columnas_usuarios))
        return pd.DataFrame()
    
    st.success(f"✅ Columna de usuario identificada: {columna_usuario}")
    
    for _, usuario_row in _df_usuarios.iterrows():
        usuario_nombre = usuario_row[columna_usuario]
        
        # Obtener fechas con nombres alternativos
        fecha_inicio = usuario_row.get(columna_fecha_inicio, None) if columna_fecha_inicio else None
        fecha_fin = usuario_row.get(columna_fecha_fin, None) if columna_fecha_fin else None
        semanas_baja = usuario_row.get(columna_semanas_baja, 0) if columna_semanas_baja else 0
        
        # Determinar estado
        if pd.isna(fecha_fin) or str(fecha_fin).strip() == '':
            estado = "ACTIVO"
        else:
            try:
                fecha_fin_dt = pd.to_datetime(fecha_fin, errors='coerce')
                if pd.isna(fecha_fin_dt) or fecha_fin_dt > _fecha_max:
                    estado = "ACTIVO"
                else:
                    estado = "INACTIVO"
            except:
                estado = "ACTIVO"
        
        usuarios_data.append({
            'USUARIO': usuario_nombre,
            'FECHA_INICIO': fecha_inicio,
            'FECHA_FIN': fecha_fin,
            'SEMANAS_BAJA': semanas_baja,
            'ESTADO': estado
        })
    
    df_usuarios_info = pd.DataFrame(usuarios_data)
    
    # 5. COMBINAR DATOS Y CALCULAR INDICADORES POR USUARIO
    resultados = []
    
    for _, row in despachados_por_usuario.iterrows():
        usuario = row['USUARIO']
        expedientes_despachados = row['EXPEDIENTES_DESPACHADOS']
        
        # Obtener equipos del usuario (para mostrar)
        equipos_usuario = equipos_por_usuario[equipos_por_usuario['USUARIO'] == usuario]
        equipos_str = equipos_usuario['EQUIPOS'].iloc[0] if not equipos_usuario.empty else "Sin equipo"
        
        # Buscar información del usuario
        usuario_info = None
        if not df_usuarios_info.empty:
            usuario_match = df_usuarios_info[df_usuarios_info['USUARIO'] == usuario]
            if not usuario_match.empty:
                usuario_info = usuario_match.iloc[0]
        
        # CALCULAR SEMANAS EFECTIVAS (UNA SOLA VEZ POR USUARIO)
        semanas_efectivas = 1
        estado = "INACTIVO"  # Por defecto si no se encuentra en usuarios
        
        if usuario_info is not None:
            # Usuario encontrado en archivo USUARIOS
            fecha_inicio = pd.to_datetime(usuario_info['FECHA_INICIO'], errors='coerce')
            fecha_fin = pd.to_datetime(usuario_info['FECHA_FIN'], errors='coerce')
            semanas_baja = float(usuario_info['SEMANAS_BAJA']) if pd.notna(usuario_info['SEMANAS_BAJA']) and str(usuario_info['SEMANAS_BAJA']).strip() != '' else 0
            estado = usuario_info['ESTADO']
            
            # Calcular fecha fin efectiva
            fecha_fin_efectiva = fecha_fin if pd.notna(fecha_fin) and fecha_fin <= _fecha_max else _fecha_max
            
            # Calcular semanas efectivas de trabajo
            fecha_inicio_efectiva = max(fecha_inicio, fecha_inicio_totales) if pd.notna(fecha_inicio) else fecha_inicio_totales
            
            if pd.notna(fecha_inicio_efectiva):
                dias_totales = (fecha_fin_efectiva - fecha_inicio_efectiva).days
                semanas_totales = max(dias_totales / 7, 0)
                semanas_efectivas = max(semanas_totales - semanas_baja, 0)
        
        # CALCULAR RENDIMIENTOS POR PERÍODOS
        
        # Definir períodos
        fecha_inicio_anio = _fecha_max - timedelta(days=365)
        fecha_inicio_trimestre = _fecha_max - timedelta(days=90)
        fecha_inicio_mes = _fecha_max - timedelta(days=30)
        fecha_inicio_semana = _fecha_max - timedelta(days=7)
        
        # Ajustar fechas de inicio según fecha_inicio del usuario
        if usuario_info is not None and pd.notna(usuario_info['FECHA_INICIO']):
            fecha_inicio_usuario = pd.to_datetime(usuario_info['FECHA_INICIO'])
            fecha_inicio_anio = max(fecha_inicio_anio, fecha_inicio_usuario)
            fecha_inicio_trimestre = max(fecha_inicio_trimestre, fecha_inicio_usuario)
            fecha_inicio_mes = max(fecha_inicio_mes, fecha_inicio_usuario)
            fecha_inicio_semana = max(fecha_inicio_semana, fecha_inicio_usuario)
        
        # Último año
        despachados_ultimo_anio = len(df_despachados[
            (df_despachados['USUARIO'] == usuario) & 
            (estado == 'ACTIVO') &
            (df_despachados['FECHA RESOLUCIÓN'] >= fecha_inicio_anio) &
            (df_despachados['FECHA RESOLUCIÓN'] <= _fecha_max)
        ])
        semanas_anio = min(52, ((_fecha_max - fecha_inicio_anio).days / 7)) if fecha_inicio_anio < _fecha_max else 0
        rendimiento_anual = despachados_ultimo_anio / semanas_anio if semanas_anio > 0 else 0
        
        # NUEVO: POTENCIAL ANUAL (Rendimiento anual * 52 semanas)
        potencial_anual = rendimiento_anual * 52
        
        # Últimos tres meses
        despachados_trimestre = len(df_despachados[
            (df_despachados['USUARIO'] == usuario) & 
            (estado == 'ACTIVO') &
            (df_despachados['FECHA RESOLUCIÓN'] >= fecha_inicio_trimestre) &
            (df_despachados['FECHA RESOLUCIÓN'] <= _fecha_max)
        ])
        semanas_trimestre = min(13, ((_fecha_max - fecha_inicio_trimestre).days / 7)) if fecha_inicio_trimestre < _fecha_max else 0
        rendimiento_trimestral = despachados_trimestre / semanas_trimestre if semanas_trimestre > 0 else 0
        
        # Último mes
        despachados_mes = len(df_despachados[
            (df_despachados['USUARIO'] == usuario) & 
            (estado == 'ACTIVO') &
            (df_despachados['FECHA RESOLUCIÓN'] >= fecha_inicio_mes) &
            (df_despachados['FECHA RESOLUCIÓN'] <= _fecha_max)
        ])
        semanas_mes = min(4, ((_fecha_max - fecha_inicio_mes).days / 7)) if fecha_inicio_mes < _fecha_max else 0
        rendimiento_mensual = despachados_mes / semanas_mes if semanas_mes > 0 else 0
        
        # Última semana
        despachados_semana = len(df_despachados[
            (df_despachados['USUARIO'] == usuario) & 
            (estado == 'ACTIVO') &
            (df_despachados['FECHA RESOLUCIÓN'] >= fecha_inicio_semana) &
            (df_despachados['FECHA RESOLUCIÓN'] <= _fecha_max)
        ])
        semanas_semana = min(1, ((_fecha_max - fecha_inicio_semana).days / 7)) if fecha_inicio_semana < _fecha_max else 0
        rendimiento_semanal = despachados_semana / semanas_semana if semanas_semana > 0 else 0
        
        # Rendimiento total
        rendimiento_total = expedientes_despachados / semanas_efectivas if semanas_efectivas > 0 else 0
        
        resultados.append({
            'USUARIO': usuario,
            'EQUIPOS': equipos_str,  # Mostrar todos los equipos en una columna
            'ESTADO': estado,
            'EXPEDIENTES_DESPACHADOS': expedientes_despachados,
            'SEMANAS_EFECTIVAS': round(semanas_efectivas, 1),
            'RENDIMIENTO_TOTAL': round(rendimiento_total, 2),
            'RENDIMIENTO_ANUAL': round(rendimiento_anual, 2),
            'POTENCIAL_ANUAL': round(potencial_anual, 1),  # NUEVA COLUMNA
            'RENDIMIENTO_TRIMESTRAL': round(rendimiento_trimestral, 2),
            'RENDIMIENTO_MENSUAL': round(rendimiento_mensual, 2),
            'RENDIMIENTO_SEMANAL': round(rendimiento_semanal, 2)
        })
    
    return pd.DataFrame(resultados)

# =============================================
# HANDSONTABLE - VERSIÓN CORREGIDA (VISUALIZACIÓN COMPLETA)
# =============================================
def mostrar_con_handsontable(df_filtrado):
    """
    Versión funcional que muestra todas las columnas
    """
    import io
    import streamlit.components.v1 as components
    import json
    from datetime import datetime
    
    # 1. Crear copia y formatear fechas de forma SEGURA
    df_display = df_filtrado.copy()

    df_display['ANTIGÜEDAD EXP. (DÍAS)'] = df_display['ANTIGÜEDAD EXP. (DÍAS)'].round().astype('Int64')
    
    # Función segura para formatear fechas
    def formatear_fecha(x):
        try:
            # Si es NaT o NaN
            if pd.isna(x):
                return ""
            
            # Si es string que contiene 9999
            if isinstance(x, str) and '9999' in str(x):
                return "09/09/9999"
            
            # Si es Timestamp 9999
            if isinstance(x, pd.Timestamp) and hasattr(x, 'year') and x.year == 9999:
                return "09/09/9999"
            
            # Si es datetime válido
            if isinstance(x, (pd.Timestamp, datetime)):
                return x.strftime("%d/%m/%Y")
            
            # Si es otro tipo, convertir a string
            return str(x)
            
        except Exception:
            return ""
    
    # Aplicar a columnas de fecha
    columnas_fecha = [col for col in df_display.columns if 'FECHA' in col.upper()]
    
    for col in columnas_fecha:
        if col in df_display.columns:
            df_display[col] = df_display[col].apply(formatear_fecha)
    
    # 2. Preparar datos para Handsontable CORRECTAMENTE
    # Convertir DataFrame a lista de diccionarios (formato que espera Handsontable)
    data = df_display.to_dict('records')
    
    # Preparar columnas
    columns = []
    for idx, col_name in enumerate(df_display.columns):
        col_config = {
            'data': col_name,
            'title': col_name,
            'type': 'text'
        }
        
        # Configurar columnas de fecha
        if 'FECHA' in col_name.upper():
            col_config.update({
                'type': 'date',
                'dateFormat': 'DD/MM/YYYY',
                'correctFormat': True
            })
        
        # Configurar columnas numéricas
        elif df_display[col_name].dtype in ['int64', 'float64']:
            col_config.update({
                'type': 'numeric',
                'numericFormat': {
                    'pattern': '0,0'
                }
            })
        
        columns.append(col_config)
    
    # 3. HTML/JavaScript CORREGIDO
    hot_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <script src="https://cdn.jsdelivr.net/npm/handsontable/dist/handsontable.full.min.js"></script>
        <link href="https://cdn.jsdelivr.net/npm/handsontable/dist/handsontable.full.min.css" rel="stylesheet">
        <style>
            body {{
                margin: 0;
                padding: 0;
            }}
            #hot-container {{
                width: 100%;
                height: 600px;
                overflow: hidden;
            }}
            .handsontable {{
                font-size: 11px;
            }}
            .handsontable thead th {{
                background-color: #007933 !important;
                color: white !important;
                font-weight: bold !important;
            }}
        </style>
    </head>
    <body>
        <div id="hot-container"></div>
        
        <script>
            // Datos y columnas
            var data = {json.dumps(data, default=str)};
            var columns = {json.dumps(columns)};
            
            console.log("Datos cargados:", data.length, "registros");
            console.log("Columnas:", columns.length);
            
            // Configuración
            var config = {{
                data: data,
                columns: columns,
                colHeaders: true,
                rowHeaders: true,
                height: 600,
                width: '100%',
                autoColumnSize: {{
                    useHeaders: true,       // IMPORTANTE: Considerar encabezados
                    syncLimit: 50,          // Reducir para mejor rendimiento
                    samplingRatio: 10,      // Menor ratio para cálculo más rápido
                    limitToWindowSize: false // Permitir anchos mayores que la ventana
                }},
                manualColumnResize: true,   // Permitir ajuste manual después
                manualRowResize: true,
                stretchH: 'all',           // Cambiado de 'last' a 'all' para mejor ajuste
                licenseKey: 'non-commercial-and-evaluation',
                filters: true,
                dropdownMenu: true,
                contextMenu: true,
                autoWrapRow: true,
                wordWrap: true,             // Cambiado a true para mejor visualización
                columnSorting: true         // Añadido para mejor UX
            }};
            
            // Inicializar
            var container = document.getElementById('hot-container');
            var hot = new Handsontable(container, config);
            
            // Asegurar que se renderice completamente
            setTimeout(function() {{
                hot.render();
                console.log("Handsontable renderizado");
            }}, 100);
        </script>
    </body>
    </html>
    """
    
    # 4. Mostrar Handsontable
    st.subheader("📊 Vista de expedientes")
    st.write(f"**Mostrando {len(df_display)} registros - {len(df_display.columns)} columnas**")
    
    # Mostrar Handsontable
    components.html(hot_html, height=650, scrolling=False)
    
    # 5. Exportación a Excel (igual que antes, funciona bien)
    st.markdown("---")
    st.subheader("💾 Exportar Datos")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        nombre = st.text_input(
            "Nombre del archivo:",
            value=f"RECTAUTO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            key="excel_filename_hot"
        )
    
    with col2:
        if st.button("📥 Exportar a Excel", type="primary", use_container_width=True):
            with st.spinner("Generando Excel..."):
                try:
                    # Preparar DataFrame para exportación
                    df_export = df_filtrado.copy()

                    # Identificar columnas de usuario y equipo
                    columnas_usuario = [col for col in df_export.columns if 'USUARIO' in col.upper()]
                    columnas_equipo = [col for col in df_export.columns if 'EQUIPO' in col.upper()]
                    
                    col_usuario = columnas_usuario[0] if columnas_usuario else None
                    col_equipo = columnas_equipo[0] if columnas_equipo else None
                    
                    # ============================================================
                    # SOLUCIÓN 1: REORDENAR COLUMNAS - USUARIO Y EQUIPO AL PRINCIPIO
                    # ============================================================
                    
                    # Crear lista de columnas ordenadas
                    columnas_ordenadas = []
                    
                    # 1. Primero las columnas clave que queremos al principio
                    columnas_prioridad = []
                    
                    if 'RUE' in df_export.columns:
                        columnas_prioridad.append('RUE')
                    
                    if col_usuario:
                        columnas_prioridad.append(col_usuario)
                    
                    if col_equipo:
                        columnas_prioridad.append(col_equipo)
                    
                    # 2. Otras columnas importantes
                    otras_importantes = ['FECHA DE ENTRADA', 'ANTIGÜEDAD EXP. (DÍAS)', 
                                        'ETIQ. PENÚLTIMO TRAM.', 'PENÚLTIMO TRAMITE']
                    
                    for col in otras_importantes:
                        if col in df_export.columns:
                            columnas_prioridad.append(col)
                    
                    # 3. Eliminar duplicados y añadir el resto de columnas
                    columnas_ordenadas = columnas_prioridad.copy()
                    
                    for col in df_export.columns:
                        if col not in columnas_ordenadas:
                            columnas_ordenadas.append(col)
                    
                    # Reordenar el DataFrame
                    df_export = df_export[columnas_ordenadas]
                    
                    # Función segura para exportación
                    def preparar_fecha_export(x):
                        try:
                            if pd.isna(x):
                                return None
                            if isinstance(x, pd.Timestamp) and x.year == 9999:
                                return "09/09/9999"
                            if isinstance(x, datetime) and x.year == 9999:
                                return "09/09/9999"
                            return x
                        except:
                            return None
                    
                    # Aplicar a columnas de fecha
                    for col in columnas_fecha:
                        if col in df_export.columns:
                            df_export[col] = df_export[col].apply(preparar_fecha_export)
                    
                    # APLICAR ROUND Y CONVERSIÓN A INT64 PARA ANTIGÜEDAD
                    if 'ANTIGÜEDAD EXP. (DÍAS)' in df_export.columns:
                        # Aplicar round y convertir a Int64 (que admite NaN)
                        df_export['ANTIGÜEDAD EXP. (DÍAS)'] = df_export['ANTIGÜEDAD EXP. (DÍAS)'].round().astype('Int64')
                        # Reemplazar <NA> por None para Excel
                        df_export['ANTIGÜEDAD EXP. (DÍAS)'] = df_export['ANTIGÜEDAD EXP. (DÍAS)'].where(pd.notna, None)
                    
                    # Crear Excel
                    output = io.BytesIO()
                    
                    # Usar openpyxl directamente
                    from openpyxl import Workbook
                    from openpyxl.utils import get_column_letter
                    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Expedientes"
                    
                    # ============================================================================
                    # ESCRIBIR ENCABEZADOS DE LA TABLA (FILA 1)
                    # ============================================================================
                    
                    header_fill = PatternFill(start_color="007933", end_color="007933", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    
                    for col_num, header in enumerate(df_export.columns, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                    
                    # ============================================================================
                    # ESCRIBIR DATOS DE LA TABLA (DESDE FILA 2)
                    # ============================================================================
                    
                    data_font = Font(size=10)
                    number_alignment = Alignment(horizontal="right", vertical="center")
                    text_alignment = Alignment(horizontal="left", vertical="center")
                    
                    for row_num, row in enumerate(df_export.values, 2):
                        for col_num, value in enumerate(row, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            cell.font = data_font
                            
                            # Aplicar alineación específica según tipo de dato
                            col_name = df_export.columns[col_num-1]
                            
                            # Si es ANTIGÜEDAD o otra columna numérica
                            if 'ANTIGÜEDAD' in col_name.upper() or 'DÍAS' in col_name.upper():
                                cell.alignment = number_alignment
                                # Formato sin decimales para números enteros
                                if isinstance(value, (int, float)) and not pd.isna(value):
                                    cell.number_format = '#,##0'
                            # Si es columna de fecha
                            elif 'FECHA' in col_name.upper():
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                date_format = 'DD/MM/YYYY'
                                if isinstance(value, (pd.Timestamp, datetime)):
                                    cell.number_format = date_format
                            # Si es columna numérica genérica
                            elif isinstance(value, (int, float)) and not pd.isna(value):
                                cell.alignment = number_alignment
                                # Formato con separador de miles
                                cell.number_format = '#,##0'
                            else:
                                cell.alignment = text_alignment
                    
                    # ============================================================================
                    # APLICAR FORMATO DE FECHA ESPECÍFICO
                    # ============================================================================
                    
                    date_format = 'DD/MM/YYYY'
                    for col_num, col_name in enumerate(df_export.columns, 1):
                        if 'FECHA' in col_name.upper():
                            col_letter = get_column_letter(col_num)
                            for row in ws.iter_rows(min_row=2, max_row=len(df_export)+1, min_col=col_num, max_col=col_num):
                                cell = row[0]
                                if isinstance(cell.value, (pd.Timestamp, datetime)):
                                    cell.number_format = date_format
                    
                    # ============================================================================
                    # AJUSTAR ANCHOS DE COLUMNA AUTOMÁTICAMENTE
                    # ============================================================================
                    
                    for col_num, column_title in enumerate(df_export.columns, 1):
                        max_length = 0
                        column_letter = get_column_letter(col_num)
                        
                        # Calcular longitud máxima del contenido
                        for row_num in range(1, len(df_export) + 2):  # +2 para encabezados
                            cell_value = ws.cell(row=row_num, column=col_num).value
                            if cell_value:
                                # Para números, considerar formato con separador de miles
                                if isinstance(cell_value, (int, float)):
                                    formatted_value = f"{cell_value:,.0f}"
                                    cell_length = len(formatted_value)
                                else:
                                    cell_length = len(str(cell_value))
                                max_length = max(max_length, cell_length)
                        
                        # Establecer ancho (mínimo 12, máximo 50)
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
                    
                    # ============================================================================
                    # APLICAR AUTOFILTER
                    # ============================================================================
                    
                    if len(df_export) > 0:
                        # Determinar el rango de la tabla (encabezados + datos)
                        start_col = get_column_letter(1)
                        end_col = get_column_letter(len(df_export.columns))
                        end_row = len(df_export) + 1  # +1 por los encabezados
                        
                        # Aplicar autofilter
                        ws.auto_filter.ref = f"{start_col}1:{end_col}{end_row}"
                    
                    # ============================================================================
                    # CONGELAR ENCABEZADOS (FILA 1) Y RUE (COLUMNA a)
                    # ============================================================================
                    
                    # Congelar la primera fila para que sea visible al hacer scroll
                    ws.freeze_panes = 'B2'
                    
                    # ============================================================================
                    # APLICAR BORDES A LA TABLA
                    # ============================================================================
                    
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in ws.iter_rows(min_row=1, max_row=len(df_export)+1, 
                                        min_col=1, max_col=len(df_export.columns)):
                        for cell in row:
                            cell.border = thin_border
                    
                    # ============================================================================
                    # GUARDAR Y MOSTRAR BOTÓN DE DESCARGA
                    # ============================================================================
                    
                    wb.save(output)
                    output.seek(0)
                    
                    # Botón de descarga
                    st.download_button(
                        label="⬇️ Descargar Excel",
                        data=output.read(),
                        file_name=nombre,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("✅ Excel generado correctamente")
                    
                except Exception as e:
                    st.error(f"❌ Error: {e}")
    
    return df_display

# =============================================
# PÁGINA 1: CARGA DE ARCHIVOS
# =============================================
if eleccion == "Carga de Archivos":
    st.header("📁 Carga de Archivos")
    
    # Mostrar botones de limpieza solo en esta página
    with st.sidebar:
        st.markdown("---")
        st.subheader("🛠️ Herramientas de Mantenimiento")
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("🔄 Limpiar cache", help="Limpiar toda la cache y recargar", use_container_width=True):
                st.cache_data.clear()
                # Mantener solo los datos esenciales
                keys_to_keep = ['df_combinado', 'df_usuarios', 'archivos_hash', 'filtro_estado', 'filtro_equipo', 'filtro_usuario']
                for key in list(st.session_state.keys()):
                    if key not in keys_to_keep:
                        del st.session_state[key]
                st.success("Cache limpiada correctamente")
                st.rerun()
        
        with col2:
            if st.button("🧹 Limpiar temp", help="Limpiar archivos temporales", use_container_width=True):
                user_env.cleanup()
                st.success("Archivos temporales limpiados")

    # NUEVA SECCIÓN: CARGA DE CINCO ARCHIVOS (incluyendo DOCUMENTOS)
    st.markdown("---")
    st.subheader("📁 Carga de Archivos")

    # Crear cinco columnas para los archivos
    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.markdown('<div style="text-align: center; background-color: #1f77b4; padding: 10px; border-radius: 5px; margin-bottom: 10px;">'
                    '<h4 style="color: white; margin: 0;">📊 RECTAUTO</h4>'
                    '</div>', unsafe_allow_html=True)
        archivo_rectauto = st.file_uploader(
            "Archivo principal de expedientes",
            type=["xlsx", "xls"],
            key="rectauto_upload",
            label_visibility="collapsed",
            help="Sube el archivo principal RECTAUTO"
        )
        if archivo_rectauto:
            st.success(f"✅ {archivo_rectauto.name}")
        else:
            st.info("⏳ Esperando archivo RECTAUTO")

    with col2:
        st.markdown('<div style="text-align: center; background-color: #ff7f0e; padding: 10px; border-radius: 5px; margin-bottom: 10px;">'
                    '<h4 style="color: white; margin: 0;">📨 NOTIFICA</h4>'
                    '</div>', unsafe_allow_html=True)
        archivo_notifica = st.file_uploader(
            "Archivo de notificaciones",
            type=["xlsx", "xls"],
            key="notifica_upload",
            label_visibility="collapsed",
            help="Sube el archivo NOTIFICA"
        )
        if archivo_notifica:
            st.success(f"✅ {archivo_notifica.name}")
        else:
            st.info("⏳ Esperando archivo NOTIFICA")

    with col3:
        st.markdown('<div style="text-align: center; background-color: #2ca02c; padding: 10px; border-radius: 5px; margin-bottom: 10px;">'
                    '<h4 style="color: white; margin: 0;">⚡ TRIAJE</h4>'
                    '</div>', unsafe_allow_html=True)
        archivo_triaje = st.file_uploader(
            "Archivo de triaje",
            type=["xlsx", "xls"],
            key="triaje_upload",
            label_visibility="collapsed",
            help="Sube el archivo TRIAJE"
        )
        if archivo_triaje:
            st.success(f"✅ {archivo_triaje.name}")
        else:
            st.info("⏳ Esperando archivo TRIAJE")

    with col4:
        st.markdown('<div style="text-align: center; background-color: #9467bd; padding: 10px; border-radius: 5px; margin-bottom: 10px;">'
                    '<h4 style="color: white; margin: 0;">👥 USUARIOS</h4>'
                    '</div>', unsafe_allow_html=True)
        archivo_usuarios = st.file_uploader(
            "Archivo de configuración de usuarios",
            type=["xlsx", "xls"],
            key="usuarios_upload",
            label_visibility="collapsed",
            help="Sube el archivo USUARIOS.xlsx"
        )
        if archivo_usuarios:
            st.success(f"✅ {archivo_usuarios.name}")
        else:
            st.info("⏳ Esperando archivo USUARIOS")

    with col5:
        st.markdown('<div style="text-align: center; background-color: #d62728; padding: 10px; border-radius: 5px; margin-bottom: 10px;">'
                    '<h4 style="color: white; margin: 0;">📄 DOCUMENTOS</h4>'
                    '</div>', unsafe_allow_html=True)
        archivo_documentos = st.file_uploader(
            "Archivo de documentación incorporada",
            type=["xlsx"],
            key="documentos_upload",
            label_visibility="collapsed",
            help="Sube el archivo DOCUMENTOS.xlsx"
        )
        if archivo_documentos:
            st.success(f"✅ {archivo_documentos.name}")
        else:
            st.info("⏳ Esperando archivo DOCUMENTOS")

    # Estado de carga
    st.markdown("---")
    st.subheader("📋 Estado de Carga")

    # Mostrar estado con métricas (6 columnas ahora)
    estado_col1, estado_col2, estado_col3, estado_col4, estado_col5, estado_col6 = st.columns(6)

    with estado_col1:
        rectauto_status = "✅ Cargado" if archivo_rectauto else "❌ Pendiente"
        st.metric("RECTAUTO", rectauto_status)

    with estado_col2:
        notifica_status = "✅ Cargado" if archivo_notifica else "❌ Pendiente"
        st.metric("NOTIFICA", notifica_status)

    with estado_col3:
        triaje_status = "✅ Cargado" if archivo_triaje else "❌ Pendiente"
        st.metric("TRIAJE", triaje_status)

    with estado_col4:
        usuarios_status = "✅ Cargado" if archivo_usuarios else "❌ Pendiente"
        st.metric("USUARIOS", usuarios_status)

    with estado_col5:
        documentos_status = "✅ Cargado" if archivo_documentos else "❌ Pendiente"
        st.metric("DOCUMENTOS", documentos_status)

    with estado_col6:
        archivos_cargados = sum([1 for f in [archivo_rectauto, archivo_notifica, archivo_triaje, archivo_usuarios, archivo_documentos] if f])
        st.metric("Total Cargados", f"{archivos_cargados}/5")

    # Procesar archivos cuando estén listos usando la función optimizada
    if archivo_rectauto:
        # Verificar si los archivos han cambiado
        archivos_actuales = {
            'rectauto': obtener_hash_archivo(archivo_rectauto),
            'notifica': obtener_hash_archivo(archivo_notifica) if archivo_notifica else None,
            'triaje': obtener_hash_archivo(archivo_triaje) if archivo_triaje else None,
            'usuarios': obtener_hash_archivo(archivo_usuarios) if archivo_usuarios else None,
            'documentos': obtener_hash_archivo(archivo_documentos) if archivo_documentos else None
        }
        
        archivos_guardados = st.session_state.get("archivos_hash", {})
        
        # Si los archivos son nuevos o cambiaron, procesar
        if (archivos_actuales != archivos_guardados or 
            "df_combinado" not in st.session_state):
            
            with st.spinner("🔄 Procesando archivos combinados..."):
                try:
                    # Usar la función optimizada de procesamiento combinado
                    archivos_dict = {
                        'rectauto': archivo_rectauto,
                        'notifica': archivo_notifica,
                        'triaje': archivo_triaje,
                        'usuarios': archivo_usuarios,
                        'documentos': archivo_documentos
                    }
                    
                    df_combinado, df_usuarios, datos_documentos = procesar_archivos_combinado(archivos_dict)
                    
                    # Convertir columnas de fecha
                    df_combinado = convertir_fechas(df_combinado)
                    
                    # Guardar en session_state
                    st.session_state["df_combinado"] = df_combinado
                    st.session_state["df_usuarios"] = df_usuarios
                    st.session_state["datos_documentos"] = datos_documentos
                    st.session_state["archivos_hash"] = archivos_actuales
                    
                    st.success(f"✅ Archivos procesados correctamente")
                    st.info(f"📊 Dataset final: {len(df_combinado)} registros, {len(df_combinado.columns)} columnas")
                    if df_usuarios is not None:
                        st.info(f"👥 Usuarios cargados: {len(df_usuarios)} registros")
                    if datos_documentos is not None:
                        st.info(f"📄 Documentos cargados: {len(datos_documentos['documentos'])} registros")
                    
                except Exception as e:
                    st.error(f"❌ Error procesando archivos: {e}")
                    # Fallback: usar solo RECTAUTO
                    with st.spinner("🔄 Cargando solo RECTAUTO..."):
                        df_rectauto = cargar_y_procesar_rectauto(archivo_rectauto)
                        st.session_state["df_combinado"] = df_rectauto
                        st.session_state["df_usuarios"] = None
                        st.session_state["datos_documentos"] = None
                        st.session_state["archivos_hash"] = archivos_actuales
                        st.warning("⚠️ Usando solo archivo RECTAUTO debido a errores en combinación")
        
        else:
            # Usar datos cacheados
            df_combinado = st.session_state["df_combinado"]
            df_usuarios = st.session_state.get("df_usuarios", None)
            datos_documentos = st.session_state.get("datos_documentos", None)
            st.sidebar.success("✅ Usando datos combinados cacheados")

    elif "df_combinado" in st.session_state:
        # Usar datos previamente cargados
        df_combinado = st.session_state["df_combinado"]
        df_usuarios = st.session_state.get("df_usuarios", None)
        datos_documentos = st.session_state.get("datos_documentos", None)
        st.sidebar.info("📊 Datos combinados cargados desde cache")
    else:
        st.warning("⚠️ **Carga obligatoria:** Sube al menos el archivo RECTAUTO para continuar")
        st.info("💡 **Archivos opcionales:** NOTIFICA, TRIAJE, USUARIOS y DOCUMENTOS enriquecerán el análisis")
        st.stop()

    # Mostrar información del dataset combinado
    with st.expander("📊 Información del Dataset Combinado"):
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Registros", f"{len(df_combinado):,}".replace(",", "."))
        
        with col2:
            st.metric("Total Columnas", len(df_combinado.columns))
        
        with col3:
            archivos_usados = 1
            if archivo_notifica and 'FECHA NOTIFICACIÓN' in df_combinado.columns:
                archivos_usados += 1
            if archivo_triaje and any(col in df_combinado.columns for col in ['USUARIO-CSV', 'CALIFICACIÓN', 'OBSERVACIONES', 'FECHA ASIG']):
                archivos_usados += 1
            if archivo_documentos and 'DOCUM.INCORP.' in df_combinado.columns:
                archivos_usados += 1
            st.metric("Archivos Usados", f"{archivos_usados}/4")
        
        with col4:
            documentos_status = "✅ Cargado" if datos_documentos is not None else "❌ No cargado"
            st.metric("DOCUMENTOS", documentos_status)
        
        # Mostrar primeras filas SIN formato condicional para evitar errores
        st.write("**Vista previa del dataset combinado:**")
        df_mostrar_preview = df_combinado.head(3).copy()
        for col in df_mostrar_preview.select_dtypes(include='datetime').columns:
            df_mostrar_preview[col] = df_mostrar_preview[col].dt.strftime("%d/%m/%Y")
        
        st.dataframe(df_mostrar_preview, use_container_width=True)
        
        # Mostrar columnas disponibles
        st.write("**Columnas disponibles:**")
        columnas_grupos = {}
        for col in df_combinado.columns:
            if col == 'FECHA NOTIFICACIÓN':
                grupo = 'NOTIFICA'
            elif col in ['USUARIO-CSV', 'CALIFICACIÓN', 'OBSERVACIONES', 'FECHA ASIG']:
                grupo = 'TRIAJE'
            elif col == 'DOCUM.INCORP.':
                grupo = 'DOCUMENTACIÓN'
            else:
                grupo = 'RECTAUTO'
            
            if grupo not in columnas_grupos:
                columnas_grupos[grupo] = []
            columnas_grupos[grupo].append(col)
        
        for grupo, columnas in columnas_grupos.items():
            with st.expander(f"📋 Columnas de {grupo} ({len(columnas)})"):
                st.write(columnas)

# =============================================
# PÁGINA 2: VISTA DE EXPEDIENTES
# =============================================
elif eleccion == "Vista de Expedientes":
    st.header("📋 Vista de Expedientes")
    
    if "df_combinado" not in st.session_state:
        st.warning("⚠️ Primero carga los archivos en la sección 'Carga de Archivos'")
        st.stop()
    
    # Usar df_combinado en lugar de df
    df = st.session_state["df_combinado"]
    df_usuarios = st.session_state.get("df_usuarios", None)
    datos_documentos = st.session_state.get("datos_documentos", None)
    
    # =============================================
    # FILTROS DINÁMICOS INTERCONECTADOS - VERSIÓN MEJORADA CON ETIQUETAS
    # =============================================
    
    st.sidebar.header("Filtros Interconectados")

    # Inicializar variables de sesión si no existen
    if 'filtro_estado' not in st.session_state:
        st.session_state.filtro_estado = ['Abierto'] if 'Abierto' in df['ESTADO'].values else []

    if 'filtro_equipo' not in st.session_state:
        st.session_state.filtro_equipo = []

    if 'filtro_usuario' not in st.session_state:
        st.session_state.filtro_usuario = []

    if 'filtro_etiq_penultimo' not in st.session_state:
        st.session_state.filtro_etiq_penultimo = []

    if 'filtro_etiq_ultimo' not in st.session_state:
        st.session_state.filtro_etiq_ultimo = []

    # Botón para resetear filtros
    if st.sidebar.button("🔄 Mostrar todos / Resetear filtros", use_container_width=True):
        st.session_state.filtro_estado = []
        st.session_state.filtro_equipo = []
        st.session_state.filtro_usuario = []
        st.session_state.filtro_etiq_penultimo = []
        st.session_state.filtro_etiq_ultimo = []

        #st.session_state.filtro_estado = sorted(df['ESTADO'].dropna().unique())
        #st.session_state.filtro_equipo = sorted(df['EQUIPO'].dropna().unique())
        #st.session_state.filtro_usuario = sorted(df['USUARIO'].dropna().unique())
        #if 'ETIQ. PENÚLTIMO TRAM.' in df.columns:
        #    st.session_state.filtro_etiq_penultimo = sorted(df['ETIQ. PENÚLTIMO TRAM.'].dropna().unique())
        #if 'ETIQ. ÚLTIMO TRAM.' in df.columns:
        #    st.session_state.filtro_etiq_ultimo = sorted(df['ETIQ. ÚLTIMO TRAM.'].dropna().unique())
        st.rerun()

    # 1. Primero aplicar filtros secuencialmente para calcular opciones disponibles
    df_filtrado_temp = df.copy()

    # Aplicar filtro de ESTADO primero
    if st.session_state.filtro_estado:
        df_filtrado_temp = df_filtrado_temp[df_filtrado_temp['ESTADO'].isin(st.session_state.filtro_estado)]

    # Calcular EQUIPOS disponibles basados en el filtro de ESTADO
    equipos_disponibles = sorted(df_filtrado_temp['EQUIPO'].dropna().unique())

    # Aplicar filtro de EQUIPO (si hay selección)
    equipos_seleccionados = [eq for eq in st.session_state.filtro_equipo if eq in equipos_disponibles]
    if equipos_seleccionados:
        df_filtrado_temp = df_filtrado_temp[df_filtrado_temp['EQUIPO'].isin(equipos_seleccionados)]

    # Calcular USUARIOS disponibles basados en filtros de ESTADO y EQUIPO
    usuarios_disponibles = sorted(df_filtrado_temp['USUARIO'].dropna().unique())

    # Aplicar filtro de USUARIO (si hay selección)
    usuarios_seleccionados = [us for us in st.session_state.filtro_usuario if us in usuarios_disponibles]
    if usuarios_seleccionados:
        df_filtrado_temp = df_filtrado_temp[df_filtrado_temp['USUARIO'].isin(usuarios_seleccionados)]

    # Calcular ETIQ. PENÚLTIMO TRAM. disponibles basados en filtros anteriores
    etiq_penultimo_disponibles = []
    if 'ETIQ. PENÚLTIMO TRAM.' in df_filtrado_temp.columns:
        etiq_penultimo_disponibles = sorted(df_filtrado_temp['ETIQ. PENÚLTIMO TRAM.'].dropna().unique())

    # Aplicar filtro de ETIQ. PENÚLTIMO TRAM. (si hay selección)
    etiq_penultimo_seleccionados = [etiq for etiq in st.session_state.filtro_etiq_penultimo if etiq in etiq_penultimo_disponibles]
    if etiq_penultimo_seleccionados:
        df_filtrado_temp = df_filtrado_temp[df_filtrado_temp['ETIQ. PENÚLTIMO TRAM.'].isin(etiq_penultimo_seleccionados)]

    # Calcular ETIQ. ÚLTIMO TRAM. disponibles basados en todos los filtros anteriores
    etiq_ultimo_disponibles = []
    if 'ETIQ. ÚLTIMO TRAM.' in df_filtrado_temp.columns:
        etiq_ultimo_disponibles = sorted(df_filtrado_temp['ETIQ. ÚLTIMO TRAM.'].dropna().unique())

    # Aplicar filtro de ETIQ. ÚLTIMO TRAM. (si hay selección)
    etiq_ultimo_seleccionados = [etiq for etiq in st.session_state.filtro_etiq_ultimo if etiq in etiq_ultimo_disponibles]
    if etiq_ultimo_seleccionados:
        df_filtrado_temp = df_filtrado_temp[df_filtrado_temp['ETIQ. ÚLTIMO TRAM.'].isin(etiq_ultimo_seleccionados)]

    # 2. Ahora crear los widgets de filtro con opciones actualizadas
    st.sidebar.markdown("---")
    st.sidebar.subheader("Filtros Activos")

    # FILTRO DE ESTADO (siempre muestra todas las opciones)
    opciones_estado = sorted(df['ESTADO'].dropna().unique())
    estado_sel = st.sidebar.multiselect(
        "🔘 Selecciona Estado:",
        options=opciones_estado,
        default=st.session_state.filtro_estado,
        key='filtro_estado_selector'
    )

    # FILTRO DE EQUIPO (se actualiza según estado seleccionado)
    equipo_sel = st.sidebar.multiselect(
        "👥 Selecciona Equipo:",
        options=equipos_disponibles,
        default=equipos_seleccionados,
        key='filtro_equipo_selector'
    )

    # FILTRO DE USUARIO (se actualiza según estado y equipo seleccionados)
    usuario_sel = st.sidebar.multiselect(
        "👤 Selecciona Usuario:",
        options=usuarios_disponibles,
        default=usuarios_seleccionados,
        key='filtro_usuario_selector'
    )

    # NUEVOS FILTROS: ETIQ. PENÚLTIMO TRAM. (se actualiza según filtros anteriores)
    if 'ETIQ. PENÚLTIMO TRAM.' in df.columns:
        etiq_penultimo_sel = st.sidebar.multiselect(
            "🏷️ ETIQ. PENÚLTIMO TRAM.:",
            options=etiq_penultimo_disponibles,
            default=etiq_penultimo_seleccionados,
            key='filtro_etiq_penultimo_selector'
        )
    else:
        etiq_penultimo_sel = []
        st.sidebar.info("ℹ️ Columna 'ETIQ. PENÚLTIMO TRAM.' no disponible")

    # NUEVOS FILTROS: ETIQ. ÚLTIMO TRAM. (se actualiza según todos los filtros anteriores)
    if 'ETIQ. ÚLTIMO TRAM.' in df.columns:
        etiq_ultimo_sel = st.sidebar.multiselect(
            "🏷️ ETIQ. ÚLTIMO TRAM.:",
            options=etiq_ultimo_disponibles,
            default=etiq_ultimo_seleccionados,
            key='filtro_etiq_ultimo_selector'
        )
    else:
        etiq_ultimo_sel = []
        st.sidebar.info("ℹ️ Columna 'ETIQ. ÚLTIMO TRAM.' no disponible")

    # 3. Actualizar session_state cuando cambian los filtros
    if estado_sel != st.session_state.filtro_estado:
        st.session_state.filtro_estado = estado_sel
        # Cuando cambia el estado, resetear todos los filtros dependientes
        st.session_state.filtro_equipo = []
        st.session_state.filtro_usuario = []
        st.session_state.filtro_etiq_penultimo = []
        st.session_state.filtro_etiq_ultimo = []
        st.rerun()

    if equipo_sel != st.session_state.filtro_equipo:
        st.session_state.filtro_equipo = equipo_sel
        # Cuando cambia el equipo, resetear usuario y etiquetas
        st.session_state.filtro_usuario = []
        st.session_state.filtro_etiq_penultimo = []
        st.session_state.filtro_etiq_ultimo = []
        st.rerun()

    if usuario_sel != st.session_state.filtro_usuario:
        st.session_state.filtro_usuario = usuario_sel
        # Cuando cambia el usuario, resetear etiquetas
        st.session_state.filtro_etiq_penultimo = []
        st.session_state.filtro_etiq_ultimo = []
        st.rerun()

    if 'ETIQ. PENÚLTIMO TRAM.' in df.columns and etiq_penultimo_sel != st.session_state.filtro_etiq_penultimo:
        st.session_state.filtro_etiq_penultimo = etiq_penultimo_sel
        # Cuando cambia etiq penúltimo, resetear etiq último
        st.session_state.filtro_etiq_ultimo = []
        st.rerun()

    if 'ETIQ. ÚLTIMO TRAM.' in df.columns and etiq_ultimo_sel != st.session_state.filtro_etiq_ultimo:
        st.session_state.filtro_etiq_ultimo = etiq_ultimo_sel
        st.rerun()

    # 4. Aplicar filtros finales al DataFrame principal
    df_filtrado = df.copy()

    if st.session_state.filtro_estado:
        df_filtrado = df_filtrado[df_filtrado['ESTADO'].isin(st.session_state.filtro_estado)]

    if st.session_state.filtro_equipo:
        df_filtrado = df_filtrado[df_filtrado['EQUIPO'].isin(st.session_state.filtro_equipo)]

    if st.session_state.filtro_usuario:
        df_filtrado = df_filtrado[df_filtrado['USUARIO'].isin(st.session_state.filtro_usuario)]

    if 'ETIQ. PENÚLTIMO TRAM.' in df.columns and st.session_state.filtro_etiq_penultimo:
        df_filtrado = df_filtrado[df_filtrado['ETIQ. PENÚLTIMO TRAM.'].isin(st.session_state.filtro_etiq_penultimo)]

    if 'ETIQ. ÚLTIMO TRAM.' in df.columns and st.session_state.filtro_etiq_ultimo:
        df_filtrado = df_filtrado[df_filtrado['ETIQ. ÚLTIMO TRAM.'].isin(st.session_state.filtro_etiq_ultimo)]

    # Mostrar resumen de filtros activos
    st.sidebar.markdown("---")
    st.sidebar.subheader("📊 Resumen Filtros")
    
    if st.session_state.filtro_estado:
        st.sidebar.write(f"**Estados:** {len(st.session_state.filtro_estado)} seleccionados")
    
    if st.session_state.filtro_equipo:
        st.sidebar.write(f"**Equipos:** {len(st.session_state.filtro_equipo)} seleccionados")
    
    if st.session_state.filtro_usuario:
        st.sidebar.write(f"**Usuarios:** {len(st.session_state.filtro_usuario)} seleccionados")
    
    if 'ETIQ. PENÚLTIMO TRAM.' in df.columns and st.session_state.filtro_etiq_penultimo:
        st.sidebar.write(f"**ETIQ. PENÚLTIMO:** {len(st.session_state.filtro_etiq_penultimo)} seleccionados")
    
    if 'ETIQ. ÚLTIMO TRAM.' in df.columns and st.session_state.filtro_etiq_ultimo:
        st.sidebar.write(f"**ETIQ. ÚLTIMO:** {len(st.session_state.filtro_etiq_ultimo)} seleccionados")
    
    st.sidebar.write(f"**Registros:** {len(df_filtrado):,}".replace(",", "."))

    # Mostrar detalles de filtros activos en un expander
    with st.sidebar.expander("📋 Ver detalles de filtros"):
        if st.session_state.filtro_estado:
            st.write("**Estados seleccionados:**")
            for estado in st.session_state.filtro_estado:
                st.write(f"- {estado}")
        
        if st.session_state.filtro_equipo:
            st.write("**Equipos seleccionados:**")
            for equipo in st.session_state.filtro_equipo:
                st.write(f"- {equipo}")
        
        if st.session_state.filtro_usuario:
            st.write("**Usuarios seleccionados:**")
            for usuario in st.session_state.filtro_usuario:
                st.write(f"- {usuario}")
        
        if 'ETIQ. PENÚLTIMO TRAM.' in df.columns and st.session_state.filtro_etiq_penultimo:
            st.write("**ETIQ. PENÚLTIMO TRAM. seleccionados:**")
            for etiq in st.session_state.filtro_etiq_penultimo:
                st.write(f"- {etiq}")
        
        if 'ETIQ. ÚLTIMO TRAM.' in df.columns and st.session_state.filtro_etiq_ultimo:
            st.write("**ETIQ. ÚLTIMO TRAM. seleccionados:**")
            for etiq in st.session_state.filtro_etiq_ultimo:
                st.write(f"- {etiq}")

    # NUEVO: Opciones de ordenamiento y auto-filtros
    st.sidebar.markdown("---")
    st.sidebar.subheader("Opciones de Visualización")
    
    # Checkbox para ordenar por prioridad
    ordenar_prioridad = st.sidebar.checkbox("Ordenar por prioridad (RUE amarillo primero)", value=True, key="ordenar_prioridad_checkbox")

    # Aplicar ordenamiento si está activado - CORREGIDO
    if ordenar_prioridad:
        df_filtrado = ordenar_dataframe_por_prioridad_y_antiguedad(df_filtrado)
        st.sidebar.success("✅ Ordenado por prioridad")
    else:
        # Si no está activado, ordenar solo por antigüedad descendente
        columnas_antiguedad = [col for col in df_filtrado.columns if 'ANTIGÜEDAD' in col.upper() or 'DÍAS' in col.upper()]
        if columnas_antiguedad:
            columna_antiguedad = columnas_antiguedad[0]
            df_filtrado = df_filtrado.sort_values(columna_antiguedad, ascending=False)
            st.sidebar.info("📊 Ordenado solo por antigüedad")

    # Auto-filtros para mostrar solo filas con formato condicional
    st.sidebar.markdown("**Auto-filtros:**")

    # Usar keys únicos para los checkboxes y definir las variables
    mostrar_solo_amarillos = st.sidebar.checkbox("Solo RUE prioritarios", value=False, key="filtro_amarillos")
    mostrar_solo_rojos = st.sidebar.checkbox("Solo USUARIO-CSV discrepantes", value=False, key="filtro_rojos") 
    mostrar_solo_90_incdocu = st.sidebar.checkbox("Solo con 90 INCDOCU", value=False, key="filtro_90_incdocu")

    # Aplicar auto-filtros - VERSIÓN CORREGIDA
    if mostrar_solo_amarillos or mostrar_solo_rojos or mostrar_solo_90_incdocu:
        # Crear una COPIA para los filtros sin afectar el DataFrame principal
        df_filtrado_temp = df_filtrado.copy(deep=True)
        
        filtros_aplicados = []
        
        if mostrar_solo_amarillos:
            # Filtrar solo RUE amarillos usando la función corregida
            df_priorizado_temp = identificar_filas_prioritarias(df_filtrado_temp)
            mask_amarillo = df_priorizado_temp['_prioridad'] == 1
            if mask_amarillo.any():
                df_filtrado_temp = df_filtrado_temp[mask_amarillo]
                filtros_aplicados.append(f"RUE prioritarios: {mask_amarillo.sum()}")
        
        if mostrar_solo_rojos:
            # Filtrar solo USUARIO-CSV rojos
            if 'USUARIO' in df_filtrado_temp.columns and 'USUARIO-CSV' in df_filtrado_temp.columns:
                # Comparación segura
                usuario_principal = df_filtrado_temp['USUARIO'].astype(str).str.strip().fillna('')
                usuario_csv = df_filtrado_temp['USUARIO-CSV'].astype(str).str.strip().fillna('')
                mask_rojo = usuario_principal != usuario_csv
                
                if mask_rojo.any():
                    df_filtrado_temp = df_filtrado_temp[mask_rojo]
                    filtros_aplicados.append(f"USUARIO-CSV discrepantes: {mask_rojo.sum()}")
        
        if mostrar_solo_90_incdocu:
            # Filtrar solo expedientes con "ETIQ. PENÚLTIMO TRAM." = "90 INCDOCU"
            if 'ETIQ. PENÚLTIMO TRAM.' in df_filtrado_temp.columns:
                mask_90_incdocu = df_filtrado_temp['ETIQ. PENÚLTIMO TRAM.'] == "90 INCDOCU"
                
                if mask_90_incdocu.any():
                    df_filtrado_temp = df_filtrado_temp[mask_90_incdocu]
                    filtros_aplicados.append(f"Con 90 INCDOCU: {mask_90_incdocu.sum()}")
                else:
                    st.sidebar.warning("No hay expedientes con ETIQ. PENÚLTIMO TRAM. = '90 INCDOCU'")
            else:
                st.sidebar.warning("Columna 'ETIQ. PENÚLTIMO TRAM.' no disponible")
        
        # Solo actualizar si se aplicaron filtros
        if filtros_aplicados:
            df_filtrado = df_filtrado_temp
            st.sidebar.success(" | ".join(filtros_aplicados))

    # Mostrar estadísticas de los filtros aplicados
    st.sidebar.markdown("---")
    st.sidebar.subheader("Estadísticas")
    
    # Contar filas prioritarias
    df_priorizado = identificar_filas_prioritarias(df_filtrado)
    filas_amarillas = df_priorizado['_prioridad'].sum()
    filas_totales = len(df_filtrado)
    
    st.sidebar.write(f"Total filas: {filas_totales}")
    st.sidebar.write(f"RUE prioritarios: {filas_amarillas}")
    
    # Contar USUARIO-CSV rojos
    if 'USUARIO' in df_filtrado.columns and 'USUARIO-CSV' in df_filtrado.columns:
        filas_rojas = (df_filtrado['USUARIO'] != df_filtrado['USUARIO-CSV']).sum()
        st.sidebar.write(f"USUARIO-CSV discrepantes: {filas_rojas}")
    
    # Contar 90 INCDOCU (sustituye a DOCUM.INCORP.)
    if 'ETIQ. PENÚLTIMO TRAM.' in df_filtrado.columns:
        filas_90_incdocu = (df_filtrado['ETIQ. PENÚLTIMO TRAM.'] == "90 INCDOCU").sum()
        st.sidebar.write(f"Con 90 INCDOCU: {filas_90_incdocu}")

    # Gráficos Generales - CORREGIDOS: datos siempre frescos según filtros
    st.subheader("📈 Gráficos Generales")
    columnas_graficos = st.columns(3)
    graficos = [("EQUIPO", "Expedientes por equipo"), 
                ("USUARIO", "Expedientes por usuario"), 
                ("ESTADO", "Distribución por estado")]

    for i, (col, titulo) in enumerate(graficos):
        if col in df_filtrado.columns:
            # Calcular el conteo actual (siempre fresco según los filtros)
            conteo_actual = df_filtrado[col].value_counts().reset_index()
            conteo_actual.columns = [col, "Cantidad"]
            
            # Crear gráfico con datos actualizados (SIN CACHE)
            fig = crear_grafico_dinamico(conteo_actual, col, titulo)
            if fig:
                columnas_graficos[i].plotly_chart(fig, use_container_width=True)

    # NUEVOS GRÁFICOS PARA LAS ETIQUETAS
    col1, col2 = st.columns(2)
    with col1:
        if 'ETIQ. PENÚLTIMO TRAM.' in df_filtrado.columns:
            conteo_penultimo = df_filtrado['ETIQ. PENÚLTIMO TRAM.'].value_counts().reset_index()
            conteo_penultimo.columns = ['ETIQ. PENÚLTIMO TRAM.', 'Cantidad']
            fig_penultimo = crear_grafico_dinamico(conteo_penultimo, 'ETIQ. PENÚLTIMO TRAM.', 'Distribución por ETIQ. PENÚLTIMO TRAM.')
            if fig_penultimo:
                st.plotly_chart(fig_penultimo, use_container_width=False)

    with col2:
        if 'ETIQ. ÚLTIMO TRAM.' in df_filtrado.columns:
            conteo_ultimo = df_filtrado['ETIQ. ÚLTIMO TRAM.'].value_counts().reset_index()
            conteo_ultimo.columns = ['ETIQ. ÚLTIMO TRAM.', 'Cantidad']
            fig_ultimo = crear_grafico_dinamico(conteo_ultimo, 'ETIQ. ÚLTIMO TRAM.', 'Distribución por ETIQ. ÚLTIMO TRAM.')
            if fig_ultimo:
                st.plotly_chart(fig_ultimo, use_container_width=False)


    # Obtener información de la semana si está disponible
    num_semana_info, fecha_max_str_info, _ = obtener_info_semana_actual(df_filtrado)

    # Mostrar con Handsontable
    df_mostrar = mostrar_con_handsontable(
        df_filtrado
    )

    # Estadísticas generales
    st.markdown("---")
    st.subheader("📊 Estadísticas Generales")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        registros_mostrados = f"{len(df_mostrar):,}".replace(",", ".")
        registros_totales = f"{len(df):,}".replace(",", ".")
        st.metric("Registros mostrados", f"{registros_mostrados}/{registros_totales}")

    with col2:
        # Contar RUE amarillos - CON NUEVAS CONDICIONES
        mask_amarillo = pd.Series(False, index=df_filtrado.index)
        for idx, row in df_filtrado.iterrows():
            try:
                etiq_penultimo = row.get('ETIQ. PENÚLTIMO TRAM.', '')
                fecha_notif = row.get('FECHA NOTIFICACIÓN', None)
                docum_incorp = row.get('DOCUM.INCORP.', '')
                
                # CONDICIÓN 1: "80 PROPRES" con fecha límite superada
                if (str(etiq_penultimo).strip() == "80 PROPRES" and 
                    pd.notna(fecha_notif) and 
                    isinstance(fecha_notif, (pd.Timestamp, datetime))):
                    
                    fecha_limite = fecha_notif + timedelta(days=23)
                    if datetime.now() > fecha_limite:
                        mask_amarillo[idx] = True
                
                # NUEVA CONDICIÓN 2: "50 REQUERIR" con fecha límite superada
                elif (str(etiq_penultimo).strip() == "50 REQUERIR" and 
                    pd.notna(fecha_notif) and 
                    isinstance(fecha_notif, (pd.Timestamp, datetime))):
                    
                    fecha_limite = fecha_notif + timedelta(days=23)
                    if datetime.now() > fecha_limite:
                        mask_amarillo[idx] = True
                
                # NUEVA CONDICIÓN 3: "70 ALEGACI" o "60 CONTESTA"
                elif str(etiq_penultimo).strip() in ["70 ALEGACI", "60 CONTESTA"]:
                    mask_amarillo[idx] = True
                
                # NUEVA CONDICIÓN 4: DOCUM.INCORP. no vacío Y distinto de "SOLICITUD" o "REITERA SOLICITUD"
                elif (pd.notna(docum_incorp) and 
                    str(docum_incorp).strip() != '' and
                    str(docum_incorp).strip().upper() not in ["SOLICITUD", "REITERA SOLICITUD"]):
                    mask_amarillo[idx] = True
                    
            except:
                pass
        st.metric("RUE prioritarios", f"{mask_amarillo.sum():,}".replace(",", "."))

    with col3:
        # Contar USUARIO-CSV rojos
        if 'USUARIO' in df_filtrado.columns and 'USUARIO-CSV' in df_filtrado.columns:
            mask_rojo = (df_filtrado['USUARIO'] != df_filtrado['USUARIO-CSV']).sum()
            st.metric("Discrepancias", f"{mask_rojo:,}".replace(",", "."))
        else:
            st.metric("Discrepancias", "N/A")

    with col4:
        # Contar 90 INCDOCU (sustituye a DOCUM.INCORP.)
        if 'ETIQ. PENÚLTIMO TRAM.' in df_filtrado.columns:
            mask_90_incdocu = (df_filtrado['ETIQ. PENÚLTIMO TRAM.'] == "90 INCDOCU").sum()
            st.metric("Con 90 INCDOCU", f"{mask_90_incdocu:,}".replace(",", "."))
        else:
            st.metric("Con 90 INCDOCU", "N/A")

    # NUEVA SECCIÓN: GESTIÓN DE DOCUMENTACIÓN INCORPORADA
    st.markdown("---")
    st.header("📄 Gestión de Documentación Incorporada")

    if datos_documentos:
        # Filtro para expedientes con ETIQ. PENÚLTIMO TRAM. = "90 INCDOCU"
        df_incdocu = df_filtrado[
            df_filtrado['ETIQ. PENÚLTIMO TRAM.'] == "90 INCDOCU"
        ].copy()
        
        if df_incdocu.empty:
            st.info("ℹ️ No hay expedientes con ETIQ. PENÚLTIMO TRAM. = '90 INCDOCU' en los filtros actuales")
        else:
            st.success(f"📋 Encontrados {len(df_incdocu)} expedientes con '90 INCDOCU'")
            
            # Obtener opciones del desplegable
            opciones_docu = datos_documentos['opciones']
            opciones_combo = [""] + opciones_docu  # Añadir opción vacía
            
            # Crear interfaz para editar la documentación
            st.subheader("Editar Documentación Incorporada")
            
            # Diccionario temporal para almacenar cambios (si no existe)
            if 'cambios_documentacion_temp' not in st.session_state:
                st.session_state.cambios_documentacion_temp = {}
            
            # Mostrar tabla editable
            cambios_realizados = False
            for idx, row in df_incdocu.iterrows():
                rue = row['RUE']
                docum_actual = row.get('DOCUM.INCORP.', '')
                
                col1, col2 = st.columns([2, 3])  # Reducir a 2 columnas
                
                with col1:
                    st.write(f"**RUE:** {rue}")
                
                with col2:
                    # Selectbox para elegir documentación
                    clave_docum = f"docum_{rue}"
                    nueva_docum = st.selectbox(
                        "Documentación:",
                        options=opciones_combo,
                        index=opciones_combo.index(docum_actual) if docum_actual in opciones_combo else 0,
                        key=clave_docum,
                        label_visibility="collapsed"
                    )
                    
                    # Almacenar cambio temporalmente (sin grabar aún)
                    if nueva_docum != docum_actual:
                        st.session_state.cambios_documentacion_temp[rue] = nueva_docum
                        cambios_realizados = True
                    elif rue in st.session_state.cambios_documentacion_temp:
                        # Si vuelve al valor original, eliminar del diccionario
                        del st.session_state.cambios_documentacion_temp[rue]
            
            # Botón único para guardar todos los cambios
            st.markdown("---")
            col1, col2 = st.columns([3, 1])
            
            with col2:
                if st.button("💾 Guardar Todos los Cambios", type="primary", key="guardar_todos_documentos"):
                    if not st.session_state.cambios_documentacion_temp:
                        st.warning("⚠️ No hay cambios para guardar")
                    else:
                        with st.spinner("Guardando cambios..."):
                            # Obtener DataFrame combinado
                            df_combinado = st.session_state["df_combinado"]
                            
                            # Aplicar todos los cambios al DataFrame
                            for rue, nueva_docum in st.session_state.cambios_documentacion_temp.items():
                                df_combinado.loc[df_combinado['RUE'] == rue, 'DOCUM.INCORP.'] = nueva_docum
                            
                            # Actualizar session_state
                            st.session_state["df_combinado"] = df_combinado
                            
                            # Filtrar para guardar en archivo
                            df_documentos_actualizado = df_combinado[
                                (df_combinado['ETIQ. PENÚLTIMO TRAM.'] == "90 INCDOCU") &
                                (df_combinado['DOCUM.INCORP.'].notna()) &
                                (df_combinado['DOCUM.INCORP.'] != '')
                            ][['RUE', 'DOCUM.INCORP.']].copy()
                            
                            # Guardar en archivo
                            contenido_actualizado = guardar_documentos_actualizados(
                                datos_documentos['archivo'], 
                                df_documentos_actualizado
                            )
                            
                            if contenido_actualizado:
                                st.session_state.documentos_actualizados = contenido_actualizado
                                st.session_state.mostrar_descarga = True
                                
                                # Mostrar resumen de cambios
                                st.success(f"✅ {len(st.session_state.cambios_documentacion_temp)} cambios guardados correctamente")
                                
                                # Limpiar cambios temporales
                                st.session_state.cambios_documentacion_temp = {}
                                
                                # Actualizar cache
                                st.cache_data.clear()
                                
                                # Rerun para actualizar la vista
                                st.rerun()
                            else:
                                st.error("❌ Error al guardar el archivo DOCUMENTOS.xlsx")
            
            # Mostrar botón de descarga si hay archivo actualizado
            if st.session_state.get('mostrar_descarga', False) and st.session_state.get('documentos_actualizados'):
                st.markdown("---")
                st.subheader("📥 Descargar Archivo Actualizado")
                
                # Generar nombre de archivo con timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nombre_archivo = f"DOCUMENTOS_actualizado_{timestamp}.xlsx"
                
                st.download_button(
                    label="⬇️ Descargar DOCUMENTOS.xlsx actualizado",
                    data=st.session_state.documentos_actualizados,
                    file_name=nombre_archivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="descarga_documentos"
                )
                
                # Opción para continuar sin descargar
                if st.button("Continuar sin descargar", key="continuar_sin_descargar"):
                    st.session_state.mostrar_descarga = False
                    st.rerun()
    else:
        st.warning("⚠️ Carga el archivo DOCUMENTOS.xlsx para gestionar la documentación incorporada")

# =============================================
# PÁGINA 3: INDICADORES CLAVE (KPI)
# =============================================
elif eleccion == "Indicadores clave (KPI)":
    st.header("📊 Indicadores clave (KPI)")
    
    if "df_combinado" not in st.session_state:
        st.warning("⚠️ Primero carga los archivos en la sección 'Carga de Archivos'")
        st.stop()
    
    # Usar df_combinado en lugar de df
    df = st.session_state["df_combinado"]
    
    # Obtener fecha de referencia para cálculos
    # columna_fecha = df.columns[13]
    # df[columna_fecha] = pd.to_datetime(df[columna_fecha], errors='coerce')
    # fecha_max = df[columna_fecha].max()     ya está definida anteriormente
    
    if pd.isna(fecha_max):
        st.error("No se pudo encontrar la fecha máxima en los datos")
        st.stop()
    
    # Crear rango de semanas disponibles
    fecha_inicio = pd.to_datetime("2022-11-01")
    semanas_disponibles = pd.date_range(
        start=fecha_inicio,
        end=fecha_max,
        freq='W-FRI'
    ).tolist()
    
    if not semanas_disponibles:
        st.error("No hay semanas disponibles para mostrar")
        st.stop()
    
    # Inicialización del estado para KPI
    if 'kpi_semana_index' not in st.session_state:
        st.session_state.kpi_semana_index = len(semanas_disponibles) - 1
    
    # Obtener la semana seleccionada actual
    semana_seleccionada = semanas_disponibles[st.session_state.kpi_semana_index]
    num_semana_seleccionada = ((semana_seleccionada - FECHA_REFERENCIA).days) // 7 + 1
    fecha_str = semana_seleccionada.strftime('%d/%m/%Y')
    
    # Selector de semana en el área principal
    st.markdown("---")
    st.header("🗓️ Selector de Semana")
    
    # Crear etiquetas formateadas para el slider
    opciones_slider = []
    for i, fecha in enumerate(semanas_disponibles):
        num_semana = ((fecha - FECHA_REFERENCIA).days) // 7 + 1
        fecha_str_opcion = fecha.strftime('%d/%m/%Y')
        opciones_slider.append(f"Semana {num_semana} ({fecha_str_opcion})")
    
    # Slider
    semana_index_slider = st.select_slider(
        "Selecciona la semana:",
        options=list(range(len(semanas_disponibles))),
        value=st.session_state.kpi_semana_index,
        format_func=lambda x: opciones_slider[x]
    )
    
    # Actualizar el índice si el slider cambió
    if semana_index_slider != st.session_state.kpi_semana_index:
        st.session_state.kpi_semana_index = semana_index_slider
        st.rerun()
    
    # RECALCULAR después de posibles cambios del slider
    semana_seleccionada = semanas_disponibles[st.session_state.kpi_semana_index]
    num_semana_seleccionada = ((semana_seleccionada - FECHA_REFERENCIA).days) // 7 + 1
    fecha_str = semana_seleccionada.strftime('%d/%m/%Y')
    
    # Mostrar información de la semana seleccionada
    st.info(f"**Semana seleccionada:** {fecha_str} (Semana {num_semana_seleccionada})")
    
    # Sidebar con botones de navegación
    with st.sidebar:
        st.header("🗓️ Navegación por Semanas")
        
        st.write(f"**Semana actual:**")
        st.write(f"{fecha_str}")
        st.write(f"(Semana {num_semana_seleccionada})")
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("◀️ Anterior", use_container_width=True, key="btn_anterior_kpi"):
                nuevo_indice = st.session_state.kpi_semana_index - 1
                if nuevo_indice >= 0:
                    st.session_state.kpi_semana_index = nuevo_indice
                    st.rerun()
        
        with col2:
            if st.button("Siguiente ▶️", use_container_width=True, key="btn_siguiente_kpi"):
                nuevo_indice = st.session_state.kpi_semana_index + 1
                if nuevo_indice < len(semanas_disponibles):
                    st.session_state.kpi_semana_index = nuevo_indice
                    st.rerun()
        
        st.write(f"**Posición:** {st.session_state.kpi_semana_index + 1} de {len(semanas_disponibles)}")
        
        if st.button("📅 Ir a semana actual", use_container_width=True, key="btn_actual_kpi"):
            st.session_state.kpi_semana_index = len(semanas_disponibles) - 1
            st.rerun()

    # Calcular KPIs para todas las semanas (usando cache)
    df_kpis_semanales = calcular_kpis_todas_semanas_optimizado(df, semanas_disponibles, FECHA_REFERENCIA, fecha_max)

    # Función para mostrar los nuevos KPIs principales
    def mostrar_kpis_principales(_df_kpis, _semana_seleccionada, _num_semana):
        kpis_semana = _df_kpis[_df_kpis['semana_numero'] == _num_semana].iloc[0]
        
        fecha_str = _semana_seleccionada.strftime('%d/%m/%Y')
        st.header(f"📊 KPIs de la Semana: {fecha_str} (Semana {_num_semana})")
        
        # Mostrar el rango correcto de la semana
        inicio_str = kpis_semana['inicio_semana'].strftime('%d/%m/%Y')
        fin_str = kpis_semana['fin_semana'].strftime('%d/%m/%Y')
        dias_semana = kpis_semana['dias_semana']
        es_actual = kpis_semana['es_semana_actual']
        
        if es_actual:
            st.info(f"**📅 Período de la semana (ACTUAL):** {inicio_str} (viernes) a {fin_str} (viernes) - {dias_semana} días")
            st.warning("ℹ️ **Semana actual:** Incluye todos los expedientes hasta la fecha de actualización")
        else:
            st.info(f"**📅 Período de la semana:** {inicio_str} (viernes) a {fin_str} (jueves) - {dias_semana} días")
        
        # PRIMERA FILA DE KPIs - DOS COLUMNAS (SEMANAL vs TOTALES)
        st.subheader("📈 KPIs Principales")
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Semanal")
            st.metric(
                label="💰 Nuevos Expedientes",
                value=f"{int(kpis_semana['nuevos_expedientes']):,}".replace(",", ".")
            )
            st.metric(
                label="📤 Expedientes Despachados",
                value=f"{int(kpis_semana['despachados_semana']):,}".replace(",", ".")
            )
            st.metric(
                label="🎯 Coef. Absorción (Desp/Nuevos)",
                value=f"{kpis_semana['c_abs_despachados_sem']:.2f}%".replace(".", ",")
            )
            st.metric(
                label="🛒 Expedientes Cerrados",
                value=f"{int(kpis_semana['expedientes_cerrados']):,}".replace(",", ".")
            )
            st.metric(
                label="📊 Coef. Absorción (Cer/Asig)",
                value=f"{kpis_semana['c_abs_cerrados_sem']:.2f}%".replace(".", ",")
            )
            st.metric(
                label="👥 Expedientes Abiertos",
                value=f"{int(kpis_semana['total_abiertos']):,}".replace(",", ".")
            )
            st.metric(
                label="👥 Expedientes Abiertos no Despachados",
                value=f"{int(kpis_semana['total_abiertos_no_despachados']):,}".replace(",", ".")
            )

        with col2:
            st.markdown("### Totales (desde 01/11/2022)")
            st.metric(
                label="💰 Nuevos Expedientes",
                value=f"{int(kpis_semana['nuevos_expedientes_totales']):,}".replace(",", ".")
            )
            st.metric(
                label="📤 Expedientes Despachados",
                value=f"{int(kpis_semana['despachados_totales']):,}".replace(",", ".")
            )
            st.metric(
                label="🎯 Coef. Absorción (Desp/Nuevos)",
                value=f"{kpis_semana['c_abs_despachados_tot']:.2f}%".replace(".", ",")
            )
            st.metric(
                label="🛒 Expedientes Cerrados",
                value=f"{int(kpis_semana['expedientes_cerrados_totales']):,}".replace(",", ".")
            )
            # No hay totales para abiertos (es un snapshot)
            st.metric(
                label="📊 Coef. Absorción (Cer/Asig)",
                value=f"{kpis_semana['c_abs_cerrados_tot']:.2f}%".replace(".", ",")
            )
            st.metric(
                label="👥 Total Expedientes Asignados",
                value=f"{int(kpis_semana['despachados_totales'] + kpis_semana['total_abiertos_no_despachados']):,}".replace(",", ".")
            )
            st.metric(
                label="👥 Expedientes Rehabilitados (última semana)",
                value=f"{int(kpis_semana['total_rehabilitados']):,}".replace(",", ".")
            )

        st.markdown("---")
        
        # SEGUNDA FILA - EXPEDIENTES ESPECIALES
        st.subheader("📋 Expedientes con 029, 033, PRE, RSL, pendiente de firma, de decisión o de completar trámite (ÚLTIMA SEMANA)")
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric(
                label="🔍 Expedientes Especiales",
                value=f"{int(kpis_semana['expedientes_especiales']):,}".replace(",", ".")
            )
        
        with col2:
            st.metric(
                label="📈 Porcentaje sobre Abiertos",
                value=f"{kpis_semana['porcentaje_especiales']:.2f}%".replace(".", ",")
            )
        
        st.markdown("---")
        
        # TERCERA FILA - TIEMPOS DE TRAMITACIÓN (TRES COLUMNAS)
        st.subheader("⏱️ Tiempos de Tramitación (en días)")
        
        # Expedientes Despachados
        st.markdown("#### 📤 Expedientes Despachados")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                label="📊 Tiempo Medio",
                value=f"{kpis_semana['tiempo_medio_despachados']:.0f}"
            )
        
        with col2:
            st.metric(
                label="🎯 Percentil 90",
                value=f"{kpis_semana['percentil_90_despachados']:.0f}"
            )
        
        with col3:
            col3a, col3b = st.columns(2)
            with col3a:
                st.metric(
                    label="≤180 días",
                    value=f"{kpis_semana['percentil_180_despachados']:.2f}%".replace(".", ",")
                )
            with col3b:
                st.metric(
                    label="≤120 días",
                    value=f"{kpis_semana['percentil_120_despachados']:.2f}%".replace(".", ",")
                )
        
        # Expedientes Cerrados
        st.markdown("#### 📤 Expedientes Cerrados")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                label="📊 Tiempo Medio",
                value=f"{kpis_semana['tiempo_medio_cerrados']:.0f}"
            )
        
        with col2:
            st.metric(
                label="🎯 Percentil 90",
                value=f"{kpis_semana['percentil_90_cerrados']:.0f}"
            )
        
        with col3:
            col3a, col3b = st.columns(2)
            with col3a:
                st.metric(
                    label="≤180 días",
                    value=f"{kpis_semana['percentil_180_cerrados']:.2f}%".replace(".", ",")
                )
            with col3b:
                st.metric(
                    label="≤120 días",
                    value=f"{kpis_semana['percentil_120_cerrados']:.2f}%".replace(".", ",")
                )
        
        # Expedientes Abiertos
        st.markdown("#### 📤 Expedientes Abiertos")
        col1, col2, col3 = st.columns(3)
        
        #with col1:
        #    st.metric(
        #        label="📊 Tiempo Actual",
        #        value="-"  # No hay tiempo medio para abiertos
        #    )
        
        with col2:
            st.metric(
                label="🎯 Percentil 90",
                value=f"{kpis_semana['percentil_90_abiertos']:.0f}"
            )
        
        with col3:
            col3a, col3b = st.columns(2)
            with col3a:
                st.metric(
                    label="≤180 días",
                    value=f"{kpis_semana['percentil_180_abiertos']:.2f}%".replace(".", ",")
                )
            with col3b:
                st.metric(
                    label="≤120 días",
                    value=f"{kpis_semana['percentil_120_abiertos']:.2f}%".replace(".", ",")
                )

        # Expedientes Abiertos no Despachados
        st.markdown("#### 📤 Expedientes Abiertos no Despachados")
        col1, col2, col3 = st.columns(3)
        
        #with col1:
        #    st.metric(
        #        label="📊 Tiempo Actual",
        #        value="-"  # No hay tiempo medio para abiertos
        #    )
        
        with col2:
            st.metric(
                label="🎯 Percentil 90",
                value=f"{kpis_semana['percentil_90_abiertos_no_despachados']:.0f}"
            )
        
        with col3:
            col3a, col3b = st.columns(2)
            with col3a:
                st.metric(
                    label="≤180 días",
                    value=f"{kpis_semana['percentil_180_abiertos_no_despachados']:.2f}%".replace(".", ",")
                )
            with col3b:
                st.metric(
                    label="≤120 días",
                    value=f"{kpis_semana['percentil_120_abiertos_no_despachados']:.2f}%".replace(".", ",")
                )


    # Mostrar dashboard principal
    mostrar_kpis_principales(df_kpis_semanales, semana_seleccionada, num_semana_seleccionada)

    # GRÁFICO DE EVOLUCIÓN TEMPORAL (ACTUALIZADO) - CORREGIDO
    st.markdown("---")
    st.subheader("📈 Evolución de KPIs Principales y Porcentajes")

    datos_grafico = df_kpis_semanales.copy()
    col1, col2, col3 = st.columns(3)
    num_semana_seleccionada = ((semana_seleccionada - FECHA_REFERENCIA).days) // 7 + 1

    # --- Gráfico 1: KPI principales (sin "Abiertos") ---
    with col1:
        fig1 = px.line(
            datos_grafico,
            x='semana_numero',
            y=['nuevos_expedientes', 'despachados_semana', 'expedientes_cerrados'],
            title='Evolución de Expedientes (Nuevos, Despachados, Cerrados)',
            labels={'semana_numero': 'Semana', 'value': 'Cantidad', 'variable': 'KPI'},
            color_discrete_map={
                'nuevos_expedientes': '#1f77b4',
                'despachados_semana': '#ff7f0e',
                'expedientes_cerrados': '#2ca02c'
            }
        )
        fig1.update_layout(
            height=400,
            hovermode="x unified",
            legend_title="KPI",
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
        )
        # Línea discontinua vertical indicando la semana seleccionada
        fig1.add_vline(
            x=num_semana_seleccionada,
            line_dash="dash",
            line_color="red",
            annotation_text=f"Semana {num_semana_seleccionada}",
            annotation_position="top left"
        )
        st.plotly_chart(fig1, use_container_width=True)

    # --- Gráfico 2: Solo "Expedientes Abiertos" ---
    with col2:
        fig2 = px.line(
            datos_grafico,
            x='semana_numero',
            y=['total_abiertos'],
            title='Evolución de Expedientes Abiertos',
            labels={'semana_numero': 'Semana', 'value': 'Cantidad', 'variable': 'KPI'},
            color_discrete_map={'total_abiertos': '#d62728'}
        )
        fig2.update_layout(
            height=400,
            hovermode="x unified",
            legend_title="KPI",
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
        )
        # Línea discontinua vertical indicando la semana seleccionada
        fig2.add_vline(
            x=num_semana_seleccionada,
            line_dash="dash",
            line_color="red",
            annotation_text=f"Semana {num_semana_seleccionada}",
            annotation_position="top left"
        )
        st.plotly_chart(fig2, use_container_width=True)

    # --- Gráfico 3: Porcentajes (4) ---
    with col3:
        fig3 = px.line(
            datos_grafico,
            x='semana_numero',
            y=[
                'c_abs_despachados_sem', 'c_abs_despachados_tot',
                'c_abs_cerrados_sem', 'c_abs_cerrados_tot'
            ],
            title='Evolución de Coeficientes de Absorción (%)',
            labels={'semana_numero': 'Semana', 'value': 'Porcentaje (%)', 'variable': 'Indicador'},
            color_discrete_map={
                'c_abs_despachados_sem': '#9467bd',
                'c_abs_despachados_tot': '#c5b0d5',
                'c_abs_cerrados_sem': '#8c564b',
                'c_abs_cerrados_tot': '#c49c94'
            }
        )
        fig3.update_layout(
            height=400,
            hovermode="x unified",
            legend_title="Indicador",
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
        )
        # Línea discontinua vertical indicando la semana seleccionada
        fig3.add_vline(
            x=num_semana_seleccionada,
            line_dash="dash",
            line_color="red",
            annotation_text=f"Semana {num_semana_seleccionada}",
            annotation_position="top left"
        )
        st.plotly_chart(fig3, use_container_width=True)

    # -------------------------------------------------------------
    # SEGUNDO BLOQUE: TIEMPOS DE TRAMITACIÓN
    # -------------------------------------------------------------
    st.markdown("---")
    st.subheader("⏱️ Tiempos de Tramitación")

    col1, col2 = st.columns(2)

    # --- Gráfico 1: Tiempos medios y percentiles 90 ---
    with col1:
        fig_tiempo = px.line(
            datos_grafico,
            x='semana_numero',
            y=[
                'tiempo_medio_despachados', 'tiempo_medio_cerrados',
                'percentil_90_despachados', 'percentil_90_cerrados'
            ],
            title='Tiempos Medios y Percentiles 90 (días)',
            labels={'semana_numero': 'Semana', 'value': 'Días', 'variable': 'Indicador'},
            color_discrete_map={
                'tiempo_medio_despachados': '#ff7f0e',
                'tiempo_medio_cerrados': '#2ca02c',
                'percentil_90_despachados': '#ffbb78',
                'percentil_90_cerrados': '#98df8a'
            }
        )
        fig_tiempo.update_layout(
            height=400,
            hovermode="x unified",
            legend_title="Indicador",
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
        )
        # Línea discontinua vertical indicando la semana seleccionada
        fig_tiempo.add_vline(
            x=num_semana_seleccionada,
            line_dash="dash",
            line_color="red",
            annotation_text=f"Semana {num_semana_seleccionada}",
            annotation_position="top left"
        )
        st.plotly_chart(fig_tiempo, use_container_width=True)

    # --- Gráfico 2: Porcentajes 120/180 ---
    with col2:
        fig_percentiles = px.line(
            datos_grafico,
            x='semana_numero',
            y=[
                'percentil_180_despachados', 'percentil_120_despachados',
                'percentil_180_cerrados', 'percentil_120_cerrados'
            ],
            title='Porcentaje de Expedientes ≤120 y ≤180 días (%)',
            labels={'semana_numero': 'Semana', 'value': 'Porcentaje (%)', 'variable': 'Indicador'},
            color_discrete_map={
                'percentil_180_despachados': '#ff7f0e',
                'percentil_120_despachados': '#ffddaa',
                'percentil_180_cerrados': '#2ca02c',
                'percentil_120_cerrados': '#98df8a'
            }
        )
        fig_percentiles.update_layout(
            height=400,
            hovermode="x unified",
            legend_title="Indicador",
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5)
        )
        # Línea discontinua vertical indicando la semana seleccionada
        fig_percentiles.add_vline(
            x=num_semana_seleccionada,
            line_dash="dash",
            line_color="red",
            annotation_text=f"Semana {num_semana_seleccionada}",
            annotation_position="top left"
        )
        st.plotly_chart(fig_percentiles, use_container_width=True)

# =============================================
# PÁGINA 4: ANÁLISIS DEL RENDIMIENTO - MODIFICADO CON HANDSONTABLE
# =============================================
elif eleccion == "Análisis del Rendimiento":
    st.header("📈 Análisis del Rendimiento")
    
    if "df_combinado" not in st.session_state:
        st.warning("⚠️ Primero carga los archivos en la sección 'Carga de Archivos'")
        st.stop()
    
    # Usar df_combinado en lugar de df
    df = st.session_state["df_combinado"]
    df_usuarios = st.session_state.get("df_usuarios", None)
    
    # Verificar si tenemos el archivo de usuarios
    if df_usuarios is None or df_usuarios.empty:
        st.error("❌ No se ha cargado el archivo USUARIOS.xlsx o está vacío")
        st.info("💡 Necesitas cargar el archivo USUARIOS.xlsx en la sección 'Carga de Archivos' para usar esta funcionalidad")
        st.stop()
    
    # Obtener fecha máxima para cálculos
    if pd.isna(fecha_max):
        st.error("No se pudo encontrar la fecha máxima en los datos")
        st.stop()
    
    fecha_max_str = fecha_max.strftime("%d/%m/%Y")
    
    st.info(f"📅 **Fecha de referencia para cálculos:** {fecha_max_str}")
    
    # =============================================
    # CÁLCULOS PREVIOS - VERSIÓN MODIFICADA PARA INCLUIR POTENCIAL ANUAL
    # =============================================
    
    # Calcular datos de rendimiento (AGRUPDOS POR USUARIO)
    with st.spinner("📊 Calculando indicadores de rendimiento (agrupados por usuario)..."):
        df_rendimiento = calcular_rendimiento_usuarios_agrupado(df, df_usuarios, fecha_max)
    
    if df_rendimiento.empty:
        st.warning("⚠️ No se encontraron datos de rendimiento para mostrar")
        st.info("💡 Esto puede deberse a que no hay expedientes despachados o no coinciden los nombres de usuarios entre los archivos")
        
        # Mostrar información de depuración
        with st.expander("🔍 Información de depuración"):
            st.write("**Usuarios en RECTAUTO:**", df['USUARIO'].dropna().unique().tolist() if 'USUARIO' in df.columns else "Columna USUARIO no encontrada")
            if df_usuarios is not None:
                columna_usuario = None
                for col in ['USUARIOS', 'USUARIO', 'NOMBRE']:
                    if col in df_usuarios.columns:
                        columna_usuario = col
                        break
                if columna_usuario:
                    st.write(f"**Usuarios en USUARIOS.xlsx ({columna_usuario}):**", df_usuarios[columna_usuario].dropna().unique().tolist())
        
        st.stop()
    
    # =============================================
    # FILTROS INTERCONECTADOS EN SIDEBAR - MODIFICADOS PARA USUARIOS ÚNICOS
    # =============================================
    
    st.sidebar.header("🔍 Filtros de Rendimiento")
    
    # Inicializar variables de sesión para filtros
    if 'filtro_estado_rendimiento' not in st.session_state:
        st.session_state.filtro_estado_rendimiento = ['ACTIVO'] if 'ACTIVO' in df_rendimiento['ESTADO'].values else []
    
    if 'filtro_equipo_rendimiento' not in st.session_state:
        st.session_state.filtro_equipo_rendimiento = []
    
    if 'filtro_usuario_rendimiento' not in st.session_state:
        st.session_state.filtro_usuario_rendimiento = []
    
    # Botón para resetear filtros
    if st.sidebar.button("🔄 Resetear filtros", use_container_width=True, key="reset_rendimiento"):
        st.session_state.filtro_estado_rendimiento = []
        st.session_state.filtro_equipo_rendimiento = []
        st.session_state.filtro_usuario_rendimiento = []
        st.rerun()
    
    # 1. Aplicar filtros secuencialmente para calcular opciones disponibles
    df_filtrado_temp = df_rendimiento.copy()
    
    # Aplicar filtro de ESTADO primero
    if st.session_state.filtro_estado_rendimiento:
        df_filtrado_temp = df_filtrado_temp[df_filtrado_temp['ESTADO'].isin(st.session_state.filtro_estado_rendimiento)]
    
    # MODIFICACIÓN: Para filtrar por EQUIPO, necesitamos buscar en la columna EQUIPOS (que contiene strings)
    equipos_disponibles = []
    if not df_filtrado_temp.empty:
        # Extraer todos los equipos únicos de la columna EQUIPOS
        todos_equipos = set()
        for equipos_str in df_filtrado_temp['EQUIPOS'].dropna():
            # Dividir por comas y limpiar
            equipos_lista = [eq.strip() for eq in str(equipos_str).split(',')]
            todos_equipos.update(equipos_lista)
        equipos_disponibles = sorted(todos_equipos)
    
    # Aplicar filtro de EQUIPO (búsqueda en string)
    if st.session_state.filtro_equipo_rendimiento:
        # Filtrar usuarios que tengan al menos uno de los equipos seleccionados
        mask_equipo = pd.Series(False, index=df_filtrado_temp.index)
        for i, row in df_filtrado_temp.iterrows():
            equipos_usuario = str(row['EQUIPOS']).split(',') if pd.notna(row['EQUIPOS']) else []
            equipos_usuario = [eq.strip() for eq in equipos_usuario]
            # Verificar si hay intersección entre equipos del usuario y equipos seleccionados
            if any(eq in equipos_usuario for eq in st.session_state.filtro_equipo_rendimiento):
                mask_equipo[i] = True
        df_filtrado_temp = df_filtrado_temp[mask_equipo]
    
    # Calcular USUARIOS disponibles basados en filtros anteriores
    usuarios_disponibles = sorted(df_filtrado_temp['USUARIO'].dropna().unique())
    
    # 2. Crear widgets de filtro con opciones actualizadas
    st.sidebar.markdown("---")
    
    # FILTRO DE ESTADO
    opciones_estado = sorted(df_rendimiento['ESTADO'].dropna().unique())
    estado_sel = st.sidebar.multiselect(
        "🔘 Estado:",
        options=opciones_estado,
        default=st.session_state.filtro_estado_rendimiento,
        key='filtro_estado_rendimiento_selector'
    )
    
    # FILTRO DE EQUIPO (ahora busca en columna EQUIPOS)
    equipo_sel = st.sidebar.multiselect(
        "👥 Equipo:",
        options=equipos_disponibles,
        default=st.session_state.filtro_equipo_rendimiento,
        key='filtro_equipo_rendimiento_selector'
    )
    
    # FILTRO DE USUARIO
    usuario_sel = st.sidebar.multiselect(
        "👤 Usuario:",
        options=usuarios_disponibles,
        default=st.session_state.filtro_usuario_rendimiento,
        key='filtro_usuario_rendimiento_selector'
    )
    
    # 3. Actualizar session_state cuando cambian los filtros
    if estado_sel != st.session_state.filtro_estado_rendimiento:
        st.session_state.filtro_estado_rendimiento = estado_sel
        st.session_state.filtro_equipo_rendimiento = []
        st.session_state.filtro_usuario_rendimiento = []
        st.rerun()
    
    if equipo_sel != st.session_state.filtro_equipo_rendimiento:
        st.session_state.filtro_equipo_rendimiento = equipo_sel
        st.session_state.filtro_usuario_rendimiento = []
        st.rerun()
    
    if usuario_sel != st.session_state.filtro_usuario_rendimiento:
        st.session_state.filtro_usuario_rendimiento = usuario_sel
        st.rerun()
    
    # 4. Aplicar filtros finales al DataFrame principal
    df_filtrado = df_rendimiento.copy()
    
    if st.session_state.filtro_estado_rendimiento:
        df_filtrado = df_filtrado[df_filtrado['ESTADO'].isin(st.session_state.filtro_estado_rendimiento)]
    
    # Aplicar filtro de EQUIPO (búsqueda en columna EQUIPOS)
    if st.session_state.filtro_equipo_rendimiento:
        mask_equipo = pd.Series(False, index=df_filtrado.index)
        for i, row in df_filtrado.iterrows():
            equipos_usuario = str(row['EQUIPOS']).split(',') if pd.notna(row['EQUIPOS']) else []
            equipos_usuario = [eq.strip() for eq in equipos_usuario]
            if any(eq in equipos_usuario for eq in st.session_state.filtro_equipo_rendimiento):
                mask_equipo[i] = True
        df_filtrado = df_filtrado[mask_equipo]
    
    if st.session_state.filtro_usuario_rendimiento:
        df_filtrado = df_filtrado[df_filtrado['USUARIO'].isin(st.session_state.filtro_usuario_rendimiento)]
    
    # Mostrar resumen de filtros
    st.sidebar.markdown("---")
    st.sidebar.subheader("📊 Resumen Filtros")
    
    if st.session_state.filtro_estado_rendimiento:
        st.sidebar.write(f"**Estados:** {len(st.session_state.filtro_estado_rendimiento)}")
    
    if st.session_state.filtro_equipo_rendimiento:
        st.sidebar.write(f"**Equipos:** {len(st.session_state.filtro_equipo_rendimiento)}")
    
    if st.session_state.filtro_usuario_rendimiento:
        st.sidebar.write(f"**Usuarios:** {len(st.session_state.filtro_usuario_rendimiento)}")
    
    st.sidebar.write(f"**Usuarios mostrados:** {len(df_filtrado):,}".replace(",", "."))
    
    # =============================================
    # CÁLCULO DE TOTALES AGRUPADOS (POR USUARIO)
    # =============================================
    
    def calcular_totales_agrupados_usuarios(df_agrupar):
        """Calcula los totales agrupados correctamente para usuarios únicos"""
        if df_agrupar.empty:
            return None
            
        # Sumas directas
        total_expedientes = df_agrupar['EXPEDIENTES_DESPACHADOS'].sum()
        total_semanas = df_agrupar['SEMANAS_EFECTIVAS'].sum()
        
        # Calcular rendimiento total correcto
        rendimiento_total_agrupado = total_expedientes / total_semanas if total_semanas > 0 else 0
        
        # Para los rendimientos por período, calcular la media
        rendimiento_anual_agrupado = df_agrupar['RENDIMIENTO_ANUAL'].mean()
        potencial_anual_agrupado = df_agrupar['POTENCIAL_ANUAL'].mean()  # NUEVO
        rendimiento_trimestral_agrupado = df_agrupar['RENDIMIENTO_TRIMESTRAL'].mean()
        rendimiento_mensual_agrupado = df_agrupar['RENDIMIENTO_MENSUAL'].mean()
        rendimiento_semanal_agrupado = df_agrupar['RENDIMIENTO_SEMANAL'].mean()
        
        # Obtener lista de equipos únicos
        todos_equipos = set()
        for equipos_str in df_agrupar['EQUIPOS'].dropna():
            equipos_lista = [eq.strip() for eq in str(equipos_str).split(',')]
            todos_equipos.update(equipos_lista)
        equipos_str = ', '.join(sorted(todos_equipos))
        
        return {
            'USUARIO': 'TOTAL',
            'EQUIPOS': equipos_str,
            'ESTADO': 'TOTAL',
            'EXPEDIENTES_DESPACHADOS': total_expedientes,
            'SEMANAS_EFECTIVAS': round(total_semanas, 1),
            'RENDIMIENTO_TOTAL': round(rendimiento_total_agrupado, 2),
            'RENDIMIENTO_ANUAL': round(rendimiento_anual_agrupado, 2),
            'POTENCIAL_ANUAL': round(potencial_anual_agrupado, 1),  # NUEVO
            'RENDIMIENTO_TRIMESTRAL': round(rendimiento_trimestral_agrupado, 2),
            'RENDIMIENTO_MENSUAL': round(rendimiento_mensual_agrupado, 2),
            'RENDIMIENTO_SEMANAL': round(rendimiento_semanal_agrupado, 2)
        }
    
    # Calcular total general
    total_general = calcular_totales_agrupados_usuarios(df_filtrado)
    
    # =============================================
    # FUNCIÓN PARA MOSTRAR CON HANDSONTABLE - ESPECÍFICA PARA RENDIMIENTO
    # =============================================
    
    def mostrar_rendimiento_con_handsontable(df_rendimiento, incluir_totales=True):
        """
        Versión funcional que muestra la tabla de rendimiento con Handsontable
        Incluye formato específico para números
        """
        import io
        import streamlit.components.v1 as components
        import json
        from datetime import datetime
        
        # 1. Crear copia y formatear datos según especificaciones
        df_display = df_rendimiento.copy()
        
        # Añadir fila de total general si corresponde
        if incluir_totales and total_general:
            df_display = pd.concat([df_display, pd.DataFrame([total_general])], ignore_index=True)
        
        # 2. Preparar datos para Handsontable con formato específico
        # Función para formatear números según especificaciones
        def formatear_numero(valor, es_decimal=True):
            try:
                if pd.isna(valor):
                    return ""
                
                # Convertir a float para formatear
                num_val = float(valor)
                
                if es_decimal:
                    # Formato para números con decimales: 1.000,00
                    return f"{num_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                else:
                    # Formato para números enteros: 1.000
                    return f"{num_val:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except:
                return str(valor) if not pd.isna(valor) else ""
        
        # Aplicar formatos específicos a cada columna
        for col in df_display.columns:
            if col in ['EXPEDIENTES_DESPACHADOS', 'POTENCIAL_ANUAL']:
                # Números enteros: 1.000
                df_display[col] = df_display[col].apply(lambda x: formatear_numero(x, es_decimal=False))
            elif col in ['SEMANAS_EFECTIVAS', 'RENDIMIENTO_TOTAL', 'RENDIMIENTO_ANUAL', 
                        'RENDIMIENTO_TRIMESTRAL', 'RENDIMIENTO_MENSUAL', 'RENDIMIENTO_SEMANAL']:
                # Números con decimales: 1.000,00 (2 decimales)
                df_display[col] = df_display[col].apply(lambda x: formatear_numero(x, es_decimal=True))
        
        # 3. Preparar datos para Handsontable
        data = df_display.to_dict('records')
        
        # 4. Preparar columnas con configuración específica
        columns = []
        for idx, col_name in enumerate(df_display.columns):
            col_config = {
                'data': col_name,
                'title': col_name,
                'type': 'text',
                'width': 100  # Ancho mínimo
            }
            
            # Configurar columnas numéricas con formato específico
            if col_name in ['EXPEDIENTES_DESPACHADOS', 'POTENCIAL_ANUAL']:
                col_config.update({
                    'type': 'numeric',
                    'numericFormat': {
                        'pattern': '0,0',  # Enteros: 1.000
                        'culture': 'es-ES'  # Usar formato español
                    }
                })
            elif col_name in ['SEMANAS_EFECTIVAS', 'RENDIMIENTO_TOTAL', 'RENDIMIENTO_ANUAL', 
                             'RENDIMIENTO_TRIMESTRAL', 'RENDIMIENTO_MENSUAL', 'RENDIMIENTO_SEMANAL']:
                col_config.update({
                    'type': 'numeric',
                    'numericFormat': {
                        'pattern': '0,0.00',  # Decimales: 1.000,00
                        'culture': 'es-ES'    # Usar formato español
                    }
                })
            
            columns.append(col_config)
        
        # 5. HTML/JavaScript para Handsontable
        hot_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <script src="https://cdn.jsdelivr.net/npm/handsontable/dist/handsontable.full.min.js"></script>
            <link href="https://cdn.jsdelivr.net/npm/handsontable/dist/handsontable.full.min.css" rel="stylesheet">
            <style>
                body {{
                    margin: 0;
                    padding: 0;
                }}
                #hot-container {{
                    width: 100%;
                    height: 600px;
                    overflow: hidden;
                }}
                .handsontable {{
                    font-size: 11px;
                }}
                .handsontable thead th {{
                    background-color: #007933 !important;
                    color: white !important;
                    font-weight: bold !important;
                }}
                /* Estilo para la fila TOTAL */
                .handsontable tbody tr:last-child td {{
                    background-color: #e6f3ff !important;
                    font-weight: bold !important;
                }}
                /* Estilo para números */
                .htRight {{
                    text-align: right !important;
                }}
            </style>
        </head>
        <body>
            <div id="hot-container"></div>
            
            <script>
                // Datos y columnas
                var data = {json.dumps(data, default=str)};
                var columns = {json.dumps(columns)};
                
                console.log("Datos de rendimiento cargados:", data.length, "registros");
                console.log("Columnas:", columns.length);
                
                // Calcular anchos basados en contenido (mínimo 100px)
                function calcularAnchosColumnas() {{
                    var anchos = [];
                    
                    // Para cada columna
                    for (var i = 0; i < columns.length; i++) {{
                        var maxAncho = 80; // Mínimo 60px
                        var colName = columns[i].title;
                        
                        // Medir encabezado
                        var anchoEncabezado = colName.length * 8; // Estimación
                        maxAncho = Math.max(maxAncho, anchoEncabezado);
                        
                        // Medir contenido de las primeras filas
                        for (var j = 0; j < Math.min(data.length, 10); j++) {{
                            var valor = data[j][colName] || '';
                            var anchoValor = valor.toString().length * 7; // Estimación
                            maxAncho = Math.max(maxAncho, anchoValor);
                        }}
                        
                        // Limitar máximo a 300px
                        anchos.push(Math.min(maxAncho, 300));
                    }}
                    return anchos;
                }}
                
                var anchosColumnas = calcularAnchosColumnas();
                console.log("Anchos calculados:", anchosColumnas);
                
                // Configurar anchos en las columnas
                for (var i = 0; i < 4; i++) {{
                    columns[i].width = anchosColumnas[i];
                }}
                for (var i = 4; i < columns.length; i++) {{
                    columns[i].width = 125;
                }}

                // Configuración
                var config = {{
                    data: data,
                    columns: columns,
                    colHeaders: true,
                    rowHeaders: true,
                    height: 600,
                    width: '100%',
                    stretchH: 'last',
                    licenseKey: 'non-commercial-and-evaluation',
                    filters: true,
                    dropdownMenu: true,
                    contextMenu: true,
                    autoWrapRow: true,
                    wordWrap: true,
                    columnSorting: true,
                    manualColumnResize: true,
                    manualRowResize: true,
                    // Configuración de formato numérico
                    numericFormat: {{
                        pattern: '0,0.00',
                        culture: 'es-ES'
                    }}
                }};
                
                // Inicializar
                var container = document.getElementById('hot-container');
                var hot = new Handsontable(container, config);
                
                // Aplicar alineación derecha a partir de la 4ta columna (índice 3)
                for (var col = 3; col < hot.countCols(); col++) {{
                    // Verificar si es columna numérica
                    var columnConfig = hot.getSettings().columns[col];
                    if (columnConfig && columnConfig.type === 'numeric') {{
                        // Aplicar a TODAS las filas de esta columna
                        for (var row = 0; row < hot.countRows(); row++) {{
                            hot.setCellMeta(row, col, 'className', 'htRight');
                        }}
                    }}
                }}
                
                // Renderizar todas las filas
                setTimeout(function() {{
                    hot.render();
                    console.log("Handsontable de rendimiento renderizado");
                    
                    // Aplicar estilo a la última fila (TOTAL)
                    var totalFilas = hot.countRows();
                    if (totalFilas > 0) {{
                        for (var col = 0; col < hot.countCols(); col++) {{
                            hot.getCellMeta(totalFilas - 1, col).className = 'htRight htBold htHighlight';
                        }}
                        hot.render();
                    }}
                }}, 100);
            </script>
        </body>
        </html>
        """
        
        # 6. Mostrar Handsontable
        st.subheader("📊 Vista de Rendimiento por Usuario")
        st.write(f"**Mostrando {len(df_display)} registros - {len(df_display.columns)} columnas**")
        
        # Mostrar Handsontable
        components.html(hot_html, height=650, scrolling=False)
        
        # 7. Exportación a Excel y CSV
        st.markdown("---")
        st.subheader("💾 Exportar Datos de Rendimiento")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            nombre = st.text_input(
                "Nombre del archivo:",
                value=f"RENDIMIENTO_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                key="excel_filename_rendimiento"
            )
        
        with col2:
            formato = st.selectbox(
                "Formato de exportación:",
                options=["Excel (.xlsx)", "CSV (.csv)"],
                key="formato_export_rendimiento"
            )
        
        col_export1, col_export2 = st.columns(2)
        
        with col_export1:
            if st.button("📊 Exportar a Excel", type="primary", use_container_width=True):
                with st.spinner("Generando Excel..."):
                    try:
                        # Preparar DataFrame para exportación
                        df_export = df_rendimiento.copy()
                        
                        # Añadir fila de total general si corresponde
                        if incluir_totales and total_general:
                            df_export = pd.concat([df_export, pd.DataFrame([total_general])], ignore_index=True)
                        
                        # Crear Excel con formato
                        output = io.BytesIO()
                        
                        # Usar openpyxl directamente para mejor control
                        from openpyxl import Workbook
                        from openpyxl.utils import get_column_letter
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                        
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Rendimiento"
                        
                        # Escribir encabezados con formato
                        header_fill = PatternFill(start_color="007933", end_color="007933", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        for col_num, header in enumerate(df_export.columns, 1):
                            cell = ws.cell(row=1, column=col_num, value=header)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        
                        # Escribir datos con formato
                        data_font = Font(size=10)
                        number_alignment = Alignment(horizontal="right", vertical="center")
                        text_alignment = Alignment(horizontal="left", vertical="center")
                        
                        for row_num, row in enumerate(df_export.values, 2):
                            for col_num, value in enumerate(row, 1):
                                cell = ws.cell(row=row_num, column=col_num, value=value)
                                cell.font = data_font
                                
                                # Aplicar formato específico según tipo de dato
                                col_name = df_export.columns[col_num-1]
                                if col_name in ['EXPEDIENTES_DESPACHADOS', 'POTENCIAL_ANUAL']:
                                    cell.alignment = number_alignment
                                    cell.number_format = '#,##0'
                                elif col_name in ['SEMANAS_EFECTIVAS']:
                                    cell.alignment = number_alignment
                                    cell.number_format = '#,##0.0'
                                elif col_name in ['RENDIMIENTO_TOTAL', 'RENDIMIENTO_ANUAL', 
                                                'RENDIMIENTO_TRIMESTRAL', 'RENDIMIENTO_MENSUAL', 'RENDIMIENTO_SEMANAL']:
                                    cell.alignment = number_alignment
                                    cell.number_format = '#,##0.00'
                                else:
                                    cell.alignment = text_alignment
                        
                        # Aplicar estilo a la fila TOTAL
                        if incluir_totales and total_general:
                            last_row = len(df_export)+1
                            total_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                            total_font = Font(bold=True)
                            
                            for col_num in range(1, len(df_export.columns) + 1):
                                cell = ws.cell(row=last_row, column=col_num)
                                cell.fill = total_fill
                                cell.font = total_font
                        
                        # Ajustar anchos de columna
                        for col_num, column_title in enumerate(df_export.columns, 1):
                            max_length = 0
                            column_letter = get_column_letter(col_num)
                            
                            # Calcular longitud máxima del contenido
                            for row_num in range(1, len(df_export) + 2):  # +2 para encabezados
                                cell_value = ws.cell(row=row_num, column=col_num).value
                                if cell_value:
                                    cell_length = len(str(cell_value))
                                    max_length = max(max_length, cell_length)
                            
                            # Establecer ancho (mínimo 10, máximo 50)
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)
                        
                        # Aplicar bordes
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for row in ws.iter_rows(min_row=1, max_row=len(df_export)+1, 
                                              min_col=1, max_col=len(df_export.columns)):
                            for cell in row:
                                cell.border = thin_border
                        
                        wb.save(output)
                        output.seek(0)
                        
                        # Botón de descarga
                        nombre_completo = f"{nombre}.xlsx"
                        st.download_button(
                            label="⬇️ Descargar Excel",
                            data=output.read(),
                            file_name=nombre_completo,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel_rendimiento"
                        )
                        
                        st.success("✅ Excel generado correctamente con formato")
                        
                    except Exception as e:
                        st.error(f"❌ Error al generar Excel: {e}")
        
        with col_export2:
            if st.button("📄 Exportar a CSV", type="secondary", use_container_width=True):
                with st.spinner("Generando CSV..."):
                    try:
                        # Preparar DataFrame para exportación CSV
                        df_export = df_rendimiento.copy()
                        
                        # Añadir fila de total general si corresponde
                        if incluir_totales and total_general:
                            df_export = pd.concat([df_export, pd.DataFrame([total_general])], ignore_index=True)
                        
                        # Convertir a CSV con separador punto y coma (;)
                        csv_data = df_export.to_csv(index=False, sep=';', decimal=',')
                        
                        # Botón de descarga
                        nombre_completo = f"{nombre}.csv"
                        st.download_button(
                            label="⬇️ Descargar CSV",
                            data=csv_data,
                            file_name=nombre_completo,
                            mime="text/csv",
                            key="download_csv_rendimiento"
                        )
                        
                        st.success("✅ CSV generado correctamente")
                        
                    except Exception as e:
                        st.error(f"❌ Error al generar CSV: {e}")
        
        return df_display
    
    # =============================================
    # MOSTRAR TABLA CON HANDSONTABLE
    # =============================================
    
    # Mostrar estadísticas rápidas
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    
    with col1:
        total_usuarios = df_filtrado['USUARIO'].nunique()
        st.metric("Usuarios únicos", total_usuarios)
    
    with col2:
        # Contar equipos únicos
        todos_equipos = set()
        for equipos_str in df_filtrado['EQUIPOS'].dropna():
            equipos_lista = [eq.strip() for eq in str(equipos_str).split(',')]
            todos_equipos.update(equipos_lista)
        st.metric("Equipos únicos", len(todos_equipos))
    
    with col3:
        total_despachados = df_filtrado['EXPEDIENTES_DESPACHADOS'].sum() + 6
        st.metric("Expedientes despachados", f"{total_despachados:,}".replace(",", "."))
    
    with col4:
        if total_general:
            st.metric("Rendimiento total", f"{total_general['RENDIMIENTO_TOTAL']:.2f}".replace(".", ","))
        else:
            st.metric("Rendimiento total", "0,00")
    
    with col5:
        if total_general:
            st.metric("Rendimiento anual", f"{total_general['RENDIMIENTO_ANUAL']:.2f}".replace(".", ","))
        else:
            st.metric("Rendimiento anual", "0,00")
    
    with col6:
        if total_general:
            st.metric("Potencial anual", f"{total_general['POTENCIAL_ANUAL']:.0f}")
        else:
            st.metric("Potencial anual", "0")
    
    with col7:
        if total_general:
            st.metric("Potencial anual conjunto", f"{total_general['POTENCIAL_ANUAL']*total_usuarios:,.0f}".replace(",", "."))
        else:
            st.metric("Potencial anual conjunto", "0")
    
    # Mostrar tabla con Handsontable
    df_mostrar = mostrar_rendimiento_con_handsontable(df_filtrado, incluir_totales=True)
    
    # =============================================
    # GRÁFICOS DE ANÁLISIS
    # =============================================
    
    st.markdown("---")
    st.subheader("📊 Gráficos de Análisis")
    
    if not df_filtrado.empty and len(df_filtrado) > 1:
        col1, col2 = st.columns(2)
        
        with col1:
            # Gráfico de rendimiento por usuario (top 10)
            df_top_rendimiento = df_filtrado.sort_values('RENDIMIENTO_TOTAL', ascending=False).head(10)
            
            fig_rendimiento = px.bar(
                df_top_rendimiento,
                x='USUARIO',
                y='RENDIMIENTO_TOTAL',
                title='Top 10 Rendimiento por Usuario',
                labels={'RENDIMIENTO_TOTAL': 'Rendimiento (expedientes/semana)', 'USUARIO': 'Usuario'},
                color='RENDIMIENTO_TOTAL',
                color_continuous_scale='Viridis',
                text='RENDIMIENTO_TOTAL'
            )
            fig_rendimiento.update_traces(texttemplate='%{text:.2f}', textposition='outside')
            fig_rendimiento.update_layout(height=400, xaxis_tickangle=-45)
            st.plotly_chart(fig_rendimiento, use_container_width=True)
        
        with col2:
            # Gráfico de distribución por estado
            conteo_estado = df_filtrado['ESTADO'].value_counts().reset_index()
            conteo_estado.columns = ['ESTADO', 'CANTIDAD']
            
            fig_estado = px.pie(
                conteo_estado,
                values='CANTIDAD',
                names='ESTADO',
                title='Distribución por Estado',
                color='ESTADO',
                color_discrete_map={'ACTIVO': '#2ca02c', 'INACTIVO': '#d62728'}
            )
            fig_estado.update_layout(height=400)
            st.plotly_chart(fig_estado, use_container_width=True)
    
    # =============================================
    # INFORMACIÓN ADICIONAL
    # =============================================
    
    with st.expander("ℹ️ Información sobre los indicadores (Modo Agrupado)"):
        st.markdown("""
        **📊 Explicación de los indicadores (MODO AGRUPADO POR USUARIO):**
        
        - **USUARIO**: Nombre del usuario (aparece una sola vez aunque esté en varios equipos)
        - **EQUIPOS**: Lista de equipos en los que trabaja el usuario (separados por comas)
        - **ESTADO**: Estado del usuario (ACTIVO/INACTIVO)
        - **EXPEDIENTES_DESPACHADOS**: Número total de expedientes despachados por el usuario en TODOS sus equipos
        - **SEMANAS EFECTIVAS**: Semanas de trabajo efectivas del usuario (calculadas UNA SOLA VEZ, no se duplican por equipo)
        - **RENDIMIENTO_TOTAL**: Expedientes despachados / Semanas efectivas de trabajo
        - **RENDIMIENTO_ANUAL**: Expedientes despachados en el último año / semanas reales del período
        - **POTENCIAL_ANUAL**: Rendimiento anual proyectado a 52 semanas (RENDIMIENTO_ANUAL × 52)
        - **RENDIMIENTO_TRIMESTRAL**: Expedientes despachados en los últimos 3 meses / semanas reales del período  
        - **RENDIMIENTO_MENSUAL**: Expedientes despachados en el último mes / semanas reales del período
        - **RENDIMIENTO_SEMANAL**: Expedientes despachados en la última semana / semanas reales del período
        
        **🔍 Criterios de estado:**
        - **ACTIVO**: FECHA_FIN vacía o posterior a la fecha máxima de análisis
        - **INACTIVO**: FECHA_FIN anterior o igual a la fecha máxima de análisis
        
        **📈 Formato de números:**
        - **Números enteros**: Formato 1.000 (sin decimales)
        - **Números con decimales**: Formato 1.000,00 (dos decimales)
        - **Exportación**: Excel y CSV mantienen estos formatos
        """)

# =============================================
# PÁGINA 5: INFORMES Y CORREOS - MODIFICADO PARA INCLUIR PDF DE RENDIMIENTO
# =============================================
elif eleccion == "Informes y Correos":
    st.header("📧 Informes y Correos")
    
    if "df_combinado" not in st.session_state:
        st.warning("⚠️ Primero carga los archivos en la sección 'Carga de Archivos'")
        st.stop()
    
    # Usar df_combinado en lugar de df
    df = st.session_state["df_combinado"]
    df_usuarios = st.session_state.get("df_usuarios", None)
    
    # Obtener información de la semana actual
    # columna_fecha = df.columns[13]
    # df[columna_fecha] = pd.to_datetime(df[columna_fecha], errors='coerce')
    # fecha_max = df[columna_fecha].max()         ya está definida anteriormente
    # dias_transcurridos = (fecha_max - FECHA_REFERENCIA).days
    # num_semana = dias_transcurridos // 7 + 1
    # fecha_max_str = fecha_max.strftime("%d/%m/%Y") if pd.notna(fecha_max) else "Sin fecha"
    
    # Descarga de informes
    st.subheader("📄 Generación de Informes PDF")
    
    df_pendientes = df[df["ESTADO"].isin(ESTADOS_PENDIENTES)].copy()
    usuarios_pendientes = df_pendientes["USUARIO"].dropna().unique()

    # NUEVO: Generar también PDFs por equipo (solo prioritarios) y resumen KPI y RENDIMIENTO
    equipos_pendientes = df_pendientes["EQUIPO"].dropna().unique()

    # --- Advertencia expedientes sin asignar (usuario "NI") ---
    exp_ni = df_pendientes[df_pendientes["USUARIO"].astype(str).str.strip().str.upper() == "NI"]
    if not exp_ni.empty:
        equipos_ni = exp_ni["EQUIPO"].dropna().unique()
        equipos_str = ", ".join(sorted(equipos_ni)) if len(equipos_ni) > 0 else "desconocido"
        st.warning(
            f"⚠️ Hay **{len(exp_ni)} expediente{'s' if len(exp_ni) > 1 else ''}** "
            f"asignado{'s' if len(exp_ni) > 1 else ''} al usuario **NI** (sin asignar). "
            f"Se generará su PDF pero no se enviará por correo. "
            f"Equipo{'s' if len(equipos_ni) > 1 else ''}: **{equipos_str}**."
        )
    
    if st.button(f"Generar {len(usuarios_pendientes)} Informes PDF + Equipos + Resumen KPI + Rendimiento", key="generar_pdfs_completos"):
        if usuarios_pendientes.size == 0:
            st.info("No se encontraron expedientes pendientes para generar informes.")
        else:
            with st.spinner('Generando PDFs y comprimiendo...'):
                zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # 1. PDFs por usuario (todos los pendientes)
                for usuario in usuarios_pendientes:
                    pdf_data = generar_pdf_usuario(usuario, df_pendientes, num_semana, fecha_max_str)
                    if pdf_data:
                        file_name = f"{num_semana}{usuario}.pdf"
                        zip_file.writestr(file_name, pdf_data)
                
                # 2. PDFs por equipo (solo expedientes prioritarios)
                for equipo in equipos_pendientes:
                    pdf_data = generar_pdf_equipo_prioritarios(equipo, df_pendientes, num_semana, fecha_max_str)
                    if pdf_data:
                        file_name = f"{num_semana}{equipo.replace('.', '_')}_PRIORITARIOS.pdf"
                        zip_file.writestr(file_name, pdf_data)
                
                # 3. PDF de resumen de KPIs
                # Calcular KPIs para la semana actual
                # columna_fecha = df.columns[13]
                # df[columna_fecha] = pd.to_datetime(df[columna_fecha], errors='coerce')
                # fecha_max = df[columna_fecha].max()
                
                # Crear rango de semanas disponibles
                fecha_inicio = pd.to_datetime("2022-11-01")
                semanas_disponibles = pd.date_range(
                    start=fecha_inicio,
                    end=fecha_max,
                    freq='W-FRI'
                ).tolist()
                
                # Calcular KPIs para todas las semanas
                df_kpis_semanales = calcular_kpis_todas_semanas_optimizado(df, semanas_disponibles, FECHA_REFERENCIA, fecha_max)
                
                # Generar PDF de resumen KPI
                pdf_resumen = generar_pdf_resumen_kpi_optimizado(
                    df_kpis_semanales, 
                    num_semana, 
                    fecha_max_str, 
                    df, 
                    semanas_disponibles, 
                    FECHA_REFERENCIA, 
                    fecha_max
                )
                if pdf_resumen:
                    file_name = f"{num_semana}RESUMEN_KPI.pdf"
                    zip_file.writestr(file_name, pdf_resumen)
                
                # 4. NUEVO: PDF de rendimiento por usuario (SOLO ACTIVOS)
                if df_usuarios is not None and not df_usuarios.empty:
                    # Calcular datos de rendimiento
                    with st.spinner("Calculando datos de rendimiento para PDF..."):
                        df_rendimiento_completo = calcular_rendimiento_usuarios_agrupado(df, df_usuarios, fecha_max)
                    
                    if not df_rendimiento_completo.empty:
                        # FILTRAR SOLO USUARIOS ACTIVOS
                        df_rendimiento_activos = df_rendimiento_completo[df_rendimiento_completo['ESTADO'] == 'ACTIVO']
                        
                        if not df_rendimiento_activos.empty:
                            pdf_rendimiento = generar_pdf_rendimiento(df_rendimiento_activos, num_semana, fecha_max_str)
                            if pdf_rendimiento:
                                file_name = f"{num_semana}RENDIMIENTO_USUARIOS_ACTIVOS.pdf"
                                zip_file.writestr(file_name, pdf_rendimiento)
                                st.success(f"✅ PDF de rendimiento generado ({len(df_rendimiento_activos)} usuarios activos)")
                        else:
                            st.warning("⚠️ No hay usuarios activos para generar PDF de rendimiento")

            zip_buffer.seek(0)
            zip_file_name = f"Informes_Completos_Semana_{num_semana}.zip"
            st.download_button(
                label=f"⬇️ Descargar {len(usuarios_pendientes)} Informes PDF + Equipos + Resumen KPI + Rendimiento (ZIP)",
                data=zip_buffer.read(),
                file_name=zip_file_name,
                mime="application/zip",
                help="Descarga todos los informes PDF listos.",
                key='pdf_download_button_completo'
            )

    # SECCIÓN: ENVÍO DE CORREOS INTEGRADA - VERSIÓN CORREGIDA Y MEJORADA
    st.markdown("---")
    st.subheader("📧 Envío de Correos Electrónicos")
    
    # CORRECCIÓN: Filtrar usuarios activos de forma más robusta
    try:
        if df_usuarios is None or df_usuarios.empty:
            st.error("❌ No se ha cargado el archivo USUARIOS o está vacío")
            st.stop()

        df_usuarios = df_usuarios.copy()

        # Normalizar campos de control
        for col in ['ENVIAR', 'ENVÍO RESUMEN']:
            if col not in df_usuarios.columns:
                df_usuarios[col] = ""
            df_usuarios[col] = df_usuarios[col].astype(str).str.upper().str.strip()

        valores_si = ['SÍ', 'SI', 'S', 'YES', 'Y', 'TRUE', '1', 'VERDADERO']

        # Usuarios que reciben su listado individual
        usuarios_enviar = df_usuarios[df_usuarios['ENVIAR'].isin(valores_si)]

        # Usuarios que reciben resumen KPI
        usuarios_resumen = df_usuarios[df_usuarios['ENVÍO RESUMEN'].isin(valores_si)]

        st.success(f"✅ {len(usuarios_enviar)} usuarios recibirán expedientes individuales")
        st.success(f"📊 {len(usuarios_resumen)} usuarios recibirán resumen KPI")

        if usuarios_enviar.empty and usuarios_resumen.empty:
            st.warning("⚠️ No hay usuarios con ENVIAR o ENVÍO RESUMEN = 'SÍ'")
            st.stop()

    except Exception as e:
        st.error(f"❌ Error procesando usuarios: {e}")
        st.stop()

    # Función para verificar envío de resumen - MÁS ROBUSTA
    def verificar_envio_resumen(usuario_row):
        """Verifica si el usuario debe recibir resumen KPI"""
        try:
            if 'ENVÍO RESUMEN' not in usuario_row.index:
                return False
            
            valor = str(usuario_row['ENVÍO RESUMEN']).upper().strip() if pd.notna(usuario_row['ENVÍO RESUMEN']) else ""
            valores_positivos = ['SÍ', 'SI', 'S', 'YES', 'Y', 'TRUE', '1', 'VERDADERO', 'OK', 'X', '✔', '✅']
            return valor in valores_positivos
        except:
            return False

    # Procesar usuarios para envío - SEPARAR CLARAMENTE LOS DOS GRUPOS
    usuarios_para_envio_individual = []  # Usuarios con expedientes que reciben su PDF
    usuarios_para_resumen_solo = []       # Usuarios que solo reciben resumen (jefes sin expedientes)

    # INICIALIZAR Y DEFINIR usuarios_con_pendientes DE FORMA SEGURA
    usuarios_con_pendientes = []
    try:
        # Primero, identificar todos los usuarios con expedientes pendientes
        if not df_pendientes.empty and 'USUARIO' in df_pendientes.columns:
            usuarios_con_pendientes = df_pendientes['USUARIO'].dropna().unique().tolist()
            st.info(f"📋 {len(usuarios_con_pendientes)} usuarios tienen expedientes pendientes")
        else:
            st.warning("ℹ️ No se encontraron expedientes pendientes o falta la columna 'USUARIO'")
            usuarios_con_pendientes = []
    except Exception as e:
        st.error(f"❌ Error al obtener usuarios con expedientes pendientes: {e}")
        usuarios_con_pendientes = []

    # VERIFICAR COLUMNAS REQUERIDAS EN EL ARCHIVO DE USUARIOS
    columnas_requeridas = ['USUARIOS', 'EMAIL']
    columnas_faltantes = [col for col in columnas_requeridas if col not in df_usuarios.columns]

    if columnas_faltantes:
        st.error(f"❌ Faltan columnas requeridas en el archivo USUARIOS: {', '.join(columnas_faltantes)}")
        st.stop()

    # PROCESAR CADA USUARIO ACTIVO - CORRECCIÓN IMPORTANTE
    for _, usuario_row in usuarios_enviar.iterrows():
        try:
            usuario_nombre = usuario_row['USUARIOS']  # Del archivo USUARIOS
            usuario_email = usuario_row['EMAIL']
            
            # Verificar datos básicos
            if pd.isna(usuario_nombre) or pd.isna(usuario_email):
                st.warning(f"⚠️ Usuario con nombre o email vacío: {usuario_nombre}")
                continue
                
            # Normalizar para comparación
            usuario_normalizado = str(usuario_nombre).strip().upper()
            
            # Verificar si tiene expedientes - COMPARAR CON 'USUARIO' de df_pendientes
            tiene_expedientes = False
            num_expedientes = 0
            
            # CORREGIDO: Verificar de forma segura si usuarios_con_pendientes está definido y no está vacío
            if usuarios_con_pendientes and len(usuarios_con_pendientes) > 0:
                try:
                    # CORREGIDO: Comparar con 'USUARIO' de los expedientes pendientes
                    tiene_expedientes = any(
                        str(user_pendiente).strip().upper() == usuario_normalizado 
                        for user_pendiente in usuarios_con_pendientes
                    )
                    
                    if tiene_expedientes:
                        # Contar expedientes del usuario - USANDO 'USUARIO' de df_pendientes
                        num_expedientes = len(df_pendientes[
                            df_pendientes['USUARIO'].apply(
                                lambda x: str(x).strip().upper() if pd.notna(x) else ''
                            ) == usuario_normalizado
                        ])
                except Exception as e:
                    st.error(f"❌ Error al verificar expedientes para {usuario_nombre}: {e}")
                    tiene_expedientes = False
                    num_expedientes = 0
            
            # Verificar si debe recibir resumen - CORRECCIÓN IMPORTANTE
            # Usamos la función verificar_envio_resumen para determinar esto
            recibir_resumen = verificar_envio_resumen(usuario_row)
            
            # Procesar asunto y mensaje
            asunto_template = usuario_row['ASUNTO'] if pd.notna(usuario_row.get('ASUNTO', '')) else f"Situación RECTAUTO asignados en la semana {num_semana} a {fecha_max_str}"
            asunto_procesado = asunto_template.replace("&num_semana&", str(num_semana)).replace("&fecha_max&", fecha_max_str)
            
            # Generar cuerpo del mensaje
            mensaje_base = ""
            if pd.notna(usuario_row.get('MENSAJE1', '')):
                mensaje_base += f"{usuario_row['MENSAJE1']}\n\n"
            if pd.notna(usuario_row.get('MENSAJE2', '')):
                mensaje_base += f"{usuario_row['MENSAJE2']}\n\n"
            if pd.notna(usuario_row.get('MENSAJE3', '')):
                mensaje_base += f"{usuario_row['MENSAJE3']}\n\n"
            if pd.notna(usuario_row.get('DESPEDIDA', '')):
                mensaje_base += f"{usuario_row['DESPEDIDA']}\n\n"
            
            if not mensaje_base.strip():
                mensaje_base = "Se adjunta informe de expedientes pendientes."
            
            mensaje_base += "__________________\n\nEquipo RECTAUTO."
            cuerpo_mensaje = f"{obtener_saludo()},\n\n{mensaje_base}"
            
            # CORRECCIÓN CRÍTICA: DETERMINAR A QUÉ LISTA PERTENECE
            # Primero verificar si el usuario está marcado para recibir resumen
            debe_recibir_resumen = verificar_envio_resumen(usuario_row)
            
            # Ahora determinar a qué lista va
            if tiene_expedientes:
                # Usuario con expedientes: va a la lista de envío individual
                usuarios_para_envio_individual.append({
                    'usuario': usuario_nombre,
                    'email': usuario_email,
                    'cc': usuario_row.get('CC', ''),
                    'bcc': usuario_row.get('BCC', ''),
                    'expedientes': num_expedientes,
                    'asunto': asunto_procesado,
                    'cuerpo_mensaje': cuerpo_mensaje,
                    'recibir_resumen': debe_recibir_resumen  # Puede que también reciba resumen
                })
                st.success(f"✅ {usuario_nombre} - Tiene {num_expedientes} expedientes" + 
                        (" y recibirá resumen" if debe_recibir_resumen else ""))
            
            # CORRECCIÓN: Usuarios que NO tienen expedientes pero SÍ deben recibir resumen
            elif debe_recibir_resumen:
                # Usuario sin expedientes pero que debe recibir resumen
                usuarios_para_resumen_solo.append({
                    'usuario': usuario_nombre,
                    'email': usuario_email,
                    'cc': usuario_row.get('CC', ''),
                    'bcc': usuario_row.get('BCC', ''),
                    'asunto': f"Resumen KPI y Rendimiento Semana {num_semana} - {fecha_max_str}",
                    'cuerpo_mensaje': f"{obtener_saludo()},\n\nSe adjunta el resumen de KPIs de la semana {num_semana}, el informe de rendimiento por usuario y los listados de expedientes prioritarios de todos los equipos.\n\n__________________\n\nEquipo RECTAUTO.",
                    'recibir_resumen': True
                })
                st.info(f"📊 {usuario_nombre} - Recibirá resumen KPI + informe rendimiento + expedientes prioritarios de todos los equipos")
            
            else:
                # Usuario sin expedientes y sin marcar para resumen
                st.warning(f"⚠️ {usuario_nombre} - No tiene expedientes ni está marcado para resumen")
                
        except Exception as e:
            st.error(f"❌ Error procesando usuario {usuario_row.get('USUARIOS', 'Desconocido')}: {e}")
            continue

    # Procesar también usuarios que solo reciben resumen (no están en ENVIAR pero sí en ENVÍO RESUMEN)
    for _, usuario_row in usuarios_resumen.iterrows():
        try:
            usuario_nombre = usuario_row['USUARIOS']
            usuario_email = usuario_row['EMAIL']
            
            # Verificar que no esté ya en la lista de envío individual
            if any(u['usuario'] == usuario_nombre for u in usuarios_para_envio_individual):
                continue  # Ya está procesado
                
            if any(u['usuario'] == usuario_nombre for u in usuarios_para_resumen_solo):
                continue  # Ya está procesado
                
            # Verificar datos básicos
            if pd.isna(usuario_nombre) or pd.isna(usuario_email):
                st.warning(f"⚠️ Usuario con nombre o email vacío: {usuario_nombre}")
                continue
            
            # Solo agregar si no tiene expedientes (para evitar duplicados)
            usuario_normalizado = str(usuario_nombre).strip().upper()
            tiene_expedientes = any(
                str(user_pendiente).strip().upper() == usuario_normalizado 
                for user_pendiente in usuarios_con_pendientes
            ) if usuarios_con_pendientes else False
            
            if not tiene_expedientes:
                usuarios_para_resumen_solo.append({
                    'usuario': usuario_nombre,
                    'email': usuario_email,
                    'cc': usuario_row.get('CC', ''),
                    'bcc': usuario_row.get('BCC', ''),
                    'asunto': f"Resumen KPI y Rendimiento Semana {num_semana} - {fecha_max_str}",
                    'cuerpo_mensaje': f"{obtener_saludo()},\n\nSe adjunta el resumen de KPIs de la semana {num_semana}, el informe de rendimiento por usuario y los listados de expedientes prioritarios de todos los equipos.\n\n__________________\n\nEquipo RECTAUTO.",
                    'recibir_resumen': True
                })
                st.info(f"📊 {usuario_nombre} - Recibirá solo resumen KPI + informe rendimiento (sin expedientes propios)")
                
        except Exception as e:
            st.error(f"❌ Error procesando usuario resumen {usuario_row.get('USUARIOS', 'Desconocido')}: {e}")
            continue

    # MOSTRAR RESUMEN DE LO QUE SE VA A ENVIAR
    st.markdown("---")
    st.subheader("📋 Resumen de Envíos Programados")

    if usuarios_para_envio_individual:
        st.success(f"✅ {len(usuarios_para_envio_individual)} usuarios recibirán sus expedientes pendientes")
        with st.expander("📋 Ver usuarios con expedientes"):
            df_envio = pd.DataFrame(usuarios_para_envio_individual)
            st.dataframe(df_envio[['usuario', 'email', 'expedientes', 'recibir_resumen']], use_container_width=True)
            # BOTÓN DE DESCARGA SIMPLE
            if not df_envio.empty:
                # Convertir a Excel en memoria
                from io import BytesIO
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_envio[['usuario', 'email', 'expedientes', 'recibir_resumen']].to_excel(writer, index=False, sheet_name='Usuarios')
                
                # Botón de descarga
                st.download_button(
                    label="📥 Descargar Excel",
                    data=output.getvalue(),
                    file_name="usuarios_expedientes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            

    if usuarios_para_resumen_solo:
        st.success(f"📊 {len(usuarios_para_resumen_solo)} usuarios recibirán el resumen KPI + informe rendimiento")
        with st.expander("📋 Ver usuarios para resumen"):
            df_resumen = pd.DataFrame(usuarios_para_resumen_solo)
            st.dataframe(df_resumen[['usuario', 'email']], use_container_width=True)

    if not usuarios_para_envio_individual and not usuarios_para_resumen_solo:
        st.warning("⚠️ No hay usuarios para enviar correos")
        st.stop()

    # BOTÓN DE ENVÍO - SIEMPRE VISIBLE SI HAY USUARIOS EN ALGUNA LISTA
    st.markdown("---")
    st.subheader("🚀 Envío de Correos")

    st.warning("""
    **⚠️ Importante antes de enviar:**
    - Se usará la cuenta de Outlook predeterminada
    - No es necesario tener Outlook abierto
    - Los correos se enviarán inmediatamente
    - **Usuarios con expedientes pendientes:** Recibirán su PDF individual
    - **Usuarios Gerente y Jefes de Equipo:** Recibirán el resumen KPI, informe de rendimiento y los Expedientes Prioritarios de todos los equipos
    - **Verifica que los datos sean correctos**
    """)

    if st.button("📤 Enviar Todos los Correos", type="primary", key="enviar_correos"):
        progress_bar = st.progress(0)
        status_text = st.empty()
        results_container = st.container()
        
        correos_enviados = 0
        correos_fallidos = 0
        
        # Generar PDF de resumen KPI (una sola vez para todos)
        pdf_resumen = None
        pdf_rendimiento = None  # NUEVO: PDF de rendimiento
        
        with st.spinner("Generando resumen KPI..."):
            # Calcular KPIs para la semana actual
            columna_fecha = df.columns[13]
            df[columna_fecha] = pd.to_datetime(df[columna_fecha], errors='coerce')
            fecha_max = df[columna_fecha].max()
            
            fecha_inicio = pd.to_datetime("2022-11-01")
            semanas_disponibles = pd.date_range(
                start=fecha_inicio,
                end=fecha_max,
                freq='W-FRI'
            ).tolist()
            
            df_kpis_semanales = calcular_kpis_todas_semanas_optimizado(df, semanas_disponibles, FECHA_REFERENCIA, fecha_max)
            pdf_resumen = generar_pdf_resumen_kpi_optimizado(
                df_kpis_semanales, 
                num_semana, 
                fecha_max_str, 
                df, 
                semanas_disponibles, 
                FECHA_REFERENCIA, 
                fecha_max
            )
        
        # NUEVO: Generar PDF de rendimiento (solo una vez, SOLO ACTIVOS)
        with st.spinner("Generando informe de rendimiento..."):
            if df_usuarios is not None and not df_usuarios.empty:
                df_rendimiento_completo = calcular_rendimiento_usuarios_agrupado(df, df_usuarios, fecha_max)
                if not df_rendimiento_completo.empty:
                    # FILTRAR SOLO USUARIOS ACTIVOS
                    df_rendimiento_activos = df_rendimiento_completo[df_rendimiento_completo['ESTADO'] == 'ACTIVO']
                    
                    if not df_rendimiento_activos.empty:
                        pdf_rendimiento = generar_pdf_rendimiento(df_rendimiento_activos, num_semana, fecha_max_str)
                        st.success(f"📊 PDF de rendimiento generado con {len(df_rendimiento_activos)} usuarios activos")
                    else:
                        st.warning("⚠️ No hay usuarios activos para generar PDF de rendimiento")
                        pdf_rendimiento = None
        
        total_a_procesar = len(usuarios_para_envio_individual) + len(usuarios_para_resumen_solo)
        
        # PRIMERO: Enviar a usuarios con expedientes
        for i, usuario_info in enumerate(usuarios_para_envio_individual):
            status_text.text(f"📨 Enviando a: {usuario_info['usuario']}")
            
            # Generar PDF individual
            pdf_individual = generar_pdf_usuario(usuario_info['usuario'], df_pendientes, num_semana, fecha_max_str)
            
            if pdf_individual:
                archivos_adjuntos = []
                
                # 1. PDF individual
                nombre_individual = f"Expedientes_Pendientes_{usuario_info['usuario']}_Semana_{num_semana}.pdf"
                archivos_adjuntos.append((nombre_individual, pdf_individual))
                
                # 2. Resumen KPI si corresponde
                if usuario_info['recibir_resumen'] and pdf_resumen:
                    nombre_resumen = f"Resumen_KPI_Semana_{num_semana}.pdf"
                    archivos_adjuntos.append((nombre_resumen, pdf_resumen))
                    
                    # 3. NUEVO: PDF de rendimiento si corresponde
                    if pdf_rendimiento:
                        nombre_rendimiento = f"Rendimiento_Usuarios_Activos_Semana_{num_semana}.pdf"
                        archivos_adjuntos.append((nombre_rendimiento, pdf_rendimiento))
                
                # Enviar correo
                exito = enviar_correo_outlook(
                    destinatario=usuario_info['email'],
                    asunto=usuario_info['asunto'],
                    cuerpo_mensaje=usuario_info['cuerpo_mensaje'],
                    archivos_adjuntos=archivos_adjuntos,
                    cc=usuario_info.get('cc'),
                    bcc=usuario_info.get('bcc')
                )
                
                if exito:
                    correos_enviados += 1
                    with results_container:
                        adjuntos = []
                        if usuario_info['recibir_resumen'] and pdf_resumen:
                            adjuntos.append("Resumen")
                        if usuario_info['recibir_resumen'] and pdf_rendimiento:
                            adjuntos.append("Rendimiento")
                        
                        mensaje = f"✅ {usuario_info['usuario']} - Expedientes"
                        if adjuntos:
                            mensaje += " + " + " + ".join(adjuntos)
                        st.success(mensaje)
                else:
                    correos_fallidos += 1
                    with results_container:
                        st.error(f"❌ Falló: {usuario_info['usuario']}")
            else:
                correos_fallidos += 1
                with results_container:
                    st.error(f"❌ No se pudo generar PDF para {usuario_info['usuario']}")
            
            progress_bar.progress((i + 1) / total_a_procesar)
        
        # SEGUNDO: Enviar solo resúmenes a usuarios sin expedientes - CON TODOS LOS ADJUNTOS
        for i, usuario_info in enumerate(usuarios_para_resumen_solo):
            status_text.text(f"📊 Enviando resumen a: {usuario_info['usuario']}")
            
            # Crear lista de adjuntos FRESCA para cada usuario
            archivos_adjuntos = []
            
            if pdf_resumen:
                # 1. Adjuntar resumen KPI
                nombre_resumen = f"Resumen_KPI_Semana_{num_semana}.pdf"
                archivos_adjuntos.append((nombre_resumen, pdf_resumen))
                
                # 2. NUEVO: Adjuntar PDF de rendimiento
                if pdf_rendimiento:
                    nombre_rendimiento = f"Rendimiento_Usuarios_Activos_Semana_{num_semana}.pdf"
                    archivos_adjuntos.append((nombre_rendimiento, pdf_rendimiento))
                
                # 3. Adjuntar expedientes prioritarios de TODOS los equipos
                # Obtener la lista de equipos únicos con expedientes pendientes
                equipos = df_pendientes['EQUIPO'].dropna().unique()
                
                for equipo in equipos:
                    pdf_prioritarios_equipo = generar_pdf_equipo_prioritarios(
                        equipo, 
                        df_pendientes, 
                        num_semana, 
                        fecha_max_str
                    )
                    
                    if pdf_prioritarios_equipo:
                        nombre_prioritarios = f"Expedientes_Prioritarios_{equipo.replace('.', '_')}_Semana_{num_semana}.pdf"
                        archivos_adjuntos.append((nombre_prioritarios, pdf_prioritarios_equipo))
                
                # Actualizar el cuerpo del mensaje para reflejar los nuevos adjuntos
                cuerpo_mensaje_actualizado = f"{obtener_saludo()},\n\nSe adjunta el resumen de KPIs de la semana {num_semana}, el informe de rendimiento por usuario y los listados de expedientes prioritarios de todos los equipos.\n\n__________________\n\nEquipo RECTAUTO."
                
                # Enviar correo con todos los adjuntos
                exito = enviar_correo_outlook(
                    destinatario=usuario_info['email'],
                    asunto=usuario_info['asunto'],
                    cuerpo_mensaje=cuerpo_mensaje_actualizado,
                    archivos_adjuntos=archivos_adjuntos,
                    cc=usuario_info.get('cc'),
                    bcc=usuario_info.get('bcc')
                )
                
                if exito:
                    correos_enviados += 1
                    with results_container:
                        st.success(f"📊 Resumen KPI + Rendimiento + {len(equipos)} equipos de expedientes prioritarios enviados a {usuario_info['usuario']}")
                else:
                    correos_fallidos += 1
                    with results_container:
                        st.error(f"❌ Falló resumen: {usuario_info['usuario']}")
            else:
                correos_fallidos += 1
                with results_container:
                    st.error(f"❌ No hay resumen para {usuario_info['usuario']}")
            
            progress_bar.progress((len(usuarios_para_envio_individual) + i + 1) / total_a_procesar)
        
        status_text.text("")
        
        # RESUMEN FINAL
        st.markdown("---")
        st.subheader("📊 Resumen del Envío")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total procesados", total_a_procesar)
        with col2:
            st.metric("Correos enviados", correos_enviados)
        with col3:
            st.metric("Correos fallidos", correos_fallidos)
        
        # Calcular resúmenes enviados
        resumenes_enviados = sum(1 for u in usuarios_para_envio_individual if u.get('recibir_resumen', False))
        resumenes_enviados += len(usuarios_para_resumen_solo)
        
        st.info(f"📊 Resúmenes KPI enviados: {resumenes_enviados}")
        if pdf_rendimiento:
            st.info(f"📈 Informes de rendimiento enviados: {resumenes_enviados}")
        
        if correos_enviados > 0:
            st.balloons()
            st.success("🎉 ¡Envío de correos completado!")